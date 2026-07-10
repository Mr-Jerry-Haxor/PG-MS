from decimal import Decimal, InvalidOperation
from datetime import date
import json

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail
from django.db import IntegrityError, transaction
from django.db.models import Count, Exists, Min, OuterRef, Prefetch, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_date
from urllib.parse import urlencode

try:
    from allauth.account.models import EmailAddress
except Exception:  # allauth not strictly required at import time
    EmailAddress = None

from bookings.models import Booking, ResidentApplication, Room, RoomShareStatus, ReferralCredit, RoomSwap
from core.audit import log
from core.drive import drive_delete
from core.models import Notification
from core.push_notifications import send_push_to_user, send_push_to_users
from finance.models import Fees
from django.urls import reverse
from django.http import HttpResponse
import calendar
from io import BytesIO
from bookings.utils import pending_booking_share_keys

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
except Exception:
    openpyxl = None
from .forms import PGForm, RoomForm, ShareStatusForm
from .models import PG, PGAdmin

import logging
_logger = logging.getLogger(__name__)


def _get_swap_chain_order(pending_swaps):
    """
    Order pending swaps to handle chain dependencies correctly.
    Example: A leaves, B->A's bed, C->B's bed, D->C's bed
    Must execute in order that respects dependencies (who's moving where).
    
    Returns: list of swaps in correct execution order
    """
    if not pending_swaps:
        return []
    
    swaps_list = list(pending_swaps)
    
    # Build dependency graph:
    # A swap S depends on another swap T if S.from_room/from_share == T.to_room/to_share
    # (S needs to move FROM a bed that T is moving TO)
    
    # Map: (room_id, share_no) -> swap that targets this bed
    target_map = {}
    for swap in swaps_list:
        key = (swap.to_room_id, swap.to_share_no)
        target_map[key] = swap
    
    # Find dependencies
    # swap_id -> set of swap_ids it depends on
    dependencies = {swap.id: set() for swap in swaps_list}
    
    for swap in swaps_list:
        # This swap moves FROM (from_room, from_share)
        # If another swap is moving TO (from_room, from_share), that swap must execute first
        source_key = (swap.from_room_id, swap.from_share_no)
        if source_key in target_map:
            dep_swap = target_map[source_key]
            if dep_swap.id != swap.id:
                # This swap depends on dep_swap executing first
                # Actually, it's the opposite: if someone is moving INTO my source bed,
                # they depend on ME moving out first
                dependencies[dep_swap.id].add(swap.id)
    
    # Topological sort
    result = []
    visited = set()
    temp_mark = set()
    
    def visit(swap):
        if swap.id in temp_mark:
            # Cycle detected - shouldn't happen in valid swaps
            return
        if swap.id in visited:
            return
        temp_mark.add(swap.id)
        # Visit all swaps that must happen before this one
        for dep_id in dependencies[swap.id]:
            dep_swap = next((s for s in swaps_list if s.id == dep_id), None)
            if dep_swap:
                visit(dep_swap)
        temp_mark.remove(swap.id)
        visited.add(swap.id)
        result.append(swap)
    
    for swap in swaps_list:
        if swap.id not in visited:
            visit(swap)
    
    return result


def _calculate_bed_status_after_swaps(pg):
    """
    Calculate what the final bed status should be after considering all pending future swaps.
    This helps determine which beds will actually be vacant vs occupied after chains execute.
    
    Returns: dict of {(room_id, share_no): {'status': str, 'vacant_from': date or None}}
    """
    today = timezone.now().date()
    
    # Get all pending/approved future swaps for this PG
    pending_swaps = RoomSwap.objects.filter(
        status__in=[RoomSwap.PENDING, RoomSwap.APPROVED],
        is_future_swap=True,
        booking__room__pg=pg
    ).select_related('booking', 'from_room', 'to_room', 'booking__user')
    
    # Build a map of bed movements
    # Key: (room_id, share_no)
    # Value: {'incoming_swap': swap or None, 'outgoing_swap': swap or None}
    bed_movements = {}
    
    for swap in pending_swaps:
        from_key = (swap.from_room_id, swap.from_share_no)
        to_key = (swap.to_room_id, swap.to_share_no)
        
        if from_key not in bed_movements:
            bed_movements[from_key] = {'incoming_swap': None, 'outgoing_swap': None, 'effective_date': None}
        if to_key not in bed_movements:
            bed_movements[to_key] = {'incoming_swap': None, 'outgoing_swap': None, 'effective_date': None}
        
        bed_movements[from_key]['outgoing_swap'] = swap
        bed_movements[from_key]['effective_date'] = swap.effective_date
        bed_movements[to_key]['incoming_swap'] = swap
        if not bed_movements[to_key]['effective_date'] or swap.effective_date < bed_movements[to_key]['effective_date']:
            bed_movements[to_key]['effective_date'] = swap.effective_date
    
    # Now calculate final status for each affected bed
    # A bed will be:
    # - VACANT/VACANT_FROM if it has outgoing but no incoming swap (chain endpoint)
    # - OCCUPIED/RESERVED if it has incoming swap
    # - OCCUPIED if no movement (current occupant stays)
    
    bed_final_status = {}
    
    for bed_key, movements in bed_movements.items():
        room_id, share_no = bed_key
        has_incoming = movements['incoming_swap'] is not None
        has_outgoing = movements['outgoing_swap'] is not None
        effective_date = movements['effective_date']
        
        if has_outgoing and not has_incoming:
            # This bed will become vacant after the swap executes
            # Check if effective_date is in the future
            if effective_date and effective_date > today:
                bed_final_status[bed_key] = {
                    'status': RoomShareStatus.VACANT_FROM,
                    'vacant_from': effective_date
                }
            else:
                bed_final_status[bed_key] = {
                    'status': RoomShareStatus.VACANT,
                    'vacant_from': None
                }
        elif has_incoming:
            # This bed will be occupied by someone coming in
            # Status depends on when the incoming swap executes
            incoming_swap = movements['incoming_swap']
            if incoming_swap.effective_date > today:
                # Future incoming - bed is RESERVED
                bed_final_status[bed_key] = {
                    'status': RoomShareStatus.RESERVED,
                    'vacant_from': None
                }
            else:
                # Immediate - bed is OCCUPIED
                bed_final_status[bed_key] = {
                    'status': RoomShareStatus.OCCUPIED,
                    'vacant_from': None
                }
    
    return bed_final_status


def _auto_sync_on_page_load(pg, user):
    """
    Auto-sync bed statuses and execute pending future swaps for a PG.
    Called when PG admin loads key pages (tenants, dashboard, etc).
    Returns: {'synced': bool, 'swaps_executed': int, 'swaps_failed': int}
    """
    if not pg:
        return {'synced': False, 'swaps_executed': 0, 'swaps_failed': 0}
    
    today = timezone.now().date()
    results = {'synced': False, 'swaps_executed': 0, 'swaps_failed': 0}
    
    try:
        # 1. Execute pending future swaps that are due (in correct chain order)
        pending_swaps = RoomSwap.objects.filter(
            status=RoomSwap.PENDING,
            is_future_swap=True,
            effective_date__lte=today,
            booking__room__pg=pg
        ).select_related('booking', 'from_room', 'to_room', 'booking__user').order_by('effective_date', 'requested_at')
        
        # Order swaps to handle chains correctly
        ordered_swaps = _get_swap_chain_order(pending_swaps)
        
        for swap in ordered_swaps:
            try:
                result = _execute_future_swap_auto(swap, user)
                if result['success']:
                    results['swaps_executed'] += 1
                else:
                    results['swaps_failed'] += 1
            except Exception as e:
                _logger.warning(f"Auto-execute swap #{swap.id} failed: {e}")
                results['swaps_failed'] += 1
        
        # 2. Sync bed statuses
        from bookings.utils import sync_room_share_statuses
        sync_room_share_statuses(pg=pg)
        results['synced'] = True
        
    except Exception as e:
        _logger.warning(f"Auto-sync for PG {pg.id} failed: {e}")
    
    return results


def _execute_future_swap_auto(swap, executor):
    """
    Execute a pending future swap automatically (silent, no UI messages).
    Returns: {'success': bool, 'error': str or None}
    """
    try:
        with transaction.atomic():
            # Re-fetch with lock
            swap = RoomSwap.objects.select_for_update().get(pk=swap.id)
            
            if swap.status != RoomSwap.PENDING:
                return {'success': False, 'error': f'Swap status is {swap.status}'}
            
            booking = Booking.objects.select_for_update().get(pk=swap.booking_id, status=Booking.APPROVED)
            from_room = Room.objects.select_for_update().get(pk=swap.from_room_id)
            to_room = Room.objects.select_for_update().get(pk=swap.to_room_id)
            from_share = RoomShareStatus.objects.select_for_update().get(room=from_room, share_no=swap.from_share_no)
            to_share = RoomShareStatus.objects.select_for_update().get(room=to_room, share_no=swap.to_share_no)
            
            # Verify booking is still at source location
            if booking.room_id != from_room.id or booking.share_no != swap.from_share_no:
                swap.status = RoomSwap.CANCELLED
                swap.reason += f" | Auto-cancelled: booking moved"
                swap.processed_at = timezone.now()
                swap.save(update_fields=['status', 'reason', 'processed_at'])
                return {'success': False, 'error': 'Booking moved'}
            
            today = timezone.now().date()
            
            # Check if target bed is actually available (by checking real bookings, not just status flag)
            # Allow if there's a leaving booking with leaving_date <= today (they've left/leaving)
            blocking_booking = Booking.objects.filter(
                room=to_room,
                share_no=swap.to_share_no,
                status=Booking.APPROVED,
                joining_date__lte=today
            ).filter(
                Q(leaving_date__isnull=True) | Q(leaving_date__gt=today)
            ).exclude(pk=booking.pk).first()
            
            if blocking_booking:
                swap.status = RoomSwap.CANCELLED
                blocker_name = blocking_booking.user.get_full_name() or blocking_booking.user.email
                swap.reason += f" | Auto-cancelled: bed occupied by {blocker_name}"
                swap.processed_at = timezone.now()
                swap.save(update_fields=['status', 'reason', 'processed_at'])
                return {'success': False, 'error': f'Bed occupied by {blocker_name}'}
            
            # Mark any leaving booking at the target bed as COMPLETED
            # (booking with leaving_date <= today that hasn't been marked completed yet)
            leaving_bookings = Booking.objects.filter(
                room=to_room,
                share_no=swap.to_share_no,
                status=Booking.APPROVED,
                leaving_date__lte=today
            ).exclude(pk=booking.pk)
            
            for leaving_booking in leaving_bookings:
                leaving_booking.status = Booking.COMPLETED
                leaving_booking.save(update_fields=['status'])
            
            # Execute the swap
            booking.room = to_room
            booking.share_no = swap.to_share_no
            try:
                booking.pg = to_room.pg
            except Exception:
                pass
            booking.save(update_fields=['room', 'pg', 'share_no'])
            
            # Update application if exists
            app = getattr(booking, 'application', None)
            if app and app.room_id != to_room.id:
                app.room = to_room
                app.save(update_fields=['room'])
            
            # Update share statuses
            from_share.status = RoomShareStatus.VACANT
            from_share.vacant_from = None
            from_share.save(update_fields=['status', 'vacant_from'])
            
            to_share.status = RoomShareStatus.OCCUPIED
            to_share.vacant_from = None
            to_share.save(update_fields=['status', 'vacant_from'])
            
            # Mark swap completed
            swap.status = RoomSwap.COMPLETED
            swap.processed_at = timezone.now()
            swap.processed_by = executor
            swap.save(update_fields=['status', 'processed_at', 'processed_by'])
            
            # Log
            log(
                executor,
                'future_swap_auto_executed',
                'RoomSwap',
                swap.id,
                f"Auto-executed swap: {booking.user.get_full_name() or booking.user.email} from room {from_room.room_no} bed {swap.from_share_no} to room {to_room.room_no} bed {swap.to_share_no}"
            )
            
            return {'success': True, 'error': None}
            
    except Booking.DoesNotExist:
        return {'success': False, 'error': 'Booking not found'}
    except Room.DoesNotExist:
        return {'success': False, 'error': 'Room not found'}
    except RoomShareStatus.DoesNotExist:
        return {'success': False, 'error': 'Share not found'}
    except Exception as e:
        return {'success': False, 'error': str(e)}

@login_required
def booking_joining_update(request, booking_id):
    from bookings.models import Booking
    booking = get_object_or_404(Booking, pk=booking_id)
    # Authorization: must be a PG Admin and admin of the booking's PG
    u = request.user
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    redirect_url = request.META.get('HTTP_REFERER') or 'pg_resident_applications'

    if not _require_pg_admin(u) or not _admin_pgs(u).filter(id=getattr(booking, 'pg_id', None)).exists():
        message = 'PG Admin access required for this PG.'
        if is_ajax:
            return JsonResponse({'ok': False, 'error': message}, status=403)
        messages.error(request, message)
        return redirect('pg_resident_applications')

    if request.method != 'POST':
        message = 'Unsupported request method.'
        if is_ajax:
            return JsonResponse({'ok': False, 'error': message}, status=405)
        messages.error(request, message)
        return redirect('pg_resident_applications')
    date_str = (request.POST.get('joining_date') or '').strip()
    if not date_str:
        message = 'Joining date is required.'
        if is_ajax:
            return JsonResponse({'ok': False, 'error': message}, status=400)
        messages.error(request, message)
        return redirect(redirect_url)
    dt = parse_date(date_str)
    if not dt:
        message = 'Invalid date format. Use YYYY-MM-DD.'
        if is_ajax:
            return JsonResponse({'ok': False, 'error': message}, status=400)
        messages.error(request, message)
        return redirect(redirect_url)
    try:
        booking.joining_date = dt
        # If payment_date is not explicitly set or equals old joining_date, update it too
        old_payment_date = booking.payment_date
        old_joining_date = Booking.objects.filter(pk=booking_id).values_list('joining_date', flat=True).first()
        
        update_fields = ['joining_date']
        # Auto-update payment_date if it was previously synced with joining_date or not set
        if not old_payment_date or old_payment_date == old_joining_date:
            booking.payment_date = dt
            update_fields.append('payment_date')
        
        booking.save(update_fields=update_fields)
        success_message = f'Joining date updated to {dt}.'
        if is_ajax:
            return JsonResponse({
                'ok': True,
                'message': success_message,
                'value': dt.isoformat(),
                'value_iso': dt.isoformat(),
                'display': dt.isoformat(),
            })
        messages.success(request, success_message)
    except Exception as e:
        error_message = f'Could not update joining date: {e}'
        if is_ajax:
            return JsonResponse({'ok': False, 'error': error_message}, status=500)
        messages.error(request, error_message)
    return redirect(redirect_url)


@login_required
@transaction.atomic
def booking_leave_direct(request, booking_id: int) -> JsonResponse:
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Invalid request method.'}, status=405)
    if not _require_pg_admin(request.user):
        return JsonResponse({'ok': False, 'error': 'PG Admin access required.'}, status=403)
    booking = get_object_or_404(Booking.objects.select_for_update(), pk=booking_id, status=Booking.APPROVED)
    pg_id = getattr(booking, 'pg_id', None) or getattr(getattr(booking, 'room', None), 'pg_id', None)
    if not _admin_pgs(request.user).filter(id=pg_id).exists():
        return JsonResponse({'ok': False, 'error': 'PG Admin access required for this PG.'}, status=403)
    date_str = (request.POST.get('leaving_date') or '').strip()
    if not date_str:
        return JsonResponse({'ok': False, 'error': 'Leaving date is required.'}, status=400)
    leaving_date = parse_date(date_str)
    if not leaving_date:
        return JsonResponse({'ok': False, 'error': 'Invalid leaving date.'}, status=400)
    min_allowed = booking.joining_date or booking.start_date
    if min_allowed and leaving_date < min_allowed:
        return JsonResponse({'ok': False, 'error': f'Leaving date must be on or after {min_allowed}.'}, status=400)

    share = get_object_or_404(
        RoomShareStatus.objects.select_for_update(),
        room=booking.room,
        share_no=booking.share_no,
    )

    previous_status = share.status

    # If the requested leaving date equals the booking's joining date and it's in the future,
    # treat this as an immediate cancellation: delete the booking and free the share.
    today = timezone.now().date()
    if booking.joining_date and leaving_date == booking.joining_date and leaving_date > today:
        # Delete the booking and free the share
        bid = booking.id
        room_id = booking.room_id
        share_no = booking.share_no

        # Delete booking record
        booking.delete()

        # Update share to VACANT
        share.status = RoomShareStatus.VACANT
        share.vacant_from = None
        share.save(update_fields=['status', 'vacant_from'])

        room_counts = _room_share_counts(share.room)

        log(
            request.user,
            'booking_cancelled',
            'Booking',
            bid,
            f"Booking {bid} cancelled and deleted for room {share.room.room_no} bed {share_no}",
        )

        return JsonResponse({
            'ok': True,
            'action': 'booking_deleted',
            'booking_id': bid,
            'room_id': room_id,
            'share_status': share.status,
            'vacant_from': '',
            'previous_status': previous_status,
            'room_counts': room_counts,
            'message': 'Booking cancelled and deleted.',
        })

    # Otherwise, treat as a leave request (existing behavior)
    update_fields = ['leaving_date']
    booking.leaving_date = leaving_date
    # Capture optional leaving reason submitted by PG admin
    leaving_reason = (request.POST.get('leaving_reason') or '').strip()
    if leaving_reason:
        booking.leaving_reason = leaving_reason
        update_fields.append('leaving_reason')

    # Ask whether advance is eligible (checkbox on modal). Default behavior in UI is checked.
    advance_eligible = request.POST.get('advance_eligible') == 'on'
    booking.advance_eligible = bool(advance_eligible)
    update_fields.append('advance_eligible')

    # Record that a leave was initiated by PG admin now
    booking.leaving_initiated_at = timezone.localtime(timezone.now())
    update_fields.append('leaving_initiated_at')

    if booking.leaving_confirmed_date:
        booking.leaving_confirmed_date = None
        update_fields.append('leaving_confirmed_date')
    booking.save(update_fields=update_fields)

    share.status = RoomShareStatus.VACANT_FROM
    share.vacant_from = leaving_date
    share.save(update_fields=['status', 'vacant_from'])

    room_counts = _room_share_counts(booking.room)

    log(
        request.user,
        'booking_leave_requested',
        'Booking',
        booking.id,
        f"Leave requested for room {booking.room.room_no} bed {booking.share_no} on {leaving_date}",
    )

    # Notify all PG admins (including those managing multiple PGs) with a filtered deep link,
    # and notify tenant that the leave request is now pending confirmation.
    try:
        admin_path, admin_payload = _leave_admin_path_and_payload(booking)
        admin_users = _pg_admin_users_for_pg(pg_id)
        tenant_name = booking.user.get_full_name() or booking.user.email

        for admin_user in admin_users:
            Notification.objects.create(
                user=admin_user,
                title="Leave Request Received",
                message=(
                    f"{tenant_name} requested leave for Room {booking.room.room_no}, "
                    f"Bed {booking.share_no} on {leaving_date}."
                ),
            )

        send_push_to_users(
            admin_users,
            title="Leave Request Received",
            body=f"{tenant_name} requested leave for Room {booking.room.room_no}, Bed {booking.share_no}.",
            url=admin_path,
            extra_data={**admin_payload, 'type': 'leave_requested', 'source': 'pg_admin_direct_leave'},
        )

        send_push_to_user(
            booking.user,
            title="Leave Request Created",
            body=f"Leave request for Room {booking.room.room_no}, Bed {booking.share_no} is pending admin confirmation.",
            url=reverse('booking_detail', args=[booking.id]),
            extra_data={'type': 'leave_requested', 'booking_id': booking.id, 'pg_id': pg_id},
        )
    except Exception:
        _logger.exception('Leave request notification dispatch failed for booking %s', booking.id)

    return JsonResponse({
        'ok': True,
        'action': 'booking_leave_requested',
        'booking_id': booking.id,
        'room_id': booking.room_id,
        'leaving_date': leaving_date.isoformat(),
        'share_status': share.status,
        'vacant_from': share.vacant_from.isoformat() if share.vacant_from else '',
        'previous_status': previous_status,
        'room_counts': room_counts,
        'message': 'Leave request recorded and pending confirmation.',
        'leave_message': f'Leave requested for {leaving_date.isoformat()} (awaiting confirmation).',
    })


@login_required
@transaction.atomic
def booking_swap_room(request, booking_id: int) -> JsonResponse:
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Invalid request method.'}, status=405)
    if not _require_pg_admin(request.user):
        return JsonResponse({'ok': False, 'error': 'PG Admin access required.'}, status=403)
    booking = get_object_or_404(Booking.objects.select_for_update(), pk=booking_id, status=Booking.APPROVED)
    pg_id = getattr(booking, 'pg_id', None) or getattr(getattr(booking, 'room', None), 'pg_id', None)
    if not _admin_pgs(request.user).filter(id=pg_id).exists():
        return JsonResponse({'ok': False, 'error': 'PG Admin access required for this PG.'}, status=403)

    try:
        room_id = int(request.POST.get('room_id'))
        share_no = int(request.POST.get('share_no'))
    except (TypeError, ValueError):
        return JsonResponse({'ok': False, 'error': 'Room and bed selections are required.'}, status=400)

    if room_id == booking.room_id and share_no == booking.share_no:
        return JsonResponse({'ok': False, 'error': 'Select a different room or bed to swap.'}, status=400)

    new_room = get_object_or_404(Room.objects.select_for_update(), pk=room_id, pg_id=pg_id)
    new_share = get_object_or_404(RoomShareStatus.objects.select_for_update(), room=new_room, share_no=share_no)
    
    # Check if swap_with_occupied checkbox was checked
    swap_with_occupied = request.POST.get('swap_with_occupied') == 'on'
    
    # Check if this is a future swap (has swap_date in the POST)
    swap_date_str = request.POST.get('swap_date', '').strip()
    today = timezone.now().date()
    swap_date = parse_date(swap_date_str) if swap_date_str else None
    is_future_swap = swap_date and swap_date > today
    
    # For future swaps, be more lenient with status checks since the bed might
    # become available by the swap date (e.g., occupant has outgoing future swap)
    if is_future_swap:
        # For future swaps, allow any status except RESERVED (which means someone else is coming)
        # The detailed availability check happens later using the swap_date
        if new_share.status == RoomShareStatus.RESERVED:
            # Check if the bed is reserved due to an incoming swap - that's a conflict
            # But if it's reserved due to a future booking, we need more context
            pass  # Will be validated in the future swap section below
        # Don't block based on current OCCUPIED status for future swaps
        # The actual conflict checking happens in the future swap validation section
    else:
        # Immediate swap - validate the target share status strictly
        if new_share.status == RoomShareStatus.OCCUPIED and not swap_with_occupied:
            return JsonResponse({'ok': False, 'error': 'Cannot swap to occupied bed. Please check "Swap with occupied" checkbox if you want to exchange rooms with another tenant.'}, status=400)
        
        if swap_with_occupied:
            # Allow VACANT, VACANT_FROM, or OCCUPIED
            if new_share.status not in [RoomShareStatus.VACANT, RoomShareStatus.VACANT_FROM, RoomShareStatus.OCCUPIED]:
                return JsonResponse({'ok': False, 'error': 'Selected bed is not available for swap.'}, status=400)
        else:
            # Only allow VACANT or VACANT_FROM
            if new_share.status not in [RoomShareStatus.VACANT, RoomShareStatus.VACANT_FROM]:
                return JsonResponse({'ok': False, 'error': 'Selected bed is not available for swap.'}, status=400)

    old_room = booking.room
    old_share_no = booking.share_no
    old_share = get_object_or_404(RoomShareStatus.objects.select_for_update(), room=old_room, share_no=old_share_no)
    
    # Check if this is a future swap (has swap_date in the POST)
    if swap_date_str:
        if swap_date and swap_date > today:
            # This is a future swap - validate and create pending swap record
            
            # Validate that the target bed will be available on that date
            conflicts = []
            
            # Get bookings that have outgoing future swaps (they will vacate the bed)
            # These bookings should be excluded from the "occupied" check
            bookings_with_outgoing_swaps = RoomSwap.objects.filter(
                from_room_id=room_id,
                from_share_no=share_no,
                effective_date__lte=swap_date,
                status__in=[RoomSwap.PENDING, RoomSwap.APPROVED],
                is_future_swap=True
            ).values_list('booking_id', flat=True)
            
            # Check if bed has active booking on that date (excluding those with outgoing swaps)
            active_booking = Booking.objects.filter(
                room_id=room_id,
                share_no=share_no,
                status=Booking.APPROVED,
                joining_date__lte=swap_date
            ).filter(
                Q(leaving_date__isnull=True) | Q(leaving_date__gt=swap_date)
            ).exclude(
                pk__in=bookings_with_outgoing_swaps
            ).exists()
            
            if active_booking:
                conflicts.append('Bed is already occupied on the selected date')
            
            # Check for other pending future swaps targeting this bed
            pending_swap = RoomSwap.objects.filter(
                to_room_id=room_id,
                to_share_no=share_no,
                effective_date__lte=swap_date,
                status__in=[RoomSwap.PENDING, RoomSwap.APPROVED],
                is_future_swap=True
            ).exists()
            
            if pending_swap:
                conflicts.append('Another future swap is already scheduled for this bed on or before this date')
            
            # Check for future bookings (excluding those with outgoing swaps)
            future_booking = Booking.objects.filter(
                room_id=room_id,
                share_no=share_no,
                status=Booking.APPROVED,
                joining_date__lte=swap_date,
                joining_date__gt=today
            ).exclude(
                pk__in=bookings_with_outgoing_swaps
            ).exists()
            
            if future_booking:
                conflicts.append('Bed has a future booking starting on or before the selected date')
            
            if conflicts:
                return JsonResponse({'ok': False, 'error': '; '.join(conflicts)}, status=400)
            
            # Create pending future swap record
            future_swap = RoomSwap.objects.create(
                booking=booking,
                from_room=old_room,
                from_share_no=old_share_no,
                to_room=new_room,
                to_share_no=share_no,
                effective_date=swap_date,
                is_future_swap=True,
                status=RoomSwap.PENDING,
                reason=f"Scheduled room swap for {swap_date.strftime('%Y-%m-%d')}",
                processed_by=request.user
            )
            
            log(
                request.user,
                'future_swap_scheduled',
                'RoomSwap',
                future_swap.id,
                f"Future swap scheduled: {booking.user.get_full_name() or booking.user.email} from room {old_room.room_no} bed {old_share_no} to room {new_room.room_no} bed {share_no} on {swap_date.strftime('%Y-%m-%d')}"
            )
            
            return JsonResponse({
                'ok': True,
                'action': 'future_swap_scheduled',
                'booking_id': booking.id,  # Added for JS modal handling
                'swap_id': future_swap.id,
                'swap_date': swap_date.strftime('%Y-%m-%d'),
                'message': f'Room swap scheduled for {swap_date.strftime("%Y-%m-%d")}. The swap will be executed automatically when you click Refresh on or after that date.',
                'from_room': old_room.room_no,
                'from_share': old_share_no,
                'to_room': new_room.room_no,
                'to_share': share_no,
                'trigger_sync': True,  # Signal to frontend to trigger sync
            })
    
    # If we get here, this is an immediate swap (not future)
    swap_type = "regular"  # regular, exchange, or occupied

    # Check if this is a swap with an OCCUPIED share (regular tenant)
    if new_share.status == RoomShareStatus.OCCUPIED:
        # Find the booking occupying this share
        occupying_booking = Booking.objects.filter(
            room=new_room,
            share_no=share_no,
            status=Booking.APPROVED,
            leaving_date__isnull=True
        ).select_for_update().first()
        
        if occupying_booking:
            swap_type = "occupied"
            # SWAP: Exchange the two tenants' rooms
            # Create swap log for the occupying tenant first
            RoomSwap.objects.create(
                booking=occupying_booking,
                from_room=new_room,
                from_share_no=share_no,
                to_room=old_room,
                to_share_no=old_share_no,
                effective_date=today,
                is_future_swap=False,
                status=RoomSwap.COMPLETED,
                reason=f"Exchanged with {booking.user.get_full_name() or booking.user.email} during room swap",
                processed_at=timezone.now(),
                processed_by=request.user
            )
            
            # Move occupying tenant to current tenant's old room
            occupying_booking.room = old_room
            occupying_booking.share_no = old_share_no
            # Ensure booking.pg stays in sync with room.pg
            try:
                occupying_booking.pg = old_room.pg
            except Exception:
                pass
            occupying_booking.save(update_fields=['room', 'pg', 'share_no'])
            
            # Update occupying tenant's application if exists
            occupying_app = getattr(occupying_booking, 'application', None)
            if occupying_app and occupying_app.room_id != old_room.id:
                occupying_app.room = old_room
                occupying_app.save(update_fields=['room'])
            
            # Both shares remain OCCUPIED, just swapped tenants
            # Old share status doesn't change (still OCCUPIED)
            # New share status doesn't change (still OCCUPIED)
            
            log(request.user, 'booking_swap', 'Booking', booking.id, 
                f"Swapped with occupied tenant: moved to room {new_room.room_no} bed {share_no}, " +
                f"occupied tenant moved to room {old_room.room_no} bed {old_share_no}")
        else:
            return JsonResponse({'ok': False, 'error': 'No active booking found for the selected occupied bed.'}, status=400)
    
    # Check if this is a swap with a VACANT_FROM share (occupied by a leaving tenant)
    elif new_share.status == RoomShareStatus.VACANT_FROM:
        # Find the booking that will be leaving from this share
        leaving_booking = Booking.objects.filter(
            room=new_room,
            share_no=share_no,
            status=Booking.APPROVED,
            leaving_date__isnull=False,
            leaving_confirmed_date__isnull=False
        ).select_for_update().first()
        
        if leaving_booking:
            swap_type = "exchange"
            # SWAP: Exchange the two tenants' rooms
            # Create swap log for the leaving tenant first
            RoomSwap.objects.create(
                booking=leaving_booking,
                from_room=new_room,
                from_share_no=share_no,
                to_room=old_room,
                to_share_no=old_share_no,
                effective_date=today,
                is_future_swap=False,
                status=RoomSwap.COMPLETED,
                reason=f"Exchanged with {booking.user.get_full_name() or booking.user.email} during room swap",
                processed_at=timezone.now(),
                processed_by=request.user
            )
            
            # Move leaving tenant to current tenant's old room
            leaving_booking.room = old_room
            leaving_booking.share_no = old_share_no
            # Ensure booking.pg stays in sync with room.pg
            try:
                leaving_booking.pg = old_room.pg
            except Exception:
                pass
            leaving_booking.save(update_fields=['room', 'pg', 'share_no'])
            
            # Update leaving tenant's application if exists
            leaving_app = getattr(leaving_booking, 'application', None)
            if leaving_app and leaving_app.room_id != old_room.id:
                leaving_app.room = old_room
                leaving_app.save(update_fields=['room'])
            
            # Old share becomes VACANT_FROM (occupied by leaving tenant)
            old_share.status = RoomShareStatus.VACANT_FROM
            old_share.vacant_from = leaving_booking.leaving_date
            old_share.save(update_fields=['status', 'vacant_from'])
            
            # New share becomes OCCUPIED (by current tenant, no longer leaving)
            new_share.status = RoomShareStatus.OCCUPIED
            new_share.vacant_from = None
            new_share.save(update_fields=['status', 'vacant_from'])
            
            log(request.user, 'booking_swap', 'Booking', booking.id, 
                f"Swapped with leaving tenant: moved to room {new_room.room_no} bed {share_no}, " +
                f"leaving tenant moved to room {old_room.room_no} bed {old_share_no}")
        else:
            # VACANT_FROM but no leaving booking found, treat as regular vacant
            old_share.status = RoomShareStatus.VACANT
            old_share.vacant_from = None
            old_share.save(update_fields=['status', 'vacant_from'])
            
            new_share.status = RoomShareStatus.OCCUPIED
            new_share.vacant_from = None
            new_share.save(update_fields=['status', 'vacant_from'])
            
            log(request.user, 'booking_swap', 'Booking', booking.id, 
                f"Moved to room {new_room.room_no} bed {share_no} (VACANT_FROM with no active leaving booking)")
    else:
        # Regular swap to VACANT share
        # Free old share
        old_share.status = RoomShareStatus.VACANT
        old_share.vacant_from = None
        old_share.save(update_fields=['status', 'vacant_from'])

        # Occupy new share
        new_share.status = RoomShareStatus.OCCUPIED
        new_share.vacant_from = None
        new_share.save(update_fields=['status', 'vacant_from'])
        
        log(request.user, 'booking_swap', 'Booking', booking.id, 
            f"Swapped to room {new_room.room_no} bed {share_no}")

    # Create swap log for the current tenant
    reason_map = {
        "regular": "Room swap (regular move to vacant bed)",
        "exchange": "Exchanged with leaving tenant",
        "occupied": "Exchanged with occupied tenant"
    }
    RoomSwap.objects.create(
        booking=booking,
        from_room=old_room,
        from_share_no=old_share_no,
        to_room=new_room,
        to_share_no=share_no,
        effective_date=today,
        is_future_swap=False,
        status=RoomSwap.COMPLETED,
        reason=reason_map.get(swap_type, "Room swap"),
        processed_at=timezone.now(),
        processed_by=request.user
    )

    # Update current tenant's booking to new room
    booking.room = new_room
    booking.share_no = share_no
    # Keep booking.pg consistent with the new room
    try:
        booking.pg = new_room.pg
    except Exception:
        pass
    booking.save(update_fields=['room', 'pg', 'share_no'])

    # Update current tenant's application if exists
    app = getattr(booking, 'application', None)
    if app and app.room_id != new_room.id:
        app.room = new_room
        app.save(update_fields=['room'])

    old_share.refresh_from_db()
    new_share.refresh_from_db()

    old_card_html = render_to_string(
        'pgadmin/_tenant_share_card.html',
        {'sd': _build_share_detail(old_room, old_share), 'today': today},
        request=request,
    )
    new_card_html = render_to_string(
        'pgadmin/_tenant_share_card.html',
        {'sd': _build_share_detail(new_room, new_share), 'today': today},
        request=request,
    )

    return JsonResponse({
        'ok': True,
        'action': 'booking_swap',
        'booking_id': booking.id,
        'old_room_id': old_room.id,
        'new_room_id': new_room.id,
        'old_share_no': old_share.share_no,
        'new_share_no': share_no,
        'old_card_html': old_card_html,
        'new_card_html': new_card_html,
        'old_room_counts': _room_share_counts(old_room),
        'new_room_counts': _room_share_counts(new_room),
        'trigger_sync': True,  # Signal to frontend to trigger sync
    })


@login_required
def booking_swap_rooms_api(request, booking_id: int) -> JsonResponse:
    if request.method != 'GET':
        return JsonResponse({'ok': False, 'error': 'Invalid request method.'}, status=405)
    if not _require_pg_admin(request.user):
        return JsonResponse({'ok': False, 'error': 'PG Admin access required.'}, status=403)
    booking = get_object_or_404(Booking, pk=booking_id, status=Booking.APPROVED)
    pg_id = getattr(booking, 'pg_id', None) or getattr(getattr(booking, 'room', None), 'pg_id', None)
    if not _admin_pgs(request.user).filter(id=pg_id).exists():
        return JsonResponse({'ok': False, 'error': 'PG Admin access required for this PG.'}, status=403)

    # Check if we should include occupied shares
    include_occupied = request.GET.get('include_occupied') == 'true'
    
    if include_occupied:
        rooms = (
            Room.objects.filter(pg_id=pg_id)
            .annotate(
                vacant_count=Count('shares', filter=Q(shares__status=RoomShareStatus.VACANT)),
                vacant_from_count=Count('shares', filter=Q(shares__status=RoomShareStatus.VACANT_FROM)),
                occupied_count=Count('shares', filter=Q(shares__status=RoomShareStatus.OCCUPIED))
            )
            .order_by('room_no')
        )
    else:
        rooms = (
            Room.objects.filter(pg_id=pg_id)
            .annotate(
                vacant_count=Count('shares', filter=Q(shares__status=RoomShareStatus.VACANT)),
                vacant_from_count=Count('shares', filter=Q(shares__status=RoomShareStatus.VACANT_FROM))
            )
            .order_by('room_no')
        )

    pending_booking_keys = set(
        Booking.objects.filter(room__pg_id=pg_id, status=Booking.PENDING)
        .values_list('room_id', 'share_no')
    )
    incoming_swap_keys = set(
        RoomSwap.objects.filter(
            to_room__pg_id=pg_id,
            is_future_swap=True,
            status__in=[RoomSwap.PENDING, RoomSwap.APPROVED],
        ).values_list('to_room_id', 'to_share_no')
    )

    room_data = []
    for room in rooms:
        selectable_statuses = [RoomShareStatus.VACANT, RoomShareStatus.VACANT_FROM]
        if include_occupied:
            selectable_statuses.append(RoomShareStatus.OCCUPIED)
        selectable_shares = [
            share for share in room.shares.filter(status__in=selectable_statuses)
            if (room.id, share.share_no) not in pending_booking_keys
            and (room.id, share.share_no) not in incoming_swap_keys
        ]
        vacant_count = sum(share.status == RoomShareStatus.VACANT for share in selectable_shares)
        vacant_from_count = sum(share.status == RoomShareStatus.VACANT_FROM for share in selectable_shares)
        data = {
            'id': room.id,
            'room_no': room.room_no,
            'vacant_count': vacant_count,
            'vacant_from_count': vacant_from_count,
            'total_beds': room.total_shares,
        }
        if include_occupied:
            data['occupied_count'] = sum(share.status == RoomShareStatus.OCCUPIED for share in selectable_shares)
        if selectable_shares:
            room_data.append(data)

    return JsonResponse({
        'ok': True,
        'rooms': room_data,
        'current_room_id': booking.room_id,
        'current_share_no': booking.share_no,
    })



@login_required
def booking_payment_date_update(request, booking_id: int):
    """Allow PG admin to update the monthly payment_date for a booking from the PG admin UI.
    Mirrors finance.monthly_update_payment_date but scoped to PG admin area and redirects back.
    """
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    redirect_url = request.META.get('HTTP_REFERER') or 'pg_tenants'

    if request.method != 'POST':
        message = 'Invalid request method.'
        if is_ajax:
            return JsonResponse({'ok': False, 'error': message}, status=405)
        messages.error(request, message)
        return redirect('pg_tenants')
    if not _require_pg_admin(request.user):
        message = 'PG Admin access required.'
        if is_ajax:
            return JsonResponse({'ok': False, 'error': message}, status=403)
        messages.error(request, message)
        return redirect('pg_my')

    payment_raw = (request.POST.get('payment_date') or '').strip()
    payment_date = parse_date(payment_raw) if payment_raw else None
    if not payment_date:
        message = 'Select a valid payment date.'
        if is_ajax:
            return JsonResponse({'ok': False, 'error': message}, status=400)
        messages.error(request, message)
        return redirect(redirect_url)

    booking = get_object_or_404(Booking, pk=booking_id)
    if not _admin_pgs(request.user).filter(id=getattr(booking, 'pg_id', None)).exists():
        message = 'You do not have access to the requested PG.'
        if is_ajax:
            return JsonResponse({'ok': False, 'error': message}, status=403)
        messages.error(request, message)
        return redirect(redirect_url)

    booking.payment_date = payment_date
    booking.save(update_fields=['payment_date'])
    log(request.user, 'booking_payment_date_updated', 'Booking', booking.id)
    success_message = 'Payment date updated.'
    if is_ajax:
        return JsonResponse({
            'ok': True,
            'message': success_message,
            'value': payment_date.isoformat(),
            'value_iso': payment_date.isoformat(),
            'display': payment_date.isoformat(),
        })
    messages.success(request, success_message)
    return redirect(redirect_url)

@login_required
def pg_referrals(request):
    """PG Admin view: list referral credits for PGs the user administers."""
    user = request.user
    if not _require_pg_admin(user):
        messages.error(request, 'PG Admin access required.')
        return redirect('pg_my')

    # Get PGs the user administers
    admin_pgs = list(_admin_pgs(user))
    if not admin_pgs:
        messages.error(request, 'You are not an admin for any PG.')
        return redirect('pg_my')

    pg_id = request.GET.get('pg')
    if pg_id:
        try:
            pg_id = int(pg_id)
        except (TypeError, ValueError):
            pg_id = None

    # If pg_id provided and user not admin for it, ignore
    if pg_id and not any(p.id == pg_id for p in admin_pgs):
        pg_id = None

    # Prefer explicit pg selection or default to first admin pg
    selected_pg = None
    if pg_id:
        selected_pg = next((p for p in admin_pgs if p.id == pg_id), None)
    else:
        selected_pg = admin_pgs[0]

    credits_qs = ReferralCredit.objects.filter(pg=selected_pg).select_related('referrer_user', 'referred_user', 'referrer_booking', 'referred_booking', 'application').order_by('-created_at')

    redeemed = credits_qs.filter(redeemed_on__isnull=False)
    pending = credits_qs.filter(redeemed_on__isnull=True)

    # Totals
    total_redeemed = redeemed.aggregate(total=Sum('redeemed_amount'))['total'] or 0
    total_pending = pending.aggregate(total=Sum('amount'))['total'] or 0

    context = {
        'pgs': admin_pgs,
        'pg': selected_pg,
        'redeemed': redeemed,
        'pending': pending,
        'total_redeemed': total_redeemed,
        'total_pending': total_pending,
    }
    return render(request, 'pgadmin/referrals.html', context)
    room_data = [
        {
            'id': room.id,
            'room_no': room.room_no,
            'vacant_count': room.vacant_count,
            'total_beds': room.total_shares,
        }
        for room in rooms
        if room.vacant_count > 0
    ]

    return JsonResponse({
        'ok': True,
        'rooms': room_data,
        'current_room_id': booking.room_id,
        'current_share_no': booking.share_no,
    })


@login_required
def booking_swap_shares_api(request, booking_id: int, room_id: int) -> JsonResponse:
    if request.method != 'GET':
        return JsonResponse({'ok': False, 'error': 'Invalid request method.'}, status=405)
    if not _require_pg_admin(request.user):
        return JsonResponse({'ok': False, 'error': 'PG Admin access required.'}, status=403)
    booking = get_object_or_404(Booking, pk=booking_id, status=Booking.APPROVED)
    pg_id = getattr(booking, 'pg_id', None) or getattr(getattr(booking, 'room', None), 'pg_id', None)
    if not _admin_pgs(request.user).filter(id=pg_id).exists():
        return JsonResponse({'ok': False, 'error': 'PG Admin access required for this PG.'}, status=403)

    room = get_object_or_404(Room, pk=room_id, pg_id=pg_id)
    
    # Check if we should include occupied shares
    include_occupied = request.GET.get('include_occupied') == 'true'
    
    if include_occupied:
        # Include VACANT, VACANT_FROM, and OCCUPIED shares (excluding current tenant's bed)
        shares = RoomShareStatus.objects.filter(
            room=room, 
            status__in=[RoomShareStatus.VACANT, RoomShareStatus.VACANT_FROM, RoomShareStatus.OCCUPIED]
        ).exclude(
            room=booking.room,
            share_no=booking.share_no
        ).order_by('share_no')
    else:
        # Include only VACANT and VACANT_FROM shares (excluding current tenant's bed)
        shares = RoomShareStatus.objects.filter(
            room=room, 
            status__in=[RoomShareStatus.VACANT, RoomShareStatus.VACANT_FROM]
        ).exclude(
            room=booking.room,
            share_no=booking.share_no
        ).order_by('share_no')

    claimed_share_numbers = set(
        Booking.objects.filter(room=room, status=Booking.PENDING)
        .values_list('share_no', flat=True)
    )
    claimed_share_numbers.update(
        RoomSwap.objects.filter(
            to_room=room,
            is_future_swap=True,
            status__in=[RoomSwap.PENDING, RoomSwap.APPROVED],
        ).values_list('to_share_no', flat=True)
    )
    shares = shares.exclude(share_no__in=claimed_share_numbers)

    data = []
    for share in shares:
        share_data = {
            'share_no': share.share_no,
            'status': share.status,
            'vacant_from': share.vacant_from.strftime('%Y-%m-%d') if share.vacant_from else None,
            'occupant_name': None,
        }
        
        # If occupied, get the occupant's name
        if share.status == RoomShareStatus.OCCUPIED or share.status == RoomShareStatus.VACANT_FROM:
            occupant_booking = Booking.objects.filter(
                room=room,
                share_no=share.share_no,
                status=Booking.APPROVED
            ).select_related('user__profile').first()
            
            if occupant_booking:
                occupant = occupant_booking.user
                share_data['occupant_name'] = occupant.get_full_name() or occupant.email
                share_data['occupant_phone'] = getattr(occupant.profile, 'phone', None) if hasattr(occupant, 'profile') else None
        
        data.append(share_data)

    return JsonResponse({'ok': True, 'room_id': room.id, 'shares': data})


@login_required
def swap_check_conflict(request):
    """
    Check if there are conflicts for scheduling a swap to a specific bed on a specific date.
    Returns: {valid: bool, conflicts: [messages]}
    """
    if request.method != 'GET':
        return JsonResponse({'valid': False, 'conflicts': ['Invalid request method']})
    
    if not _require_pg_admin(request.user):
        return JsonResponse({'valid': False, 'conflicts': ['Unauthorized']})
    
    try:
        room_id = int(request.GET.get('room', 0))
        share_no = int(request.GET.get('share', 0))
        date_str = request.GET.get('date', '').strip()
    except (ValueError, TypeError):
        return JsonResponse({'valid': False, 'conflicts': ['Invalid parameters']})
    
    if not date_str:
        return JsonResponse({'valid': False, 'conflicts': ['Date is required']})
    
    swap_date = parse_date(date_str)
    if not swap_date:
        return JsonResponse({'valid': False, 'conflicts': ['Invalid date format']})
    
    # Get the PG to verify authorization
    try:
        room = Room.objects.get(pk=room_id)
        if not _admin_pgs(request.user).filter(id=room.pg_id).exists():
            return JsonResponse({'valid': False, 'conflicts': ['Unauthorized for this PG']})
    except Room.DoesNotExist:
        return JsonResponse({'valid': False, 'conflicts': ['Room not found']})
    
    conflicts = []
    today = timezone.now().date()
    
    # Don't allow past dates
    if swap_date < today:
        conflicts.append('Cannot schedule swap for past date')
    
    # Get bookings that have outgoing future swaps (they will vacate the bed)
    # These bookings should be excluded from the "occupied" check
    bookings_with_outgoing_swaps = RoomSwap.objects.filter(
        from_room_id=room_id,
        from_share_no=share_no,
        effective_date__lte=swap_date,
        status__in=[RoomSwap.PENDING, RoomSwap.APPROVED],
        is_future_swap=True
    ).values_list('booking_id', flat=True)
    
    # 1. Check if bed has active booking on that date (excluding those with outgoing swaps)
    active_booking = Booking.objects.filter(
        room_id=room_id,
        share_no=share_no,
        status=Booking.APPROVED,
        joining_date__lte=swap_date
    ).filter(
        Q(leaving_date__isnull=True) | Q(leaving_date__gt=swap_date)
    ).exclude(
        pk__in=bookings_with_outgoing_swaps
    ).exists()
    
    if active_booking:
        conflicts.append('Bed is already occupied on the selected date')
    
    # 2. Check if bed has pending future swap targeting it on or before that date
    pending_swap = RoomSwap.objects.filter(
        to_room_id=room_id,
        to_share_no=share_no,
        effective_date__lte=swap_date,
        status__in=[RoomSwap.PENDING, RoomSwap.APPROVED],
        is_future_swap=True
    ).exists()
    
    if pending_swap:
        conflicts.append('Another future swap is already scheduled for this bed on or before this date')
    
    # 3. Check if bed is reserved for future booking starting on/before swap date (excluding those with outgoing swaps)
    future_booking = Booking.objects.filter(
        room_id=room_id,
        share_no=share_no,
        status=Booking.APPROVED,
        joining_date__lte=swap_date,
        joining_date__gt=today
    ).exclude(
        pk__in=bookings_with_outgoing_swaps
    ).exists()
    
    if future_booking:
        conflicts.append('Bed has a future booking starting on or before the selected date')
    
    # 4. Check share status
    try:
        share = RoomShareStatus.objects.get(room_id=room_id, share_no=share_no)
        if share.status == RoomShareStatus.OCCUPIED:
            # Double-check with booking (already done above, but for consistency)
            pass
        elif share.status == RoomShareStatus.VACANT_FROM:
            if share.vacant_from and swap_date < share.vacant_from:
                conflicts.append(f'Bed will only be vacant from {share.vacant_from.strftime("%Y-%m-%d")}')
    except RoomShareStatus.DoesNotExist:
        conflicts.append('Bed not found')
    
    return JsonResponse({
        'valid': len(conflicts) == 0,
        'conflicts': conflicts
    })


def _require_pg_admin(user):
    # Superusers and website admins can access PG-admin area
    if getattr(user, 'is_superuser', False):
        return True
    if hasattr(user, 'profile') and getattr(user.profile, 'is_website_admin', False):
        return True
    # Fall back to explicit PGAdmin membership (do not rely solely on profile flags)
    try:
        if PGAdmin.objects.filter(user=user).exists():
            return True
    except Exception:
        pass
    return hasattr(user, 'profile') and getattr(user.profile, 'is_pg_admin', False) and getattr(user.profile, 'status', 'active') == 'active'


def _admin_pgs(user):
    # All users (including superusers/website admins) only see PGs they have explicit PGAdmin access to
    return PG.objects.filter(admins__user=user).order_by('name')


def _pg_admin_users_for_pg(pg_id):
    """Return unique user objects for all admins assigned to the given PG."""
    users_by_id = {}
    for admin_rec in PGAdmin.objects.filter(pg_id=pg_id).select_related('user'):
        admin_user = getattr(admin_rec, 'user', None)
        if admin_user and getattr(admin_user, 'id', None):
            users_by_id[admin_user.id] = admin_user
    return list(users_by_id.values())


def _leave_admin_path_and_payload(booking):
    """Build leave-page deep link and push payload with PG and user filters."""
    pg_id = getattr(booking, 'pg_id', None) or getattr(getattr(booking, 'room', None), 'pg_id', None)
    user_obj = getattr(booking, 'user', None)
    search_value = (
        (getattr(user_obj, 'email', '') or '').strip()
        or ((user_obj.get_full_name() or '').strip() if user_obj and hasattr(user_obj, 'get_full_name') else '')
    )

    params = {
        'pg': pg_id,
        'booking_id': booking.id,
        'user_id': getattr(booking, 'user_id', ''),
    }
    if search_value:
        params['search'] = search_value

    path = f"{reverse('pg_leaving_requests_enhanced')}?{urlencode(params)}"

    payload = {
        'pg_id': pg_id,
        'booking_id': booking.id,
        'user_id': getattr(booking, 'user_id', ''),
    }
    if search_value:
        payload['search'] = search_value
    if getattr(user_obj, 'email', None):
        payload['user_email'] = user_obj.email

    return path, payload


def _active_pg(request):
    """Resolve the active PG for a PG Admin user via ?pg=, session, or first available."""
    qs = _admin_pgs(request.user)
    pg = None
    pg_id = request.GET.get('pg') or request.session.get('active_pg_id')
    if pg_id:
        # Only allow selecting PGs within authorized set
        pg = qs.filter(id=pg_id).first()
    if not pg:
        pg = qs.first()
    if pg:
        request.session['active_pg_id'] = pg.id
    return pg


def _booking_drive_urls(booking: Booking) -> set[str]:
    urls: set[str] = set()
    app = getattr(booking, 'application', None)
    if app:
        urls.update([u for u in [getattr(app, 'selfie_url', ''), getattr(app, 'aadhaar_file_url', ''), getattr(app, 'aadhaar_file_url_2', '')] if u])
    profile = getattr(getattr(booking, 'user', None), 'profile', None)
    if profile:
        urls.update([u for u in [getattr(profile, 'selfie_url', ''), getattr(profile, 'aadhaar_file_url', '')] if u])
    return urls


def _build_share_detail(room: Room, share: RoomShareStatus) -> dict:
    booking = None
    occupant = None
    application = None
    future_booking = None
    future_occupant = None
    future_application = None
    today = timezone.now().date()
    
    if share.status in [RoomShareStatus.OCCUPIED, RoomShareStatus.VACANT_FROM]:
        # Current occupant: joined already and hasn't left yet
        booking = (
            Booking.objects.filter(
                room=room, 
                share_no=share.share_no, 
                status=Booking.APPROVED,
                joining_date__lte=today
            )
            .filter(
                Q(leaving_date__isnull=True) | Q(leaving_date__gte=today)
            )
            .select_related('user', 'user__profile')
            .order_by('-created_at')
            .first()
        )
        # If status is VACANT_FROM or there's a leaving_date, check for future reserved booking
        if share.status == RoomShareStatus.VACANT_FROM or (booking and booking.leaving_date):
            future_booking = (
                Booking.objects.filter(
                    room=room,
                    share_no=share.share_no,
                    status=Booking.APPROVED,
                    joining_date__gt=today
                )
                .select_related('user', 'user__profile')
                .order_by('joining_date', '-created_at')
                .first()
            )
            if future_booking:
                future_occupant = getattr(future_booking, 'user', None)
                try:
                    future_application = future_booking.application
                except Exception:
                    future_application = None
    
    elif share.status == RoomShareStatus.RESERVED:
        # First, check if there's a current occupant (booking with leaving_date > today)
        booking = (
            Booking.objects.filter(
                room=room, 
                share_no=share.share_no, 
                status=Booking.APPROVED,
                joining_date__lte=today
            )
            .filter(
                Q(leaving_date__isnull=True) | Q(leaving_date__gte=today)
            )
            .select_related('user', 'user__profile')
            .order_by('-created_at')
            .first()
        )
        
        # If current occupant exists or we should check for future anyway, look for future reserved booking
        future_booking = (
            Booking.objects.filter(
                room=room,
                share_no=share.share_no,
                status=Booking.APPROVED,
                joining_date__gt=today
            )
            .select_related('user', 'user__profile')
            .order_by('joining_date', '-created_at')
            .first()
        )
        
        # If no current occupant found, use the future booking as the main booking
        if not booking:
            booking = future_booking
            future_booking = None  # Don't show as future if it's the main booking
        else:
            # Current occupant exists, so keep future_booking for display
            if future_booking:
                future_occupant = getattr(future_booking, 'user', None)
                try:
                    future_application = future_booking.application
                except Exception:
                    future_application = None
        
        # If still no booking, fall back to pending bookings
        if not booking:
            booking = (
                Booking.objects.filter(room=room, share_no=share.share_no, status=Booking.PENDING)
                .select_related('user', 'user__profile')
                .order_by('-created_at')
                .first()
            )
    
    if booking is not None:
        occupant = getattr(booking, 'user', None)
        try:
            application = booking.application
        except Exception:
            application = None
    
    # Check for pending future swap for this booking
    pending_swap = None
    if booking is not None and booking.status == Booking.APPROVED:
        pending_swap = (
            RoomSwap.objects.filter(
                booking=booking,
                is_future_swap=True,
                status=RoomSwap.PENDING
            )
            .select_related('to_room')
            .first()
        )
    
    return {
        'share': share,
        'booking': booking,
        'occupant': occupant,
        'application': application,
        'is_pending': bool(booking and booking.status == Booking.PENDING),
        'future_booking': future_booking,
        'future_occupant': future_occupant,
        'future_application': future_application,
        'pending_swap': pending_swap,  # Added for showing "Cancel Future Swap" button
    }


def _room_share_counts(room: Room) -> dict:
    counts = {
        'total': 0,
        'vacant': 0,
        'occupied': 0,
        'reserved': 0,
        'leaving': 0,
    }
    for status in RoomShareStatus.objects.filter(room=room).values_list('status', flat=True):
        counts['total'] += 1
        if status == RoomShareStatus.VACANT:
            counts['vacant'] += 1
        elif status == RoomShareStatus.OCCUPIED:
            counts['occupied'] += 1
        elif status == RoomShareStatus.RESERVED:
            counts['reserved'] += 1
        elif status == RoomShareStatus.VACANT_FROM:
            counts['leaving'] += 1
            counts['occupied'] += 1
    return counts


def _room_share_breakdown(room: Room) -> dict:
    counts = {
        'total': room.total_shares,
        'vacant': 0,
        'occupied': 0,
        'reserved': 0,
        'leaving': 0,
    }
    for status in room.shares.values_list('status', flat=True):
        if status == RoomShareStatus.VACANT:
            counts['vacant'] += 1
        elif status == RoomShareStatus.OCCUPIED:
            counts['occupied'] += 1
        elif status == RoomShareStatus.RESERVED:
            counts['reserved'] += 1
        elif status == RoomShareStatus.VACANT_FROM:
            counts['leaving'] += 1
            counts['occupied'] += 1
    counts['non_vacant'] = counts['occupied'] + counts['reserved']
    return counts


def _shrink_room_shares(room: Room, new_total: int) -> tuple[bool, dict | str]:
    shares = list(RoomShareStatus.objects.select_for_update().filter(room=room).order_by('share_no'))
    old_total = len(shares)
    remove_needed = old_total - new_total
    if remove_needed <= 0:
        return True, {'removed_numbers': [], 'reassignments': []}

    vacants = [s for s in shares if s.status == RoomShareStatus.VACANT]
    non_vacant = [s for s in shares if s.status != RoomShareStatus.VACANT]

    if remove_needed > len(vacants):
        shortfall = remove_needed - len(vacants)
        return False, (
            f"Need {remove_needed} vacant bed(s) to shrink, but only {len(vacants)} vacant now. "
            f"Free up {shortfall} more bed(s) or move residents before reducing."
        )

    if len(non_vacant) > new_total:
        overload = len(non_vacant) - new_total
        return False, (
            f"Cannot reduce to {new_total} bed(s) while {len(non_vacant)} bed(s) are occupied/reserved. "
            f"Vacate or relocate {overload} bed(s) first."
        )

    ordered = non_vacant + vacants
    keepers = ordered[:new_total]
    removals = ordered[new_total:]

    # Defensive: removals should be vacant and booking-free
    for share in removals:
        if share.status != RoomShareStatus.VACANT:
            return False, "Unexpected occupied bed selected for removal. Please refresh and retry."
        if Booking.objects.filter(
            room=room,
            share_no=share.share_no,
            status__in=[Booking.PENDING, Booking.APPROVED],
        ).exists():
            return False, "A pending or approved booking is still linked to a bed slated for removal."

    removed_numbers: list[int] = [share.share_no for share in removals]
    for share in removals:
        share.delete()

    target_map = {share.pk: idx + 1 for idx, share in enumerate(keepers)}
    original_numbers = {share.pk: share.share_no for share in keepers}
    change_plan = [
        (share, original_numbers[share.pk], target_map[share.pk])
        for share in keepers
        if original_numbers[share.pk] != target_map[share.pk]
    ]

    temp_base = old_total + 10
    for offset, (share, _original, _target) in enumerate(change_plan, start=1):
        share.share_no = temp_base + offset
        share.save(update_fields=['share_no'])

    for _share, original, target in change_plan:
        Booking.objects.filter(room=room, share_no=original).update(share_no=target)

    for share, _original, target in change_plan:
        share.share_no = target
        share.save(update_fields=['share_no'])

    reassigned = [
        {'from': original, 'to': target, 'status': share.status}
        for share, original, target in change_plan
    ]

    return True, {'removed_numbers': removed_numbers, 'reassignments': reassigned}


def _cleanup_booking_after_leave(booking: Booking, actor=None, origin: str = 'manual') -> dict:
    """Delete booking/application artifacts once a resident has left.
    
    Before deletion, archives basic tenant info to OldTenant for PG admin reference.
    Only creates an archive if the tenant doesn't already exist in OldTenant.
    """
    # ----- Archive tenant data to OldTenant before deletion -----
    old_tenant_created = False
    try:
        from .models import OldTenant
        
        # Check if this booking already exists in OldTenant (avoid duplicates)
        existing_archive = OldTenant.objects.filter(
            pg=booking.pg,
            original_booking_id=booking.id
        ).exists()
        
        if not existing_archive:
            app = getattr(booking, 'application', None)
            
            # Get name from application or user
            full_name = ''
            father_name = ''
            mother_name = ''
            email = ''
            phone = ''
            whatsapp_number = ''
            address = ''
            
            if app:
                full_name = app.name or ''
                father_name = app.father_name or ''
                mother_name = app.mother_name or ''
                email = app.email or ''
                phone = app.phone or ''
                whatsapp_number = app.whatsapp_number or ''
                address = app.address or ''
            
            # Fallback to user data if application data is missing
            if not full_name and booking.user:
                full_name = f"{booking.user.first_name or ''} {booking.user.last_name or ''}".strip() or booking.user.email.split('@')[0]
            if not email and booking.user:
                email = booking.user.email or ''
            
            # Only create OldTenant if we have at least a name
            if full_name:
                OldTenant.objects.create(
                    pg=booking.pg,
                    full_name=full_name,
                    father_name=father_name,
                    mother_name=mother_name,
                    email=email,
                    phone=phone,
                    whatsapp_number=whatsapp_number,
                    address=address,
                    room_no=getattr(getattr(booking, 'room', None), 'room_no', ''),
                    bed_no=str(booking.share_no) if booking.share_no else '',
                    joining_date=booking.joining_date,
                    leaving_date=booking.leaving_date,
                    leaving_reason=booking.leaving_reason or '',
                    advance_paid=booking.advance_paid or 0,
                    advance_returned=booking.advance_returned_amount if booking.advance_returned else 0,
                    original_user=booking.user,
                    original_booking_id=booking.id,
                    archived_by=actor,
                    dob=app.dob if app else None,
                    age=app.age if app else None,
                    father_phone=app.father_phone if app else '',
                    mother_phone=app.mother_phone if app else '',
                    emergency_contact=app.emergency_contact if app else '',
                    food_pref=app.food_pref if app else '',
                    marital_status=app.marital_status if app else '',
                    education=app.education if app else '',
                    occupation=app.occupation if app else '',
                    org_name=app.org_name if app else '',
                    org_address=app.org_address if app else '',
                    has_vehicle=app.has_vehicle if app else False,
                    vehicle_number=app.vehicle_number if app else '',
                    vehicle_model=app.vehicle_model if app else '',
                    aadhaar_number=app.aadhaar_number if app else '',
                    selfie_url=app.selfie_url if app else getattr(booking.user, 'profile', None).selfie_url if hasattr(booking.user, 'profile') else '',
                    aadhaar_file_url=app.aadhaar_file_url if app else getattr(booking.user, 'profile', None).aadhaar_file_url if hasattr(booking.user, 'profile') else '',
                    aadhaar_file_url_2=app.aadhaar_file_url_2 if app else getattr(booking.user, 'profile', None).aadhaar_file_url_2 if hasattr(booking.user, 'profile') else '',
                )
                old_tenant_created = True
    except Exception as e:
        # Don't fail the cleanup if archiving fails
        import logging
        logging.getLogger(__name__).warning(f"Failed to archive tenant data: {e}")
    
    # ----- Original cleanup logic -----
    share = RoomShareStatus.objects.filter(room=booking.room, share_no=booking.share_no).first()
    share_updated = False
    if share:
        share.status = RoomShareStatus.VACANT
        share.vacant_from = None
        share.save(update_fields=['status', 'vacant_from'])
        share_updated = True

    profile = getattr(booking.user, 'profile', None)
    profile_updates: list[str] = []
    if profile and getattr(profile, 'is_pg_user', True):
        profile.is_pg_user = False
        profile_updates.append('is_pg_user')

    drive_urls = [] # do not delete drive urls to keep them for 6 months
    deleted_urls: list[str] = []
    failed_urls: list[str] = []
    for url in drive_urls:
        success = drive_delete(url)
        if success:
            deleted_urls.append(url)
        else:
            failed_urls.append(url)

    if profile:
        if profile.selfie_url and profile.selfie_url in drive_urls:
            profile.selfie_url = ''
            profile_updates.append('selfie_url')
        if profile.aadhaar_file_url and profile.aadhaar_file_url in drive_urls:
            profile.aadhaar_file_url = ''
            profile_updates.append('aadhaar_file_url')
        if profile_updates:
            # Remove duplicates while preserving order
            seen = set()
            ordered_updates = []
            for field in profile_updates:
                if field not in seen:
                    ordered_updates.append(field)
                    seen.add(field)
            profile.save(update_fields=ordered_updates)
            profile_updates = ordered_updates

    app = getattr(booking, 'application', None)
    app_id = getattr(app, 'id', None)
    booking_id = booking.id
    user_id = booking.user_id
    room_no = getattr(getattr(booking, 'room', None), 'room_no', '')
    share_no = booking.share_no

    booking.delete()

    meta = {
        'origin': origin,
        'files_attempted': len(drive_urls),
        'files_deleted': len(deleted_urls),
        'files_failed': len(failed_urls),
        'application_id': app_id,
        'user_id': user_id,
        'share_updated': share_updated,
        'old_tenant_archived': old_tenant_created,
    }
    log(actor, 'leave_cleanup_deleted', 'Booking', booking_id, message=f"Leave cleanup ({origin}) for room {room_no} bed {share_no}", meta=meta)

    return {
        'share_updated': share_updated,
        'deleted_files': len(deleted_urls),
        'failed_files': len(failed_urls),
        'profile_updates': profile_updates,
        'files_attempted': len(drive_urls),
        'old_tenant_archived': old_tenant_created,
    }


@login_required
def my_pg(request):
    if not _require_pg_admin(request.user):
        messages.error(request, "PG Admin access required.")
        return redirect('dashboard')
    # Allow switching if admin of multiple PGs
    pg = _active_pg(request) or PG.objects.filter(created_by_admin=request.user).first()
    
    # AUTO-SYNC: Execute pending future swaps and sync bed statuses when page loads
    if pg:
        try:
            sync_result = _auto_sync_on_page_load(pg, request.user)
            if sync_result.get('swaps_executed', 0) > 0:
                messages.success(
                    request,
                    f"Auto-executed {sync_result['swaps_executed']} scheduled room swap(s)."
                )
        except Exception as e:
            _logger.warning(f"My PG page auto-sync failed: {e}")
    
    if request.method == 'POST':
        form = PGForm(request.POST, instance=pg)
        if form.is_valid():
            pg = form.save(commit=False)
            if not pg.pk:
                pg.created_by_admin = request.user
            pg.save()
            messages.success(request, "PG saved.")
            return redirect('pg_my')
    else:
        form = PGForm(instance=pg)
    quick_url = None
    if pg:
        try:
            quick_url = request.build_absolute_uri(reverse('pg_quick_booking', kwargs={'pgslug': pg.slug}))
        except Exception:
            quick_url = None
    fees = list(Fees.objects.filter(pg=pg)) if pg else []
    return render(request, 'pgadmin/my_pg.html', {"form": form, "pg": pg, "pgs": list(_admin_pgs(request.user)), "quick_booking_url": quick_url, "fees": fees})


@login_required
def tenants(request):
    """My PG Tenants view: shows rooms in ascending order and per-bed occupancy details.
    If admin has multiple PGs, shows a PG selector; otherwise directly shows the active PG.
    Supports server-side filtering by bed status.
    """
    if not _require_pg_admin(request.user):
        messages.error(request, "PG Admin access required.")
        return redirect('dashboard')
    # Allow switching PG via ?pg= param (already handled by _active_pg)
    pg = _active_pg(request)
    
    # AUTO-SYNC: Execute pending future swaps and sync bed statuses when page loads
    # This ensures beds reflect correct status without manual Refresh
    if pg:
        try:
            sync_result = _auto_sync_on_page_load(pg, request.user)
            if sync_result.get('swaps_executed', 0) > 0:
                messages.success(
                    request,
                    f"Auto-executed {sync_result['swaps_executed']} scheduled room swap(s)."
                )
        except Exception as e:
            import logging
            logging.getLogger(__name__).warning(f"Tenants page auto-sync failed: {e}")
    
    # Get filter parameter
    status_filter = request.GET.get('status', 'all').lower()
    valid_filters = ['all', 'vacant', 'occupied', 'reserved', 'leaving']
    if status_filter not in valid_filters:
        status_filter = 'all'
    
    rooms = []
    if pg:
        rooms = (
            Room.objects.filter(pg=pg)
            .prefetch_related('shares', 'bookings__user', 'bookings__user__profile')
            .order_by('room_no')
        )
    # Build a derived structure for template: per room -> counts and bed details
    data = []
    for room in rooms:
        shares = list(room.shares.all())
        
        # Apply server-side filtering
        if status_filter != 'all':
            if status_filter == 'vacant':
                shares = [s for s in shares if s.status == RoomShareStatus.VACANT]
            elif status_filter == 'occupied':
                shares = [s for s in shares if s.status == RoomShareStatus.OCCUPIED]
            elif status_filter == 'reserved':
                shares = [s for s in shares if s.status == RoomShareStatus.RESERVED]
            elif status_filter == 'leaving':
                shares = [s for s in shares if s.status == RoomShareStatus.VACANT_FROM]
        
        # Skip room if no shares match the filter
        if not shares:
            continue
            
        # Counts (for all shares in room, not just filtered)
        all_shares = list(room.shares.all())
        vac = sum(1 for s in all_shares if s.status == RoomShareStatus.VACANT)
        occ = sum(1 for s in all_shares if s.status in [RoomShareStatus.OCCUPIED, RoomShareStatus.VACANT_FROM])
        res = sum(1 for s in all_shares if s.status == RoomShareStatus.RESERVED)
        leaving = sum(1 for s in all_shares if s.status == RoomShareStatus.VACANT_FROM)
        
        # For each share, find latest approved booking for occupant details
        share_details = [_build_share_detail(room, s) for s in sorted(shares, key=lambda x: x.share_no)]
        data.append({
            'room': room,
            'counts': {
                'vacant': vac,
                'occupied': occ,
                'reserved': res,
                'leaving': leaving,
            },
            'shares': share_details,
        })

    ctx = {
        'pg': pg,
        'pgs': list(_admin_pgs(request.user)),
        'rooms': data,
        'today': timezone.now().date(),
        'status_filter': status_filter,
    }
    return render(request, 'pgadmin/tenants.html', ctx)


@login_required
def tenants_export_excel(request):
    """Export tenants in a formatted Excel workbook per PG and month layout described by the user.

    Layout rules implemented:
    - Leave two blank rows at top
    - One merged title row containing Month name, PG name/address/phone in bold center
    - Leave four blank rows after title
    - For each room in ascending room_no: create a block where first column is Room header and
      subsequent columns correspond to bed slots (room.total_shares). For each bed, write resident
      details in a single row (columns: room, name, joining, leaving, phone, email, father name, father phone, mother name, mother phone, ledger link)
    """
    if not _require_pg_admin(request.user):
        return HttpResponse('Forbidden', status=403)

    pg = _active_pg(request)
    if not pg:
        return HttpResponse('No active PG selected', status=400)

    if openpyxl is None:
        return HttpResponse('openpyxl not installed on server', status=500)

    # Determine month to show: current month
    today = timezone.now().date()
    month_name = calendar.month_name[today.month] + f' {today.year}'

    # Build rooms list ordered by room_no
    rooms = list(Room.objects.filter(pg=pg).order_by('room_no').prefetch_related('shares'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Tenants'

    row = 1
    # Leave two blank rows
    row += 2

    # Title: merge across some columns. We'll estimate max columns: for largest room, columns = 1 (room) + max_shares * 1 + ledger
    max_shares = max((r.total_shares or 1) for r in rooms) if rooms else 1
    cols_needed = 1 + max_shares * 1 + 10  # allocate extra for fields; final layout uses fixed columns per resident
    last_col = get_column_letter(max(6, cols_needed))

    title_cell = ws.cell(row=row, column=1)
    title_cell.value = f"{month_name} — {pg.name} • {pg.address or ''} • {pg.phone or ''}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(8, max_shares + 4))

    row += 1

    # Leave four blank rows
    row += 4

    # Column header template for each resident row
    headers = ['Room', 'Name', 'Joining', 'Leaving', 'Phone', 'Email', 'Father name', 'Father phone', 'Mother name', 'Mother phone', 'Ledger']

    # For each room, create rows. For each bed in room.total_shares, create one resident row.
    for room in rooms:
        # Room block header: write room number in bold
        ws.cell(row=row, column=1, value=f"Room {room.room_no}")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        # Write header row
        for ci, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = Font(bold=True)
        row += 1

        # For each bed number, use _build_share_detail to get both current and future bookings
        shares = list(room.shares.order_by('share_no'))
        # Ensure we iterate from 1..room.total_shares
        total = room.total_shares or len(shares) or 1
        for share_no in range(1, total + 1):
            # Get share object
            share = next((s for s in shares if s.share_no == share_no), None)
            
            # Use _build_share_detail to get booking details including future booking
            share_detail = _build_share_detail(room, share) if share else {}
            booking = share_detail.get('booking')
            future_booking = share_detail.get('future_booking')
            
            # default empty row: include room and bed number
            values = [f"{room.room_no} - (BED {share_no})"] + [''] * (len(headers) - 1)

            # If current booking exists, populate row
            if booking:
                user = booking.user
                app = share_detail.get('application')
                # Populate fields - add "(Current)" prefix if future booking exists
                name_prefix = "(Current) " if future_booking else ""
                values[1] = name_prefix + f"{user.first_name} {user.last_name}".strip()
                joining = booking.joining_date or booking.start_date
                values[2] = joining.isoformat() if joining else ''
                values[3] = booking.leaving_date.isoformat() if booking.leaving_date else ''
                # phone/email preference: application.phone -> profile.phone -> user.email
                phone = getattr(app, 'phone', None) or getattr(getattr(user, 'profile', None), 'phone', '')
                values[4] = phone
                values[5] = getattr(app, 'email', None) or user.email or ''
                # parents info from application if present
                values[6] = getattr(app, 'father_name', '') if app else ''
                values[7] = getattr(app, 'father_phone', '') if app else ''
                values[8] = getattr(app, 'mother_name', '') if app else ''
                values[9] = getattr(app, 'mother_phone', '') if app else ''
                # Ledger link: reverse finance_ledger
                try:
                    ledger_url = request.build_absolute_uri(reverse('finance_ledger', kwargs={'user_id': user.id}))
                except Exception:
                    ledger_url = ''
                values[10] = ledger_url

            # write the current booking row
            for ci, v in enumerate(values, start=1):
                cell = ws.cell(row=row, column=ci)
                if ci == 11 and v:
                    # Ledger hyperlink
                    cell.value = 'Ledger'
                    cell.hyperlink = v
                    cell.font = Font(color='0000EE', underline='single')
                    cell.alignment = Alignment(horizontal='left')
                else:
                    cell.value = v
            row += 1
            
            # If future booking exists, add an additional row for next tenant
            if future_booking:
                future_user = future_booking.user
                future_app = share_detail.get('future_application')
                future_values = [f"{room.room_no} - (BED {share_no})"] + [''] * (len(headers) - 1)
                
                # Populate future tenant fields with "(Next)" prefix
                future_values[1] = "(Next) " + f"{future_user.first_name} {future_user.last_name}".strip()
                future_joining = future_booking.joining_date or future_booking.start_date
                future_values[2] = future_joining.isoformat() if future_joining else ''
                future_values[3] = future_booking.leaving_date.isoformat() if future_booking.leaving_date else ''
                # phone/email preference: application.phone -> profile.phone -> user.email
                future_phone = getattr(future_app, 'phone', None) or getattr(getattr(future_user, 'profile', None), 'phone', '')
                future_values[4] = future_phone
                future_values[5] = getattr(future_app, 'email', None) or future_user.email or ''
                # parents info from application if present
                future_values[6] = getattr(future_app, 'father_name', '') if future_app else ''
                future_values[7] = getattr(future_app, 'father_phone', '') if future_app else ''
                future_values[8] = getattr(future_app, 'mother_name', '') if future_app else ''
                future_values[9] = getattr(future_app, 'mother_phone', '') if future_app else ''
                # Ledger link for future tenant
                try:
                    future_ledger_url = request.build_absolute_uri(reverse('finance_ledger', kwargs={'user_id': future_user.id}))
                except Exception:
                    future_ledger_url = ''
                future_values[10] = future_ledger_url
                
                # write the future booking row
                for ci, v in enumerate(future_values, start=1):
                    cell = ws.cell(row=row, column=ci)
                    if ci == 11 and v:
                        # Ledger hyperlink
                        cell.value = 'Ledger'
                        cell.hyperlink = v
                        cell.font = Font(color='0000EE', underline='single')
                        cell.alignment = Alignment(horizontal='left')
                    else:
                        cell.value = v
                row += 1

        # leave two blank rows between rooms
        row += 2

    # Adjust column widths a bit
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    # Prepare response
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = ''.join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in (pg.name or 'pg'))
    filename = f"{safe_name.replace(' ', '_')}_tenants_{today.isoformat()}.xlsx"
    resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


@login_required
def tenants_export_pdf(request):
    """Export tenants as a structured portrait PDF with room-by-room cards.
    Optimized for large datasets with streaming and async processing.

    Layout:
    - Portrait mode (A4)
    - Header: PG name (bold, large), address (medium), phone (medium bold), current month/year (bold) on separate lines
    - For each room: room number header, then resident cards (one per bed)
    - Each card: selfie (left), checkbox (right), details (center): name, phone, joining, payment, leaving
    """
    if not _require_pg_admin(request.user):
        return HttpResponse('Forbidden', status=403)

    pg = _active_pg(request)
    if not pg:
        return HttpResponse('No active PG selected', status=400)

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm, inch
        from reportlab.pdfgen import canvas
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether, Image as RLImage, Flowable
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_LEFT, TA_CENTER
        from reportlab.lib import colors
        import io as _io
        import requests
        from datetime import datetime
        from concurrent.futures import ThreadPoolExecutor, as_completed
    except Exception as e:
        return HttpResponse(f"PDF generation dependencies missing: {e}", status=500)

    today = timezone.now().date()
    current_month_year = datetime.now().strftime('%B %Y')  # e.g., "October 2025"
    
    # Calculate current month range for expected rent calculation
    import calendar
    year = today.year
    month = today.month
    m_first = date(year, month, 1)
    last_day = calendar.monthrange(year, month)[1]
    m_last = date(year, month, last_day)
    
    # PERFORMANCE FIX: Skip images for large PGs to avoid 502 timeout
    total_rooms = Room.objects.filter(pg=pg).count()
    SKIP_IMAGES = total_rooms > 50  # Skip images if more than 50 rooms
    
    # Fetch data with select_related for optimization
    rooms = list(Room.objects.filter(pg=pg).order_by('room_no').prefetch_related(
        Prefetch('shares', queryset=RoomShareStatus.objects.all())
    ))

    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=10*mm, bottomMargin=10*mm, leftMargin=12*mm, rightMargin=12*mm)
    story = []
    styles = getSampleStyleSheet()

    # Custom styles - reduced font sizes and spacing
    pg_name_style = ParagraphStyle('PGName', parent=styles['Heading1'], fontSize=14, textColor=colors.HexColor("#000000"), alignment=TA_CENTER, spaceAfter=2)
    pg_address_style = ParagraphStyle('PGAddress', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#1a1a1a'), fontName='Helvetica-Bold', alignment=TA_CENTER, spaceAfter=1)
    pg_phone_style = ParagraphStyle('PGPhone', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#1a1a1a'), fontName='Helvetica-Bold', alignment=TA_CENTER, spaceAfter=1)
    pg_month_style = ParagraphStyle('PGMonth', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#000000'), fontName='Helvetica-Bold', alignment=TA_CENTER, spaceAfter=6)
    room_header_style = ParagraphStyle('RoomHeader', parent=styles['Heading2'], fontSize=11, textColor=colors.HexColor('#2c5aa0'), spaceAfter=3, spaceBefore=5)

    # PG Header
    story.append(Paragraph(f"<b>{pg.name}</b>", pg_name_style))
    story.append(Paragraph(pg.address or '', pg_address_style))
    story.append(Paragraph(f"<b>{pg.phone or ''}</b>", pg_phone_style))
    story.append(Paragraph(f"<b>{current_month_year}</b>", pg_month_style))
    story.append(Spacer(1, 4*mm))

    # Custom Flowable for outlined checkbox
    class OutlinedCheckbox(Flowable):
        def __init__(self, size=4*mm):
            Flowable.__init__(self)
            self.size = size
            self.width = size
            self.height = size
        
        def draw(self):
            self.canv.setStrokeColor(colors.black)
            self.canv.setFillColor(colors.white)
            self.canv.setLineWidth(0.5)
            self.canv.rect(0, 0, self.size, self.size, stroke=1, fill=1)

    # Helper to download image and create RLImage - with caching
    _image_cache = {}
    def _get_image(url, default_width=18*mm, default_height=22*mm):
        # PERFORMANCE FIX: Skip images for large PGs
        if SKIP_IMAGES:
            return None
        if not url:
            return None
        if url in _image_cache:
            return _image_cache[url]
        try:
            # Normalize drive/dropbox URLs
            if 'drive.google.com' in url and '/file/d/' in url:
                fid = url.split('/file/d/')[1].split('/')[0]
                url = f'https://drive.google.com/uc?export=download&id={fid}'
            elif 'dropbox.com' in url:
                url = url.replace('www.dropbox.com', 'dl.dropboxusercontent.com').replace('?dl=0', '')
            resp = requests.get(url, timeout=2, stream=True)  # Reduced from 5s to 2s
            resp.raise_for_status()
            img_data = _io.BytesIO(resp.content)
            img = RLImage(img_data, width=default_width, height=default_height)
            _image_cache[url] = img
            return img
        except Exception:
            return None

    # Pre-fetch all bookings in bulk for optimization
    all_booking_ids = []
    room_share_map = {}
    for room in rooms:
        total_shares = room.total_shares or 1
        for share_no in range(1, total_shares + 1):
            room_share_map[(room.id, share_no)] = None
    
    # Bulk query bookings
    bookings_qs = Booking.objects.filter(
        room__in=rooms,
        status__in=[Booking.APPROVED, Booking.PENDING]
    ).select_related('user', 'user__profile', 'application').order_by('-created_at')
    
    for booking in bookings_qs:
        key = (booking.room_id, booking.share_no)
        if key in room_share_map and room_share_map[key] is None:
            if not booking.leaving_date or booking.leaving_date >= today:
                room_share_map[key] = booking

    # Pre-download images in parallel for better performance (skip if large PG)
    image_urls = set()
    if not SKIP_IMAGES:
        for booking in room_share_map.values():
            if booking:
                user = booking.user
                app = getattr(booking, 'application', None)
                selfie_url = getattr(app, 'selfie_url', None) or getattr(getattr(user, 'profile', None), 'selfie_url', None)
                if selfie_url:
                    image_urls.add(selfie_url)
        
        # Download images in parallel (max 3 concurrent, reduced from 5)
        if image_urls:
            with ThreadPoolExecutor(max_workers=3) as executor:
                future_to_url = {executor.submit(_get_image, url): url for url in image_urls}
                for future in as_completed(future_to_url, timeout=20):  # 20s max total
                    try:
                        future.result(timeout=2)  # 2s per image
                    except Exception:
                        pass  # Image download failed, will show placeholder

    # Build cards for each room
    for room in rooms:
        story.append(Paragraph(f"Room {room.room_no}", room_header_style))
        total_shares = room.total_shares or 1

        # Collect all cards for this room
        all_cards = []
        
        for share_no in range(1, total_shares + 1):
            share = room.shares.filter(share_no=share_no).first()
            share_detail = _build_share_detail(room, share) if share else {}
            booking = share_detail.get('booking')
            future_booking = share_detail.get('future_booking')

            # Build single card with 3 columns: [selfie | details | checkbox]
            if booking:
                user = booking.user
                app = getattr(booking, 'application', None)
                selfie_url = getattr(app, 'selfie_url', None) or getattr(getattr(user, 'profile', None), 'selfie_url', None)
                selfie_img = _image_cache.get(selfie_url) if selfie_url else None

                name = f"{user.first_name} {user.last_name}".strip() or user.email
                phone = getattr(app, 'phone', None) or getattr(getattr(user, 'profile', None), 'phone', '') or ''
                joining = booking.joining_date or booking.start_date
                joining_str = joining.strftime('%d/%m/%y') if joining else '—'
                payment = booking.payment_date
                payment_str = payment.strftime('%d/%m/%y') if payment else '—'
                leaving = booking.leaving_date
                leaving_str = leaving.strftime('%d/%m/%y') if leaving else '—'
                
                # Get monthly fee for this tenant (expected rent for current month)
                # This calculation works even with leaving_date - it will be pro-rated
                from finance.views import _expected_rent_for_user_pg_month
                try:
                    monthly_fee = _expected_rent_for_user_pg_month(user, pg, booking, m_first, m_last, today=today)
                    
                    # If current month expected is 0 (e.g., tenant joining next month or already left)
                    # calculate next payment cycle
                    next_cycle_fee = 0.0
                    next_payment_date = None
                    if monthly_fee == 0 or monthly_fee < 0.01:
                        # Calculate next month range
                        next_month = month + 1 if month < 12 else 1
                        next_year = year if month < 12 else year + 1
                        next_m_first = date(next_year, next_month, 1)
                        next_last_day = calendar.monthrange(next_year, next_month)[1]
                        next_m_last = date(next_year, next_month, next_last_day)
                        
                        # Only calculate next cycle if tenant will still be present
                        # (no leaving date or leaving date is after next cycle starts)
                        if not leaving or leaving >= next_m_first:
                            try:
                                next_cycle_fee = _expected_rent_for_user_pg_month(user, pg, booking, next_m_first, next_m_last, today=today)
                                # Get the payment date for next cycle
                                if booking.payment_date:
                                    next_payment_date = booking.payment_date
                            except Exception:
                                pass
                except Exception:
                    # Fallback to static fee if calculation fails
                    from finance.models import ResidentRate, Fees
                    resident_rate = ResidentRate.objects.filter(user=user, pg=pg, active=True).first()
                    if resident_rate:
                        monthly_fee = float(resident_rate.amount)
                    else:
                        room_share_type = str(room.total_shares or '')
                        fee_obj = Fees.objects.filter(pg=pg, share_type=room_share_type).first()
                        monthly_fee = float(fee_obj.amount) if fee_obj else 0.0
                    next_cycle_fee = 0.0
                    next_payment_date = None

                # Column 1: Selfie
                if selfie_img:
                    selfie_cell = selfie_img
                else:
                    selfie_cell = Paragraph("<i>No Photo</i>", ParagraphStyle('TinyText', parent=styles['Normal'], fontSize=7, alignment=TA_CENTER))
                
                # Column 2: Details
                detail_style = ParagraphStyle('CardDetail', parent=styles['Normal'], fontSize=7, leading=9, wordWrap='CJK')
                advance_amount = booking.advance_paid if booking.advance_paid is not None else Decimal('0')
                advance_str = f"Rs.{advance_amount:.2f}"

                # Build monthly fee display
                if monthly_fee > 0:
                    monthly_fee_str = f"Monthly Fee: Rs.{monthly_fee:.0f}"
                elif next_cycle_fee > 0:
                    # Show next cycle fee with payment date
                    if next_payment_date:
                        monthly_fee_str = f"Monthly Fee: Rs.{next_cycle_fee:.0f} (Next: {next_payment_date.strftime('%d/%m/%y')})"
                    else:
                        monthly_fee_str = f"Monthly Fee: Rs.{next_cycle_fee:.0f} (Next cycle)"
                else:
                    monthly_fee_str = "Monthly Fee: Rs.0"

                details_lines = [
                    f"<b>{name}</b>",
                    f"Phone: {phone}",
                    f"Join: {joining_str}",
                    f"Pay: {payment_str}",
                    f"Leave: {leaving_str}",
                    f"Advance: {advance_str}",
                    monthly_fee_str
                ]
                
                details_cell = Paragraph("<br/>".join(details_lines), detail_style)
                
                # Column 3: Checkbox
                checkbox_cell = OutlinedCheckbox(size=4*mm)
                
                # Build card for current booking
                single_card_data = [[selfie_cell, details_cell, checkbox_cell]]
                single_card = Table(single_card_data, colWidths=[18*mm, 35*mm, 7*mm], rowHeights=[26*mm])
                single_card.setStyle(TableStyle([
                    ('VALIGN', (0, 0), (0, 0), 'MIDDLE'),
                    ('VALIGN', (1, 0), (1, 0), 'TOP'),
                    ('VALIGN', (2, 0), (2, 0), 'MIDDLE'),
                    ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                    ('ALIGN', (2, 0), (2, 0), 'CENTER'),
                    ('BOX', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('LEFTPADDING', (0, 0), (-1, -1), 2),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                    ('TOPPADDING', (0, 0), (-1, -1), 2),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ]))
                all_cards.append(single_card)
                
                # Add future booking card separately if exists
                if future_booking:
                    future_user = future_booking.user
                    future_name = f"{future_user.first_name} {future_user.last_name}".strip() or future_user.email
                    future_joining = future_booking.joining_date
                    future_joining_str = future_joining.strftime('%d/%m/%y') if future_joining else '—'
                    future_leaving = future_booking.leaving_date
                    
                    # Get future tenant's photo
                    future_app = getattr(future_booking, 'application', None)
                    future_selfie_url = getattr(future_app, 'selfie_url', None) or getattr(getattr(future_user, 'profile', None), 'selfie_url', None)
                    future_selfie_img = _image_cache.get(future_selfie_url) if future_selfie_url else None
                    
                    # Get monthly fee for future tenant (expected rent for current month)
                    # This calculation works even with leaving_date - it will be pro-rated
                    try:
                        future_monthly_fee = _expected_rent_for_user_pg_month(future_user, pg, future_booking, m_first, m_last, today=today)
                        
                        # If current month expected is 0, calculate next payment cycle
                        future_next_cycle_fee = 0.0
                        future_next_payment_date = None
                        if future_monthly_fee == 0 or future_monthly_fee < 0.01:
                            # Calculate next month range
                            next_month = month + 1 if month < 12 else 1
                            next_year = year if month < 12 else year + 1
                            next_m_first = date(next_year, next_month, 1)
                            next_last_day = calendar.monthrange(next_year, next_month)[1]
                            next_m_last = date(next_year, next_month, next_last_day)
                            
                            # Only calculate next cycle if tenant will still be present
                            if not future_leaving or future_leaving >= next_m_first:
                                try:
                                    future_next_cycle_fee = _expected_rent_for_user_pg_month(future_user, pg, future_booking, next_m_first, next_m_last, today=today)
                                    # Get the payment date for next cycle
                                    if future_booking.payment_date:
                                        future_next_payment_date = future_booking.payment_date
                                except Exception:
                                    pass
                    except Exception:
                        # Fallback: try to get custom rate or use same as current tenant
                        from finance.models import ResidentRate
                        future_resident_rate = ResidentRate.objects.filter(user=future_user, pg=pg, active=True).first()
                        if future_resident_rate:
                            future_monthly_fee = float(future_resident_rate.amount)
                        else:
                            future_monthly_fee = monthly_fee  # Use same as current if not set
                        future_next_cycle_fee = 0.0
                        future_next_payment_date = None
                    
                    # Future tenant selfie
                    if future_selfie_img:
                        future_selfie_cell = future_selfie_img
                    else:
                        future_selfie_cell = Paragraph("<i>No Photo</i>", ParagraphStyle('TinyText', parent=styles['Normal'], fontSize=7, alignment=TA_CENTER))
                    
                    # Future tenant details
                    # Build monthly fee display for future tenant
                    if future_monthly_fee > 0:
                        future_fee_str = f"Monthly Fee: Rs.{future_monthly_fee:.0f}"
                    elif future_next_cycle_fee > 0:
                        # Show next cycle fee with payment date
                        if future_next_payment_date:
                            future_fee_str = f"Monthly Fee: Rs.{future_next_cycle_fee:.0f} (Next: {future_next_payment_date.strftime('%d/%m/%y')})"
                        else:
                            future_fee_str = f"Monthly Fee: Rs.{future_next_cycle_fee:.0f} (Next cycle)"
                    else:
                        future_fee_str = "Monthly Fee: Rs.0"
                    
                    future_details_lines = [
                        "<b>---NEXT---</b>",
                        f"<b>{future_name}</b>",
                        f"Join: {future_joining_str}",
                        future_fee_str
                    ]
                    future_details_cell = Paragraph("<br/>".join(future_details_lines), detail_style)
                    
                    # Future tenant checkbox
                    future_checkbox_cell = OutlinedCheckbox(size=4*mm)
                    
                    # Build separate card for future booking
                    future_card_data = [[future_selfie_cell, future_details_cell, future_checkbox_cell]]
                    future_card = Table(future_card_data, colWidths=[18*mm, 35*mm, 7*mm], rowHeights=[22*mm])
                    future_card.setStyle(TableStyle([
                        ('VALIGN', (0, 0), (0, 0), 'MIDDLE'),
                        ('VALIGN', (1, 0), (1, 0), 'TOP'),
                        ('VALIGN', (2, 0), (2, 0), 'MIDDLE'),
                        ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                        ('ALIGN', (2, 0), (2, 0), 'CENTER'),
                        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#90EE90')),
                        ('LEFTPADDING', (0, 0), (-1, -1), 2),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                        ('TOPPADDING', (0, 0), (-1, -1), 2),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                    ]))
                    all_cards.append(future_card)
            else:
                # Empty bed card
                vacant_text = Paragraph("<i>VACANT</i>", ParagraphStyle('VacantText', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER, textColor=colors.grey))
                empty_detail = Paragraph(f"<i>Bed {share_no}</i>", ParagraphStyle('EmptyDetail', parent=styles['Normal'], fontSize=7, textColor=colors.grey))
                checkbox_cell = OutlinedCheckbox(size=4*mm)
                
                empty_card_data = [[vacant_text, empty_detail, checkbox_cell]]
                empty_card = Table(empty_card_data, colWidths=[18*mm, 35*mm, 7*mm], rowHeights=[26*mm])
                empty_card.setStyle(TableStyle([
                    ('VALIGN', (0, 0), (0, 0), 'MIDDLE'),
                    ('VALIGN', (1, 0), (1, 0), 'MIDDLE'),
                    ('VALIGN', (2, 0), (2, 0), 'MIDDLE'),
                    ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                    ('ALIGN', (1, 0), (1, 0), 'LEFT'),
                    ('ALIGN', (2, 0), (2, 0), 'CENTER'),
                    ('BOX', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('LEFTPADDING', (0, 0), (-1, -1), 2),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                    ('TOPPADDING', (0, 0), (-1, -1), 2),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ]))
                all_cards.append(empty_card)
        
        # Arrange cards in rows of 3
        cards_per_row = 3
        for i in range(0, len(all_cards), cards_per_row):
            row_cards = all_cards[i:i+cards_per_row]
            # Pad row if less than 3 cards
            while len(row_cards) < cards_per_row:
                row_cards.append(Paragraph("", styles['Normal']))  # empty placeholder
            
            # Create row table
            row_table = Table([row_cards], colWidths=[60*mm, 60*mm, 60*mm])
            row_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 2),
                ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ]))
            story.append(row_table)
            story.append(Spacer(1, 2*mm))

        # Add space between rooms
        story.append(Spacer(1, 3*mm))

    # Build PDF with error handling
    try:
        doc.build(story)
    except Exception as e:
        # Fallback: minimal PDF with error
        buf2 = _io.BytesIO()
        c = canvas.Canvas(buf2, pagesize=A4)
        c.drawString(50, 800, f"PDF generation error: {e}")
        c.showPage()
        c.save()
        pdf = buf2.getvalue()
        buf2.close()
        resp = HttpResponse(pdf, content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="tenants_error.pdf"'
        return resp

    pdf = buf.getvalue()
    buf.close()
    safe_name = ''.join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in (pg.name or 'pg'))
    filename = f"{safe_name.replace(' ', '_')}_tenants_{current_month_year.replace(' ', '_')}.pdf"
    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


@login_required
def quick_booking_qr_pdf(request, pg_id: int):
    """Generate a single-page PDF containing a QR code for the PG's quick booking URL."""
    if not _require_pg_admin(request.user):
        messages.error(request, "PG Admin access required.")
        return redirect('dashboard')
    pg = get_object_or_404(PG, pk=pg_id)
    # Ensure user is admin of this PG
    if not _admin_pgs(request.user).filter(id=pg.id).exists():
        messages.error(request, 'PG Admin access required for this PG.')
        return redirect('pg_my')

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import mm
        from reportlab.graphics.barcode import qr as rl_qr
        from reportlab.graphics.shapes import Drawing
        from reportlab.graphics import renderPDF
        from reportlab.pdfbase.pdfmetrics import stringWidth
        import io as _io
    except Exception as e:
        return HttpResponse(f"PDF generation dependencies missing: {e}", status=500)

    quick_url = request.build_absolute_uri(reverse('pg_quick_booking', kwargs={'pgslug': pg.slug}))

    buf = _io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4

    # Header: PG name and label text centered at top
    top_margin = 12 * mm
    side_margin = 10 * mm
    footer_space = 10 * mm
    # Draw title (PG name) with auto-scaling if too long
    title_font = "Helvetica-Bold"
    title_size = 18.0
    max_title_width = width - 2 * side_margin
    while title_size > 10 and stringWidth(pg.name, title_font, title_size) > max_title_width:
        title_size -= 1
    c.setFont(title_font, title_size)
    title_y = height - top_margin
    c.drawCentredString(width / 2.0, title_y, pg.name)

    # Subtitle: Quick Booking
    subtitle_font = "Helvetica"
    subtitle_size = 12.0
    max_subtitle_width = width - 2 * side_margin
    while subtitle_size > 8 and stringWidth("Quick Booking", subtitle_font, subtitle_size) > max_subtitle_width:
        subtitle_size -= 1
    c.setFont(subtitle_font, subtitle_size)
    subtitle_y = title_y - 8 * mm
    c.drawCentredString(width / 2.0, subtitle_y, "Quick Booking")

    # Reserve header space and compute available area for QR
    header_space = max((height - subtitle_y) + 8 * mm, 30 * mm)  # ensure some space under subtitle
    available_height = height - header_space - footer_space

    # Create a large QR under header, within margins, centered
    qr_widget = rl_qr.QrCodeWidget(quick_url)
    bounds = qr_widget.getBounds()
    w = bounds[2] - bounds[0]
    h = bounds[3] - bounds[1]
    size = min(width - 2 * side_margin, available_height)
    d = Drawing(size, size, transform=[size / w, 0, 0, size / h, 0, 0])
    d.add(qr_widget)
    x = (width - size) / 2.0
    y = footer_space + (available_height - size) / 2.0
    renderPDF.draw(d, c, x, y)

    c.showPage()
    c.save()
    pdf = buf.getvalue()
    buf.close()
    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{pg.name}_quick_booking_qr.pdf"'
    return resp


@login_required
def rooms_list(request):
    if not _require_pg_admin(request.user):
        messages.error(request, "PG Admin access required.")
        return redirect('dashboard')
    pg = _active_pg(request)
    
    # AUTO-SYNC: Execute pending future swaps and sync bed statuses when page loads
    if pg:
        try:
            sync_result = _auto_sync_on_page_load(pg, request.user)
            if sync_result.get('swaps_executed', 0) > 0:
                messages.success(
                    request,
                    f"Auto-executed {sync_result['swaps_executed']} scheduled room swap(s)."
                )
        except Exception as e:
            _logger.warning(f"Rooms list page auto-sync failed: {e}")
    
    if pg:
        rooms = (
            Room.objects.filter(pg=pg)
            .annotate(
                occupied_count=Count('shares', filter=Q(shares__status__in=[RoomShareStatus.OCCUPIED, RoomShareStatus.VACANT_FROM])),
                reserved_count=Count('shares', filter=Q(shares__status=RoomShareStatus.RESERVED)),
                vacant_count=Count('shares', filter=Q(shares__status=RoomShareStatus.VACANT)),
                leaving_count=Count('shares', filter=Q(shares__status=RoomShareStatus.VACANT_FROM)),
                next_vacant_from=Min('shares__vacant_from', filter=Q(shares__status=RoomShareStatus.VACANT_FROM)),
            )
            .prefetch_related(
                Prefetch(
                    'shares',
                    queryset=RoomShareStatus.objects.filter(status=RoomShareStatus.VACANT_FROM).select_related('room').only('id','share_no','vacant_from','room').order_by('vacant_from'),
                    to_attr='vacant_from_shares',
                )
            )
            .order_by('room_no')
        )
        # Attach booking/user info to each vacant_from share for display
        for room in rooms:
            if hasattr(room, 'vacant_from_shares'):
                for share in room.vacant_from_shares:
                    # Find the approved booking for this share
                    booking = Booking.objects.filter(
                        room=room, 
                        share_no=share.share_no, 
                        status=Booking.APPROVED,
                        leaving_date__isnull=False
                    ).select_related('user').first()
                    share.leaving_booking = booking
                    if booking:
                        share.leaving_user_name = booking.user.get_full_name() or booking.user.email
                    else:
                        share.leaving_user_name = None
    else:
        rooms = []
    return render(request, 'pgadmin/rooms_list.html', {"pg": pg, "rooms": rooms, "pgs": list(_admin_pgs(request.user)), "active_filter": (request.GET.get('filter') or '')})


@login_required
def vehicle_search(request):
    if not _require_pg_admin(request.user):
        messages.error(request, "PG Admin access required.")
        return redirect('dashboard')
    pg = _active_pg(request)
    compiled_results: list[dict] = []
    total_results = 0
    if pg:
        today = timezone.localdate()
        qs = (
            ResidentApplication.objects.filter(pg=pg, has_vehicle=True)
            .filter(booking__isnull=False)
            .filter(Q(booking__leaving_date__isnull=True) | Q(booking__leaving_date__gt=today))
            .select_related('booking', 'booking__room', 'booking__user', 'booking__user__profile')
            .order_by('name', 'vehicle_number')
        )
        for app in qs:
            booking = getattr(app, 'booking', None)
            user = getattr(booking, 'user', None) if booking else None
            profile = getattr(user, 'profile', None) if user else None
            if app.name:
                resident_name = app.name
            elif user:
                full_name = getattr(user, 'get_full_name', lambda: '')()
                resident_name = (full_name or '').strip() or getattr(user, 'email', '')
            else:
                resident_name = ''
            room = getattr(booking, 'room', None) if booking else None
            contact_phone = app.phone or (getattr(profile, 'phone', '') if profile else '')
            compiled_results.append({
                'app': app,
                'resident_name': resident_name,
                'room_no': getattr(room, 'room_no', ''),
                'contact_phone': contact_phone,
                'vehicle_number': app.vehicle_number or '',
                'vehicle_model': app.vehicle_model or '',
                'search_blob': ' '.join(filter(None, [
                    resident_name,
                    getattr(user, 'email', '') if user else '',
                    contact_phone,
                    getattr(room, 'room_no', '') if room else '',
                    str(getattr(booking, 'share_no', '')) if booking else '',
                    app.vehicle_number or '',
                    app.vehicle_model or '',
                ])).lower(),
            })
        total_results = len(compiled_results)
    ctx = {
        'pg': pg,
        'pgs': list(_admin_pgs(request.user)),
        'compiled_results': compiled_results,
        'total_results': total_results,
    }
    return render(request, 'pgadmin/vehicle_search.html', ctx)


@login_required
@transaction.atomic
def room_create(request):
    if not _require_pg_admin(request.user):
        messages.error(request, "PG Admin access required.")
        return redirect('dashboard')
    pg = _active_pg(request)
    if request.method == 'POST':
        form = RoomForm(request.POST)
        if form.is_valid():
            room = form.save(commit=False)
            room.pg = pg
            room.save()
            # Ensure bed rows exist
            for i in range(1, room.total_shares + 1):
                RoomShareStatus.objects.get_or_create(room=room, share_no=i)
            messages.success(request, "Room created.")
            return redirect('pg_rooms')
    else:
        form = RoomForm()
    return render(request, 'pgadmin/room_form.html', {"form": form, "pg": pg, "pgs": list(_admin_pgs(request.user))})


@login_required
@transaction.atomic
def room_edit(request, pk):
    if not _require_pg_admin(request.user):
        messages.error(request, "PG Admin access required.")
        return redirect('dashboard')
    room = get_object_or_404(Room, pk=pk)
    if not _admin_pgs(request.user).filter(id=room.pg_id).exists():
        messages.error(request, "PG Admin access required for this PG.")
        return redirect('dashboard')
    share_stats = _room_share_breakdown(room)
    if request.method == 'POST':
        old_total = room.total_shares
        form = RoomForm(request.POST, instance=room)
        if form.is_valid():
            new_total = form.cleaned_data['total_shares']
            room_obj = form.save(commit=False)

            if new_total == old_total:
                room_obj.save()
                messages.success(request, "Room details updated.")
                return redirect('pg_rooms')

            if new_total > old_total:
                room_obj.save()
                for share_no in range(old_total + 1, new_total + 1):
                    RoomShareStatus.objects.get_or_create(
                        room=room_obj,
                        share_no=share_no,
                        defaults={'status': RoomShareStatus.VACANT},
                    )
                added = new_total - old_total
                log(request.user, 'room_size_increase', 'Room', room_obj.id, f'Beds increased by {added}.')
                messages.success(request, f"Room updated. Added {added} new vacant bed{'s' if added != 1 else ''}.")
                return redirect('pg_rooms')

            success, payload = _shrink_room_shares(room, new_total)
            if success:
                room_obj.save()
                removed = payload.get('removed_numbers', [])
                reassigned = payload.get('reassignments', [])
                log_msg = f"Beds reduced to {new_total}. Removed {len(removed)} bed(s)."
                if reassigned:
                    moves = ', '.join(f"{item['from']}→{item['to']}" for item in reassigned)
                    log_msg += f" Reassigned beds: {moves}."
                log(request.user, 'room_size_decrease', 'Room', room_obj.id, log_msg)
                message = f"Room updated. Removed {len(removed)} vacant bed{'s' if len(removed) != 1 else ''}."
                if reassigned:
                    message += " Occupied beds renumbered automatically."
                messages.success(request, message)
                return redirect('pg_rooms')

            form.add_error('total_shares', payload)
            room.refresh_from_db()
            share_stats = _room_share_breakdown(room)
    else:
        form = RoomForm(instance=room)
    return render(request, 'pgadmin/room_form.html', {"form": form, "room": room, "share_stats": share_stats})


@login_required
def application_pdf(request, app_id):
    """Generate a structured PDF of a resident application for PG Admins.
    Includes selfie (if available) and appends Aadhaar PDF pages when possible.
    """
    app = get_object_or_404(ResidentApplication, pk=app_id)
    # Authorization: must be PG Admin for this PG
    user = request.user
    pg_id = getattr(app, 'pg_id', None)
    if not _require_pg_admin(user) or not _admin_pgs(user).filter(id=pg_id).exists():
        messages.error(request, 'PG Admin access required for this PG.')
        return redirect('pg_resident_applications')

    # Lazy imports to avoid hard dependency at import-time
    try:
        from io import BytesIO
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
        from reportlab.lib import colors
        import requests, re
        from urllib.parse import urlparse, parse_qs
    except Exception as e:
        return HttpResponse(f"PDF generation dependencies missing: {e}", status=500)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    story = []

    title = f"Resident Application — {app.name or app.user.first_name} {app.user.last_name}"
    story.append(Paragraph(title, styles['Title']))
    story.append(Spacer(1, 6))
    story.append(Paragraph(f"PG: {app.pg.name}", styles['Normal']))
    # Payment date: prefer booking.payment_date, fallback to booking.joining_date or start_date
    booking = getattr(app, 'booking', None)
    payment_date = None
    if booking:
        payment_date = getattr(booking, 'payment_date', None) or getattr(booking, 'joining_date', None) or getattr(booking, 'start_date', None)
    story.append(Paragraph(f"Room: {getattr(app.room, 'room_no', '—')} • Share: {getattr(getattr(app, 'booking', None), 'share_no', '—')}", styles['Normal']))
    story.append(Paragraph(f"Payment date: {payment_date or '—'}", styles['Normal']))
    story.append(Spacer(1, 12))

    # Helpers for downloading
    def _normalize_direct_url(url: str) -> str:
        if not url:
            return url
        try:
            pu = urlparse(url)
            host = pu.netloc.lower()
            # Google Drive: file/d/<id>/view or open?id=<id>
            if 'drive.google.com' in host:
                m = re.search(r"/file/d/([\w-]+)/", pu.path)
                file_id = None
                if m:
                    file_id = m.group(1)
                else:
                    qs = parse_qs(pu.query)
                    file_id = (qs.get('id') or qs.get('file_id') or [None])[0]
                if file_id:
                    return f"https://drive.google.com/uc?export=download&id={file_id}"
            # Dropbox: ensure dl=1
            if 'dropbox.com' in host and 'dl=0' in url:
                return url.replace('dl=0', 'dl=1')
        except Exception:
            return url
        return url

    def _download(url: str, accept: str) -> bytes | None:
        if not url:
            return None
        url2 = _normalize_direct_url(url)
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
            'Accept': accept,
            'Referer': '',
        }
        try:
            r = requests.get(url2, headers=headers, timeout=20, allow_redirects=True)
            if r.ok and r.content:
                return r.content
        except Exception:
            return None
        return None

    def _download_with_type(url: str, accept: str):
        if not url:
            return None, None
        url2 = _normalize_direct_url(url)
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
            'Accept': accept,
            'Referer': '',
        }
        try:
            r = requests.get(url2, headers=headers, timeout=20, allow_redirects=True)
            if r.ok and r.content:
                return r.content, r.headers.get('content-type', '').lower()
        except Exception:
            return None, None
        return None, None

    # Selfie
    if getattr(app, 'selfie_url', None):
        selfie_bytes = _download(app.selfie_url, 'image/*')
        if selfie_bytes:
            try:
                img = RLImage(BytesIO(selfie_bytes))
                img._restrictSize(180, 180)
                story.append(Paragraph("Selfie:", styles['Heading4']))
                story.append(img)
                story.append(Spacer(1, 8))
            except Exception:
                story.append(Paragraph("Selfie: [Unable to load image binary]", styles['Normal']))
        else:
            # Fallback: include link when download not possible
            story.append(Paragraph(f"Selfie: <link href='{app.selfie_url}'>Open online</link>", styles['Normal']))

    # Info table
    info = [
        ["Name", f"{app.name or app.user.first_name} {app.user.last_name}"],
        ["DOB", f"{app.dob or '—'}"],
        ["Age", f"{app.age or '—'}"],
        ["Phone", f"{app.phone or getattr(getattr(app.user, 'profile', None), 'phone', '—')}"],
        ["Email", f"{app.email or app.user.email}"],
        ["Food Pref", f"{app.food_pref or '—'}"],
        ["Marital", f"{app.marital_status or '—'}"],
        ["Date of Application", f"{app.date_of_admission or '—'}"],
        ["Joining Date", f"{getattr(app.booking, 'joining_date', None) or getattr(app.booking, 'start_date', None) or '—'}"],
        ["Address", f"{app.address or '—'}"],
    ]
    table = Table(info, hAlign='LEFT', colWidths=[120, 360])
    table.setStyle(TableStyle([
        ('FONT', (0,0), (-1,-1), 'Helvetica', 9),
        ('BACKGROUND', (0,0), (-1,0), colors.whitesmoke),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.whitesmoke]),
        ('GRID', (0,0), (-1,-1), 0.25, colors.lightgrey),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING', (0,0), (-1,-1), 4),
        ('RIGHTPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('TOPPADDING', (0,0), (-1,-1), 3),
    ]))
    story.append(table)
    story.append(Spacer(1, 10))

    # Family
    fam = [
        ["Father Name", f"{app.father_name or '—'}"],
        ["Father Phone or Emergency Contact 1", f"{app.father_phone or '—'}"],
        ["Mother Name", f"{app.mother_name or '—'}"],
        ["Mother Phone or Emergency Contact 2", f"{app.mother_phone or '—'}"],
    ]
    story.append(Paragraph("Family", styles['Heading4']))
    # Make column 1 wider to accommodate long labels (e.g., "Mother Phone or Emergency Contact 2")
    story.append(Table(
        fam,
        colWidths=[180, 300],
        style=[
            ('FONT', (0,0), (-1,-1), 'Helvetica', 9),
            ('GRID', (0,0), (-1,-1), 0.25, colors.lightgrey),
            ('ROWBACKGROUNDS', (0,0), (-1,-1), [colors.white, colors.whitesmoke])
        ]
    ))
    story.append(Spacer(1, 10))

    # Education / Work
    work = [
        ["Occupation", f"{app.occupation or '—'}"],
        ["Education", f"{app.education or '—'}"],
        ["Organisation", f"{app.org_name or '—'}"],
        ["Org Address", f"{app.org_address or '—'}"],
    ]
    story.append(Paragraph("Education / Work", styles['Heading4']))
    story.append(Table(work, colWidths=[120, 360], style=[('FONT', (0,0), (-1,-1), 'Helvetica', 9), ('GRID', (0,0), (-1,-1), 0.25, colors.lightgrey), ('ROWBACKGROUNDS', (0,0), (-1,-1), [colors.white, colors.whitesmoke])]))
    story.append(Spacer(1, 10))

    # Vehicle
    vehicle_rows = [["Has Vehicle", "Yes" if app.has_vehicle else "No"]]
    if app.has_vehicle:
        vehicle_rows.append(["Number", f"{app.vehicle_number or '—'}"])
        vehicle_rows.append(["Model", f"{app.vehicle_model or '—'}"])
    story.append(Paragraph("Vehicle", styles['Heading4']))
    story.append(Table(vehicle_rows, colWidths=[120, 360], style=[('FONT', (0,0), (-1,-1), 'Helvetica', 9), ('GRID', (0,0), (-1,-1), 0.25, colors.lightgrey), ('ROWBACKGROUNDS', (0,0), (-1,-1), [colors.white, colors.whitesmoke])]))
    story.append(Spacer(1, 10))

    # Aadhaar summary and prefetch document before building main PDF
    story.append(Paragraph("Documents", styles['Heading4']))
    story.append(Paragraph(f"Aadhaar/Other Card Number: {app.aadhaar_number or '—'}", styles['Normal']))

    aadhaar_pdf_bytes = None
    if getattr(app, 'aadhaar_file_url', None):
        content, ctype = _download_with_type(app.aadhaar_file_url, 'application/pdf, image/*')
        if content:
            # Heuristics to detect PDF vs Image
            is_pdf = content[:4] == b'%PDF' or (ctype and 'pdf' in ctype)
            if is_pdf:
                aadhaar_pdf_bytes = content
                story.append(Spacer(1, 6))
                story.append(Paragraph("Document: (attached PDF will be appended)", styles['Italic']))
            else:
                try:
                    story.append(Spacer(1, 6))
                    story.append(Paragraph("Document Image:", styles['Heading4']))
                    aimg = RLImage(BytesIO(content))
                    aimg._restrictSize(420, 560)  # fit nicely on A4
                    story.append(aimg)
                except Exception:
                    story.append(Paragraph(f"Document: <link href='{app.aadhaar_file_url}'>Open online</link>", styles['Normal']))
        else:
            story.append(Paragraph(f"Document: <link href='{app.aadhaar_file_url}'>Open online</link>", styles['Normal']))

    # Optional second image (back side)
    if getattr(app, 'aadhaar_file_url_2', None):
        try:
            content2, ctype2 = _download_with_type(app.aadhaar_file_url_2, 'image/*, application/octet-stream')
            if content2:
                story.append(Spacer(1, 6))
                story.append(Paragraph("Back Side Image:", styles['Heading4']))
                img2 = RLImage(BytesIO(content2))
                img2._restrictSize(420, 560)
                story.append(img2)
            else:
                story.append(Paragraph(f"Back Side: <link href='{app.aadhaar_file_url_2}'>Open online</link>", styles['Normal']))
        except Exception:
            story.append(Paragraph(f"Back Side: <link href='{app.aadhaar_file_url_2}'>Open online</link>", styles['Normal']))

    # Declarations
    decls = [
        ["Valuables", "Agreed" if app.decl_valuables else "Not agreed"],
        ["Notice", "Agreed" if app.decl_notice else "Not agreed"],
        ["Deposit", "Agreed" if app.decl_deposit else "Not agreed"],
        ["Truth", "Agreed" if app.decl_truth else "Not agreed"],
    ]
    story.append(Spacer(1, 6))
    story.append(Paragraph("Declarations", styles['Heading4']))
    story.append(Table(decls, colWidths=[120, 360], style=[('FONT', (0,0), (-1,-1), 'Helvetica', 9), ('GRID', (0,0), (-1,-1), 0.25, colors.lightgrey), ('ROWBACKGROUNDS', (0,0), (-1,-1), [colors.white, colors.whitesmoke])]))

    # Build main doc
    doc.build(story)
    main_pdf = buf.getvalue()

    # Attempt to merge PDFs if possible
    final_bytes = main_pdf
    if aadhaar_pdf_bytes:
        try:
            from pypdf import PdfReader, PdfWriter
            main_reader = PdfReader(BytesIO(main_pdf))
            out = PdfWriter()
            for p in main_reader.pages:
                out.add_page(p)
            aadhaar_reader = PdfReader(BytesIO(aadhaar_pdf_bytes))
            for p in aadhaar_reader.pages:
                out.add_page(p)
            out_buf = BytesIO()
            out.write(out_buf)
            final_bytes = out_buf.getvalue()
        except Exception:
            # Fallback: keep main PDF only if merge fails
            final_bytes = main_pdf

    filename = f"Resident-Application-{(app.name or app.user.first_name or 'User').replace(' ', '-')}-{app.id}.pdf"
    resp = HttpResponse(final_bytes, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


@login_required
def room_shares(request, pk):
    if not _require_pg_admin(request.user):
        messages.error(request, "PG Admin access required.")
        return redirect('dashboard')
    room = get_object_or_404(Room, pk=pk)
    if not _admin_pgs(request.user).filter(id=room.pg_id).exists():
        messages.error(request, "PG Admin access required for this PG.")
        return redirect('dashboard')
    shares = room.shares.order_by('share_no')
    forms = [ShareStatusForm(prefix=f"s{rs.id}", instance=rs) for rs in shares]
    if request.method == 'POST':
        any_saved = False
        for rs in shares:
            prev_status = rs.status
            form = ShareStatusForm(request.POST, prefix=f"s{rs.id}", instance=rs)
            if form.is_valid():
                new_status = form.cleaned_data.get('status')
                # If moving from vacant/vacant_from to reserved/occupied, collect user details and create or link booking
                if prev_status in [RoomShareStatus.VACANT, RoomShareStatus.VACANT_FROM] and new_status in [RoomShareStatus.RESERVED, RoomShareStatus.OCCUPIED]:
                    email = request.POST.get(f"s{rs.id}-new-email", "").strip()
                    first_name = request.POST.get(f"s{rs.id}-new-first_name", "").strip()
                    last_name = request.POST.get(f"s{rs.id}-new-last_name", "").strip()
                    phone = request.POST.get(f"s{rs.id}-new-phone", "").strip()
                    joining_raw = request.POST.get(f"s{rs.id}-new-joining", "").strip()
                    joining_date = parse_date(joining_raw) if joining_raw else None

                    if not email:
                        messages.error(request, f"Bed {rs.share_no}: Email is required to set {new_status}.")
                        continue

                    User = get_user_model()
                    user = User.objects.filter(email__iexact=email).first()
                    if not user:
                        # Create a lightweight account; user can later sign in via Google with same email
                        # Ensure unique username
                        base_username = email.split('@')[0][:20] or 'user'
                        username = base_username
                        suffix = 1
                        while User.objects.filter(username=username).exists():
                            suffix += 1
                            username = f"{base_username}{suffix}"
                        user = User(username=username, email=email, first_name=first_name or '', last_name=last_name or '')
                        user.set_unusable_password()
                        user.save()
                        created_user = True
                    else:
                        created_user = False
                        updated = False
                        if first_name and user.first_name != first_name:
                            user.first_name = first_name
                            updated = True
                        if last_name and user.last_name != last_name:
                            user.last_name = last_name
                            updated = True
                        if updated:
                            user.save(update_fields=['first_name', 'last_name'])

                    # Ensure an EmailAddress record exists for allauth and mark verified for seamless linking on future login
                    if EmailAddress is not None:
                        ea, _ = EmailAddress.objects.get_or_create(user=user, email=user.email, defaults={'verified': True, 'primary': True})
                        # If not primary/verified, make it so
                        to_update = []
                        if not ea.primary:
                            ea.primary = True
                            to_update.append('primary')
                        if not ea.verified:
                            ea.verified = True
                            to_update.append('verified')
                        if to_update:
                            ea.save(update_fields=to_update)

                    # Ensure profile exists and update phone
                    profile = getattr(user, 'profile', None)
                    if profile:
                        if phone and profile.phone != phone:
                            profile.phone = phone
                            profile.save(update_fields=['phone'])
                    any_saved = True

                    # Create the booking according to status
                    try:
                        booking_status = Booking.APPROVED if new_status == RoomShareStatus.OCCUPIED else Booking.PENDING
                        booking = Booking(
                            user=user,
                            room=room,
                            share_no=rs.share_no,
                            status=booking_status,
                            joining_date=joining_date,
                        )
                        if booking_status == Booking.APPROVED:
                            booking.start_date = joining_date or timezone.now().date()
                        booking.save()
                    except IntegrityError:
                        messages.error(request, f"Bed {rs.share_no}: Could not create booking. User may already have an active booking in this PG.")
                        # Skip saving status change for this share
                        continue

                    # Save the bed status change after booking created and clear vacant_from if set
                    saved_rs = form.save()
                    if getattr(saved_rs, 'vacant_from', None):
                        saved_rs.vacant_from = None
                        saved_rs.save(update_fields=['vacant_from'])
                    # Feedback
                    if created_user:
                        messages.success(request, f"Bed {rs.share_no}: User created: {user.email}, booking: {booking.get_status_display()}.")
                    else:
                        messages.success(request, f"Bed {rs.share_no}: User linked: {user.email}, booking: {booking.get_status_display()}.")
                else:
                    # Normal save and inline occupant updates when already occupied
                    saved_rs = form.save()
                    # If status changed away from VACANT_FROM, clear the date
                    if prev_status == RoomShareStatus.VACANT_FROM and saved_rs.status != RoomShareStatus.VACANT_FROM and getattr(saved_rs, 'vacant_from', None):
                        saved_rs.vacant_from = None
                        saved_rs.save(update_fields=['vacant_from'])
                    any_saved = True
                    occ = (
                        Booking.objects.filter(room=room, share_no=rs.share_no, status=Booking.APPROVED)
                        .select_related('user', 'user__profile')
                        .order_by('-created_at')
                        .first()
                    )
                    if occ:
                        phone_key = f"s{rs.id}-phone"
                        leaving_key = f"s{rs.id}-leaving"
                        phone_val = request.POST.get(phone_key)
                        leaving_val = request.POST.get(leaving_key)
                        if phone_val is not None and hasattr(occ.user, 'profile'):
                            if occ.user.profile.phone != phone_val:
                                occ.user.profile.phone = phone_val
                                occ.user.profile.save(update_fields=['phone'])
                        if leaving_val is not None:
                            new_date = parse_date(leaving_val) if leaving_val else None
                            if occ.leaving_date != new_date:
                                occ.leaving_date = new_date
                                occ.save(update_fields=['leaving_date'])
        if any_saved:
            messages.success(request, "Changes saved.")
            return redirect('pg_room_shares', pk=room.id)
    # Build (share, form, occupant) where occupant is latest approved booking for this share
    rs_forms = []
    for rs, form in zip(shares, forms):
        occupant = None
        if rs.status in [RoomShareStatus.OCCUPIED, RoomShareStatus.VACANT_FROM]:
            occupant = (
                Booking.objects.filter(room=room, share_no=rs.share_no, status=Booking.APPROVED)
                .select_related('user', 'room', 'user__profile')
                .order_by('-created_at')
                .first()
            )
        # Determine application status for occupant booking
        app_exists = False
        if occupant:
            from bookings.models import ResidentApplication
            app_exists = ResidentApplication.objects.filter(booking=occupant).exists()
        rs_forms.append((rs, form, occupant, app_exists))
    return render(request, 'pgadmin/room_shares.html', {"room": room, "forms": forms, "shares": shares, "rs_forms": rs_forms})


# --- Booking approvals (PG Admin) ---

@login_required
def bookings_pending(request):
    if not _require_pg_admin(request.user):
        messages.error(request, "PG Admin access required.")
        return redirect('dashboard')
    pg = _active_pg(request)
    pending = []
    active_residents = []
    if pg:
        # Get all pending bookings (both regular and day-wise)
        pending = (
            Booking.objects.filter(status=Booking.PENDING, room__pg=pg)
            .select_related('user', 'room', 'assigned_by')
            .prefetch_related('application', 'application__status_history')
            .annotate(has_application=Exists(ResidentApplication.objects.filter(booking_id=OuterRef('pk'))))
            .order_by('booking_type', '-created_at')  # Day-wise first, then regular
        )
        
        # Get active residents for referral selection
        # Active = APPROVED bookings with no leaving_date or future leaving_date
        today = timezone.now().date()
        active_residents = (
            Booking.objects.filter(
                status=Booking.APPROVED,
                room__pg=pg
            )
            .filter(
                Q(leaving_date__isnull=True) | Q(leaving_date__gte=today)
            )
            .select_related('user', 'room')
            .order_by('user__first_name', 'user__email')
            .distinct()
        )
    
    return render(request, 'pgadmin/bookings_pending.html', {
        "pg": pg,
        "bookings": pending,
        "pgs": list(_admin_pgs(request.user)),
        "active_residents": active_residents,
    })


@login_required
def bookings_confirmed(request):
    """Show confirmed/approved bookings for the active PG.

    Lists current and future approved bookings so admins can review confirmed residents.
    """
    if not _require_pg_admin(request.user):
        messages.error(request, "PG Admin access required.")
        return redirect('dashboard')
    pg = _active_pg(request)
    confirmed = []
    referral_options = []
    if pg:
        app_qs = ResidentApplication.objects.select_related(
            'referred_by_booking',
            'referred_by_booking__user',
            'referred_by_booking__room'
        ).prefetch_related('status_history')

        confirmed = (
            Booking.objects.filter(status=Booking.APPROVED, room__pg=pg)
            .select_related('user', 'user__profile', 'room', 'assigned_by')
            .prefetch_related(Prefetch('application', queryset=app_qs))
            .annotate(has_application=Exists(ResidentApplication.objects.filter(booking_id=OuterRef('pk'))))
            .order_by('start_date', 'joining_date')
        )
        
        # Get active residents for referral selection
        today = timezone.now().date()
        referral_options = (
            Booking.objects.filter(
                status=Booking.APPROVED,
                room__pg=pg
            )
            .filter(Q(leaving_date__isnull=True) | Q(leaving_date__gt=today))
            .select_related('user', 'room', 'user__profile')
            .order_by('user__first_name', 'user__last_name', 'room__room_no', 'share_no')
            .distinct()
        )

    return render(request, 'pgadmin/bookings_confirmed.html', {
        'pg': pg,
        'bookings': confirmed,
        'pgs': list(_admin_pgs(request.user)),
        'referral_options': referral_options,
    })


@login_required
@transaction.atomic
def booking_approve(request, booking_id):
    if not _require_pg_admin(request.user):
        messages.error(request, "PG Admin access required.")
        return redirect('dashboard')
    booking = get_object_or_404(Booking, pk=booking_id, status=Booking.PENDING)
    # Enforce PG membership
    if not _admin_pgs(request.user).filter(id=(getattr(booking, 'pg_id', None) or getattr(getattr(booking, 'room', None), 'pg_id', None))).exists():
        messages.error(request, "PG Admin access required for this PG.")
        return redirect('dashboard')
    
    pg = booking.pg or booking.room.pg
    
    # Handle day-wise bookings differently - need room assignment
    if booking.booking_type == Booking.DAYWISE:
        if request.method == 'POST':
            # Process room assignment for day-wise booking
            room_id = request.POST.get('room_id')
            share_no = request.POST.get('share_no')
            payment_amount = request.POST.get('payment_amount', '0')
            payment_received = request.POST.get('payment_received') == 'on'
            
            if not room_id or not share_no:
                messages.error(request, "Please select a room and bed.")
                return redirect('pg_booking_approve', booking_id=booking.id)
            
            try:
                room = Room.objects.get(id=room_id, pg=pg)
                share_no = int(share_no)
                share = RoomShareStatus.objects.get(room=room, share_no=share_no)
                
                # Update booking with assigned room
                booking.room = room
                booking.share_no = share_no
                booking.status = Booking.APPROVED
                booking.assigned_by = request.user
                booking.assigned_at = timezone.now()
                booking.payment_received = payment_received
                if payment_amount:
                    booking.payment_amount = float(payment_amount)
                booking.save()
                
                # Update share status to OCCUPIED (day-wise bookings start immediately)
                share.status = RoomShareStatus.OCCUPIED
                share.save(update_fields=['status'])
                
                # Update application status
                if hasattr(booking, 'application'):
                    booking.application.status = ResidentApplication.CONFIRMED
                    booking.application.save(update_fields=['status'])
                
                # Create Payment record if payment was received
                if payment_received and payment_amount:
                    from finance.models import Payment
                    try:
                        amount_decimal = float(payment_amount)
                        Payment.objects.create(
                            user=booking.user,
                            pg=pg,
                            amount=amount_decimal,
                            date=timezone.now().date(),
                            status='success',
                            mode='cash',  # Default to cash, admin can edit later
                            type='daywise',  # Use 'daywise' type instead of 'fee'
                            notes=f"Day-wise booking payment for Room {room.room_no} Bed {share_no}",
                            from_date=booking.joining_date,
                            to_date=booking.leaving_date,
                            booking=booking
                        )
                    except Exception as e:
                        # Log but don't fail the approval
                        messages.warning(request, f"Booking approved but payment record failed: {str(e)}")
                
                log(request.user, 'daywise_booking_approved', 'Booking', booking.id, 
                    f"Approved day-wise booking for room {room.room_no} bed {share_no}")

                Notification.objects.create(
                    user=booking.user,
                    title="Day-wise Booking Approved",
                    message=f"Your day-wise booking has been approved and assigned to Room {room.room_no}, Bed {share_no}."
                )
                send_push_to_user(
                    booking.user,
                    title="Day-wise Booking Approved",
                    body=f"Assigned to Room {room.room_no}, Bed {share_no}.",
                    url=reverse('booking_detail', args=[booking.id]),
                    extra_data={'type': 'booking_approved', 'booking_type': 'daywise', 'booking_id': booking.id},
                )
                
                messages.success(request, f"Day-wise booking approved and assigned to Room {room.room_no}, Bed {share_no}.")
                return redirect('pg_bookings_pending')
                
            except (Room.DoesNotExist, RoomShareStatus.DoesNotExist, ValueError) as e:
                messages.error(request, f"Invalid room or bed selection: {str(e)}")
                return redirect('pg_booking_approve', booking_id=booking.id)
        else:
            # GET request - show room selection form
            from datetime import datetime, time as dt_time

            vacant_rooms = Room.objects.filter(pg=pg).prefetch_related('shares')

            # Compute booking start/end datetimes and human summary
            booking_start_dt = None
            booking_end_dt = None
            booking_summary = None
            try:
                if booking.joining_date:
                    st_time = booking.start_time or dt_time(0, 0)
                    en_time = booking.end_time or dt_time(0, 0)
                    booking_start_dt = datetime.combine(booking.joining_date, st_time)
                    booking_end_dt = datetime.combine(booking.leaving_date or booking.joining_date, en_time)
                    # If end is before start, leave booking_summary None (template will show dates)
                    if booking_end_dt >= booking_start_dt:
                        delta = booking_end_dt - booking_start_dt
                        total_minutes = int(delta.total_seconds() // 60)
                        days = total_minutes // (60 * 24)
                        hours = (total_minutes - days * 24 * 60) // 60
                        minutes = total_minutes - days * 24 * 60 - hours * 60
                        parts = []
                        if days:
                            parts.append(f"{days} day" + ("s" if days != 1 else ""))
                        if hours:
                            parts.append(f"{hours} hour" + ("s" if hours != 1 else ""))
                        if minutes:
                            parts.append(f"{minutes} minute" + ("s" if minutes != 1 else ""))
                        human = ' '.join(parts) if parts else '0 minutes'
                        total_hours = round(delta.total_seconds() / 3600, 2)
                        booking_summary = {
                            'human': human,
                            'total_hours': total_hours,
                            'start_dt': booking_start_dt,
                            'end_dt': booking_end_dt,
                        }
            except Exception:
                booking_summary = None

            # Get vacant shares but only include VACANT_FROM shares that are available by booking.start
            # Also exclude shares that have overlapping bookings (PENDING/APPROVED)
            # Also exclude shares that have pending future swaps
            from bookings.models import Booking as BookingModel, RoomSwap
            pending_booking_keys = pending_booking_share_keys(pg=pg)
            
            # Get all beds with pending/approved future swaps
            beds_with_future_swaps = set(
                RoomSwap.objects.filter(
                    to_room__pg=pg,
                    is_future_swap=True,
                    status__in=[RoomSwap.PENDING, RoomSwap.APPROVED]
                ).values_list('to_room_id', 'to_share_no')
            )
            
            vacant_shares = []
            for room in vacant_rooms:
                for share in room.shares.filter(status__in=[RoomShareStatus.VACANT, RoomShareStatus.VACANT_FROM]):
                    if (room.id, share.share_no) in pending_booking_keys:
                        continue

                    # Skip if this bed has a pending future swap
                    if (room.id, share.share_no) in beds_with_future_swaps:
                        continue
                    
                    include = False
                    if share.status == RoomShareStatus.VACANT:
                        include = True
                    elif share.status == RoomShareStatus.VACANT_FROM:
                        # Only include if vacant_from is set and is <= booking joining date
                        if share.vacant_from and booking.joining_date and share.vacant_from <= booking.joining_date:
                            include = True
                    if not include:
                        continue

                    # Now exclude if there are existing bookings on this share that overlap the requested period
                    overlap = False
                    try:
                        if booking_start_dt and booking_end_dt:
                            other_qs = BookingModel.objects.filter(
                                room=room,
                                share_no=share.share_no,
                                status__in=[BookingModel.PENDING, BookingModel.APPROVED]
                            ).exclude(pk=booking.id)
                            for other in other_qs:
                                # Determine other booking start/end datetimes
                                try:
                                    other_start = None
                                    other_end = None
                                    o_st_time = other.start_time or dt_time(0, 0)
                                    o_en_time = other.end_time or dt_time(0, 0)
                                    if other.joining_date:
                                        other_start = datetime.combine(other.joining_date, o_st_time)
                                    elif other.start_date:
                                        other_start = datetime.combine(other.start_date, o_st_time)
                                    if other.leaving_date:
                                        other_end = datetime.combine(other.leaving_date, o_en_time)
                                    # If other_end is None, treat as open-ended -> overlap
                                    if other_start and (other_end is None or other_end >= booking_start_dt) and other_start <= booking_end_dt:
                                        overlap = True
                                        break
                                except Exception:
                                    # If any parsing fails, be conservative and mark overlap
                                    overlap = True
                                    break
                    except Exception:
                        overlap = True

                    if overlap:
                        # skip this share because it conflicts with an existing booking
                        continue

                    vacant_shares.append({
                        'room': room,
                        'share': share,
                        'room_id': room.id,
                        'share_no': share.share_no,
                    })

            context = {
                'pg': pg,
                'booking': booking,
                'vacant_shares': vacant_shares,
                'booking_summary': booking_summary,
                'pgs': list(_admin_pgs(request.user)),
            }
            return render(request, 'pgadmin/booking_approve_daywise.html', context)
    
    # Regular booking approval (existing logic)
    from decimal import Decimal, InvalidOperation

    share = get_object_or_404(RoomShareStatus, room=booking.room, share_no=booking.share_no)

    confirm_application = False
    collect_advance = False
    advance_amount_value = Decimal('0')
    advance_mode = 'upi'
    advance_notes = ''

    if request.method == 'POST':
        confirm_application = request.POST.get('confirm_application') == 'on'
        collect_advance = request.POST.get('collect_advance') == 'on'
        advance_mode = (request.POST.get('advance_mode') or 'upi').lower()
        advance_notes = (request.POST.get('advance_notes') or '').strip()
        if advance_mode not in ('upi', 'cash', 'bank'):
            advance_mode = 'upi'
        if collect_advance:
            raw_amount = (request.POST.get('advance_amount') or '').strip()
            try:
                advance_amount_value = Decimal(raw_amount)
            except (InvalidOperation, ValueError):
                messages.error(request, "Enter a valid advance amount.")
                return redirect('pg_bookings_pending')
            if advance_amount_value <= 0:
                messages.error(request, "Advance amount must be greater than zero.")
                return redirect('pg_bookings_pending')

    booking.status = Booking.APPROVED
    booking.start_date = timezone.now().date()
    update_fields = ['status', 'start_date']
    if collect_advance:
        booking.advance_paid = (booking.advance_paid or Decimal('0')) + advance_amount_value
        update_fields.append('advance_paid')
    booking.save(update_fields=update_fields)
    
    # Set share status based on joining_date
    today = timezone.now().date()
    if booking.joining_date and booking.joining_date > today:
        # Future joining date - mark as RESERVED
        share.status = RoomShareStatus.RESERVED
    else:
        # Current or past joining date - mark as OCCUPIED
        share.status = RoomShareStatus.OCCUPIED
    share.save(update_fields=['status'])
    
    log(request.user, 'booking_approved', 'Booking', booking.id, f"Approved for room {booking.room.room_no} bed {booking.share_no}")
    
    # Auto-create ResidentApplication for advance/future bookings if one doesn't exist
    # This ensures all approved bookings appear in the applications page for tracking
    if not hasattr(booking, 'application') or booking.application is None:
        try:
            user_obj = booking.user
            name = (getattr(user_obj, 'first_name', '') or '').strip()
            if not name:
                name = (getattr(user_obj, 'get_full_name', lambda: '')() or user_obj.email or '').strip()
            phone = ''
            try:
                phone = getattr(getattr(user_obj, 'profile', None), 'phone', '') or ''
            except Exception:
                phone = ''
            email_addr = getattr(user_obj, 'email', '') or ''

            app = ResidentApplication.objects.create(
                user=user_obj,
                booking=booking,
                pg=pg,
                room=booking.room,
                status=ResidentApplication.PENDING,
                name=name,
                phone=phone,
                email=email_addr,
                date_of_admission=booking.joining_date,
            )
            # Refresh the booking to get the application relation
            booking.refresh_from_db()
            log(request.user, 'application_autocreated', 'ResidentApplication', app.id, 
                f"Auto-created pending application for approved booking {booking.id} (advance/future booking)")
        except Exception as exc:
            # Log but don't fail the approval
            _logger.warning(f"Could not auto-create application for booking {booking.id}: {exc}")
    
    if confirm_application and hasattr(booking, 'application') and booking.application:
        if booking.application.status != ResidentApplication.CONFIRMED:
            booking.application.status = ResidentApplication.CONFIRMED
            booking.application.save(update_fields=['status'])
            log(request.user, 'application_confirmed_during_booking', 'ResidentApplication', booking.application.id, f"Application confirmed while approving booking {booking.id}")

    # Handle referral tracking
    referral_created = False
    referral_credit_amount = Decimal('0')
    if request.method == 'POST':
        is_referred = request.POST.get('is_referred') == 'on'
        if is_referred:
            referred_by_booking_id = request.POST.get('referred_by_booking_id')
            referral_month_str = request.POST.get('referral_month')
            referral_amount_str = request.POST.get('referral_amount', '500').strip()
            
            if referred_by_booking_id and referral_month_str:
                try:
                    # Get the referrer booking
                    referrer_booking = Booking.objects.select_related('user', 'room').get(
                        id=int(referred_by_booking_id),
                        status=Booking.APPROVED,
                        room__pg=pg
                    )
                    
                    # Parse referral amount
                    try:
                        referral_credit_amount = Decimal(referral_amount_str) if referral_amount_str else Decimal('500')
                    except (InvalidOperation, ValueError):
                        referral_credit_amount = Decimal('500')
                    
                    # Parse referral month
                    referral_month = parse_date(referral_month_str)
                    if not referral_month:
                        messages.warning(request, "Invalid referral month format.")
                    else:
                        # Ensure referral month is in the future
                        today = timezone.now().date()
                        if referral_month <= today.replace(day=1):
                            messages.warning(request, "Referral month must be in the future.")
                        else:
                            # Update application with referrer
                            if hasattr(booking, 'application') and booking.application:
                                booking.application.referred_by_booking = referrer_booking
                                booking.application.save(update_fields=['referred_by_booking'])
                                log(request.user, 'referral_set', 'ResidentApplication', booking.application.id, 
                                    f"Set referrer to booking {referrer_booking.id} ({referrer_booking.user.get_full_name() or referrer_booking.user.email})")
                            
                            # Create ReferralCredit record only if there's a resident application linked
                            if hasattr(booking, 'application') and booking.application:
                                try:
                                    referral_credit = ReferralCredit.objects.create(
                                        pg=pg,
                                        referrer_user=referrer_booking.user,
                                        referrer_booking=referrer_booking,
                                        referred_user=booking.user,
                                        referred_booking=booking,
                                        application=booking.application,
                                        amount=referral_credit_amount,
                                        scheduled_month=referral_month,
                                        notes=f"Referral credit for {booking.user.get_full_name() or booking.user.email}"
                                    )
                                    referral_created = True
                                    log(request.user, 'referral_credit_created', 'ReferralCredit', referral_credit.id,
                                        f"Created referral credit of Rs.{referral_credit_amount} for {referrer_booking.user.email} scheduled for {referral_month}")
                                except Exception as exc:
                                    # Avoid leaving transaction broken by preventing IntegrityError where possible
                                    messages.warning(request, f"Referral information saved but credit record failed: {exc}")
                            else:
                                # Auto-create a minimal ResidentApplication so referral can be linked.
                                try:
                                    # Build minimal fields from user and booking
                                    user_obj = booking.user
                                    name = (getattr(user_obj, 'first_name', '') or '')
                                    if not name:
                                        # fallback to full name helper or email
                                        name = (getattr(user_obj, 'get_full_name', lambda: '')() or user_obj.email or '')
                                    phone = ''
                                    try:
                                        phone = getattr(getattr(user_obj, 'profile', None), 'phone', '') or ''
                                    except Exception:
                                        phone = ''
                                    email_addr = getattr(user_obj, 'email', '') or ''

                                    app = ResidentApplication.objects.create(
                                        user=user_obj,
                                        booking=booking,
                                        pg=pg,
                                        room=booking.room,
                                        status=ResidentApplication.PENDING,  # auto-created applications set to PENDING
                                        name=name,
                                        phone=phone,
                                        email=email_addr,
                                    )
                                    # Ensure booking.application relation is available
                                    booking.application = app
                                    log(request.user, 'application_autocreated', 'ResidentApplication', app.id, f"Auto-created minimal application for booking {booking.id} during referral setup")

                                    # Now link referrer and create ReferralCredit
                                    booking.application.referred_by_booking = referrer_booking
                                    booking.application.save(update_fields=['referred_by_booking'])
                                    try:
                                        referral_credit = ReferralCredit.objects.create(
                                            pg=pg,
                                            referrer_user=referrer_booking.user,
                                            referrer_booking=referrer_booking,
                                            referred_user=booking.user,
                                            referred_booking=booking,
                                            application=booking.application,
                                            amount=referral_credit_amount,
                                            scheduled_month=referral_month,
                                            notes=f"Referral credit for {booking.user.get_full_name() or booking.user.email}"
                                        )
                                        referral_created = True
                                        log(request.user, 'referral_credit_created', 'ReferralCredit', referral_credit.id,
                                            f"Created referral credit of Rs.{referral_credit_amount} for {referrer_booking.user.email} scheduled for {referral_month}")
                                    except Exception as exc:
                                        messages.warning(request, f"Referral information saved but credit record failed: {exc}")
                                except Exception as exc:
                                    messages.warning(request, f"Could not auto-create resident application for referral: {exc}")
                
                except Booking.DoesNotExist:
                    messages.warning(request, "Selected referrer booking not found or invalid.")
                except Exception as exc:
                    messages.warning(request, f"Error processing referral: {exc}")

    advance_payment_success = False
    payment_obj = None
    if collect_advance:
        from finance.models import Payment
        try:
            # Build notes: use custom notes if provided, otherwise use default
            payment_notes = advance_notes if advance_notes else f"Advance collected during booking approval for Room {booking.room.room_no} Bed {booking.share_no}"
            
            payment_obj = Payment.objects.create(
                user=booking.user,
                pg=pg,
                booking=booking,
                amount=advance_amount_value,
                date=timezone.now().date(),
                status='success',
                mode=advance_mode,
                type='advance',
                notes=payment_notes,
                from_date=booking.joining_date or booking.start_date,
                to_date=booking.joining_date or booking.start_date,
            )
            advance_payment_success = True
            log(request.user, 'advance_recorded', 'Payment', payment_obj.id, f"Advance of Rs.{advance_amount_value} recorded for booking {booking.id}")
            
            # Send email receipt
            try:
                from finance.signals import _build_receipt_context
                receipt_ctx = _build_receipt_context(payment_obj)
                
                # Build email content
                email_subject = f"Payment Receipt - {pg.name}"
                email_body = f"""Dear {booking.user.get_full_name() or booking.user.email},

Your advance payment has been successfully recorded.

Payment Details:
---------------
Receipt No: {receipt_ctx.get('receipt_no', 'N/A')}
Amount: Rs.{advance_amount_value}
Payment Mode: {advance_mode.upper()}
Date: {payment_obj.date.strftime('%d %B %Y')}
PG: {pg.name}

Room Details:
------------
Room: {booking.room.room_no}
Bed: {booking.share_no}
Joining Date: {booking.joining_date or booking.start_date}

{f'Notes: {payment_notes}' if payment_notes else ''}

Thank you for your payment!

---
{pg.name}
{pg.address or ''}
{pg.phone or ''}
"""
                
                send_mail(
                    subject=email_subject,
                    message=email_body,
                    from_email=None,
                    recipient_list=[booking.user.email],
                    fail_silently=True,
                )
            except Exception as e:
                # Log but don't fail - email is optional
                messages.warning(request, f"Payment recorded but email receipt could not be sent: {str(e)}")
                
        except Exception as exc:
            messages.warning(request, f"Advance amount saved but payment record could not be created: {exc}")
    # Notify user
    Notification.objects.create(user=booking.user, title="Booking approved", message=f"Your booking for {booking.room} bed {booking.share_no} was approved.")
    send_push_to_user(
        booking.user,
        title="Booking Approved",
        body=f"Your booking for Room {booking.room.room_no}, Bed {booking.share_no} was approved.",
        url=reverse('application_fill', args=[booking.id]),
        extra_data={'type': 'booking_approved', 'booking_id': booking.id},
    )
    try:
        link = request.build_absolute_uri(reverse('application_fill', args=[booking.id]))
        send_mail(
            subject="PG-MS: Booking Approved",
            message=f"Your booking for {booking.room} bed {booking.share_no} was approved.\nPlease complete your resident application here: {link}",
            from_email=None,
            recipient_list=[booking.user.email],
            fail_silently=True,
        )
    except Exception:
        pass
    success_parts = ["Booking approved and user notified."]
    if confirm_application and hasattr(booking, 'application') and booking.application.status == ResidentApplication.CONFIRMED:
        success_parts.append("Resident application marked as confirmed.")
    if collect_advance:
        if advance_payment_success:
            success_parts.append(f"Advance of Rs.{advance_amount_value:.2f} recorded.")
        else:
            success_parts.append("Advance amount noted.")
    if referral_created:
        success_parts.append(f"Referral credit of Rs.{referral_credit_amount:.2f} created and scheduled.")
    messages.success(request, ' '.join(success_parts))
    # AJAX response
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.POST.get('ajax'):
        return JsonResponse({
            'ok': True,
            'action': 'booking_approve',
            'booking_id': booking.id,
            'share_id': share.id,
            'share_status': share.status,
            'message': 'Booking approved.'
        })
    return redirect('pg_bookings_pending')


@login_required
@transaction.atomic
def booking_reject(request, booking_id):
    if not _require_pg_admin(request.user):
        messages.error(request, "PG Admin access required.")
        return redirect('dashboard')
    booking = get_object_or_404(Booking, pk=booking_id, status=Booking.PENDING)
    if not _admin_pgs(request.user).filter(id=(getattr(booking, 'pg_id', None) or getattr(getattr(booking, 'room', None), 'pg_id', None))).exists():
        messages.error(request, "PG Admin access required for this PG.")
        return redirect('dashboard')
    
    # Handle day-wise bookings differently (no share to update if not assigned yet)
    if booking.booking_type == Booking.DAYWISE:
        booking.status = Booking.REJECTED
        booking.save(update_fields=['status'])
        
        # Update application status if exists
        if hasattr(booking, 'application'):
            booking.application.status = ResidentApplication.REJECTED
            booking.application.save(update_fields=['status'])
        
        log(request.user, 'daywise_booking_rejected', 'Booking', booking.id, "Rejected day-wise booking")
        Notification.objects.create(
            user=booking.user,
            title="Day-wise Booking Rejected",
            message="Your day-wise booking request has been rejected."
        )
        send_push_to_user(
            booking.user,
            title="Day-wise Booking Rejected",
            body="Your day-wise booking request has been rejected.",
            url=reverse('booking_detail', args=[booking.id]),
            extra_data={'type': 'booking_rejected', 'booking_type': 'daywise', 'booking_id': booking.id},
        )
        messages.info(request, "Day-wise booking rejected.")
        
        if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.POST.get('ajax'):
            return JsonResponse({
                'ok': True,
                'action': 'booking_reject',
                'booking_id': booking.id,
                'message': 'Day-wise booking rejected.'
            })
        return redirect('pg_bookings_pending')
    
    # Regular booking rejection
    share = get_object_or_404(RoomShareStatus, room=booking.room, share_no=booking.share_no)
    booking.status = Booking.REJECTED
    booking.save(update_fields=['status'])
    # On rejection, revert bed to appropriate availability state.
    # Rule: if share.vacant_from is in the future -> keep VACANT_FROM (future scheduled vacancy)
    # else (date is past or today OR no date) -> mark as VACANT now.
    today = timezone.now().date()
    update_fields = ['status']
    if share.vacant_from and share.vacant_from > today:
        share.status = RoomShareStatus.VACANT_FROM
        # keep vacant_from date
    else:
        share.status = RoomShareStatus.VACANT
        if share.vacant_from:
            # Clear stale date once it's already vacant
            share.vacant_from = None
            update_fields.append('vacant_from')
    share.save(update_fields=update_fields)
    log(request.user, 'booking_rejected', 'Booking', booking.id, f"Rejected for room {booking.room.room_no} bed {booking.share_no}")
    Notification.objects.create(user=booking.user, title="Booking rejected", message=f"Your booking for {booking.room} bed {booking.share_no} was rejected.")
    send_push_to_user(
        booking.user,
        title="Booking Rejected",
        body=f"Your booking for Room {booking.room.room_no}, Bed {booking.share_no} was rejected.",
        url=reverse('booking_detail', args=[booking.id]),
        extra_data={'type': 'booking_rejected', 'booking_id': booking.id},
    )
    try:
        send_mail(
            subject="PG-MS: Booking Rejected",
            message=f"Your booking for {booking.room} bed {booking.share_no} was rejected.",
            from_email=None,
            recipient_list=[booking.user.email],
            fail_silently=True,
        )
    except Exception:
        pass
    messages.info(request, "Booking rejected.")
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.POST.get('ajax'):
        return JsonResponse({
            'ok': True,
            'action': 'booking_reject',
            'booking_id': booking.id,
            'share_id': share.id,
            'share_status': share.status,
            'vacant_from': share.vacant_from.isoformat() if getattr(share, 'vacant_from', None) else None,
            'message': 'Booking rejected.'
        })
    return redirect('pg_bookings_pending')


# ============================================================================
# OLD LEAVING REQUESTS - REPLACED BY ENHANCED VERSION (line ~3124)
# This old function is kept for backward compatibility with old leaving_requests.html template
# The new enhanced version uses leaving_requests_enhanced.html
# ============================================================================


@login_required
def application_email_send(request, booking_id):
    if not _require_pg_admin(request.user):
        messages.error(request, "PG Admin access required.")
        return redirect('dashboard')
    booking = get_object_or_404(Booking, pk=booking_id)
    if not _admin_pgs(request.user).filter(id=(getattr(booking, 'pg_id', None) or getattr(getattr(booking, 'room', None), 'pg_id', None))).exists():
        messages.error(request, "PG Admin access required for this PG.")
        return redirect('dashboard')
    try:
        from django.urls import reverse
        link = request.build_absolute_uri(reverse('application_fill', args=[booking.id]))
        sent = send_mail(
            subject="PG-MS: Complete Your Resident Application",
            message=f"Please complete your resident application here: {link}",
            from_email=None,
            recipient_list=[booking.user.email],
            fail_silently=False,
        )
        if sent:
            messages.success(request, f"Application link sent to {booking.user.email}.")
        else:
            messages.error(request, "Email was not sent. Please verify email settings.")
    except Exception as e:
        messages.error(request, f"Could not send email: {e}")
    # Redirect back to where the action was invoked from when possible
    src = request.GET.get('from')
    if src == 'applications':
        return redirect('pg_resident_applications')
    if src == 'tenants':
        return redirect('pg_tenants')
    return redirect('pg_bookings_pending')


@login_required
@transaction.atomic
def leaving_confirm(request, booking_id):
    if not _require_pg_admin(request.user):
        messages.error(request, "PG Admin access required.")
        return redirect('dashboard')
    booking = get_object_or_404(Booking, pk=booking_id, status=Booking.APPROVED)
    if not _admin_pgs(request.user).filter(id=(getattr(booking, 'pg_id', None) or getattr(getattr(booking, 'room', None), 'pg_id', None))).exists():
        messages.error(request, "PG Admin access required for this PG.")
        return redirect('dashboard')
    share = get_object_or_404(RoomShareStatus, room=booking.room, share_no=booking.share_no)
    today = timezone.now().date()
    updated_fields = []
    if not booking.leaving_confirmed_date:
        booking.leaving_confirmed_date = today
        updated_fields.append('leaving_confirmed_date')
    if booking.leaving_date and booking.leaving_date <= today:
        share.status = RoomShareStatus.VACANT
        share.vacant_from = None
    else:
        share.status = RoomShareStatus.VACANT_FROM
        share.vacant_from = booking.leaving_date
    share.save(update_fields=['status', 'vacant_from'])
    if updated_fields:
        booking.save(update_fields=updated_fields)
    try:
        if hasattr(booking.user, 'profile'):
            if getattr(booking.user.profile, 'is_pg_user', True):
                booking.user.profile.is_pg_user = False
                booking.user.profile.save(update_fields=['is_pg_user'])
    except Exception:
        pass
    log(request.user, 'leaving_confirmed', 'Booking', booking.id, f"Leaving confirmed; booking closed for room {booking.room.room_no} bed {booking.share_no}")

    try:
        admin_path, admin_payload = _leave_admin_path_and_payload(booking)
        admin_users = _pg_admin_users_for_pg(booking.room.pg_id)
        tenant_name = booking.user.get_full_name() or booking.user.email
        user_detail_path = f"{reverse('booking_detail', args=[booking.id])}?{urlencode({'pg': booking.room.pg_id})}"

        Notification.objects.create(
            user=booking.user,
            title="Leave Request Confirmed",
            message=f"Your leave request for Room {booking.room.room_no}, Bed {booking.share_no} has been confirmed.",
        )
        send_push_to_user(
            booking.user,
            title="Leave Request Confirmed",
            body=f"Leave for Room {booking.room.room_no}, Bed {booking.share_no} is confirmed.",
            url=user_detail_path,
            extra_data={'type': 'leave_confirmed', 'booking_id': booking.id, 'pg_id': booking.room.pg_id},
        )

        for admin_user in admin_users:
            Notification.objects.create(
                user=admin_user,
                title="Leave Request Confirmed",
                message=f"{tenant_name}'s leave request for Room {booking.room.room_no}, Bed {booking.share_no} has been confirmed.",
            )
        send_push_to_users(
            admin_users,
            title="Leave Request Confirmed",
            body=f"{tenant_name} leave confirmed for Room {booking.room.room_no}, Bed {booking.share_no}.",
            url=admin_path,
            extra_data={**admin_payload, 'type': 'leave_confirmed', 'source': 'legacy_leave_confirm'},
        )
    except Exception:
        _logger.exception('Legacy leave confirm notification dispatch failed for booking %s', booking.id)

    if share.status == RoomShareStatus.VACANT:
        messages.success(request, f"Leaving confirmed and bed freed for room {booking.room.room_no}.")
    else:
        messages.success(request, f"Leaving confirmed for room {booking.room.room_no}. Bed will free on {booking.leaving_date}.")
    return redirect('pg_leaving_requests')


@login_required
@transaction.atomic
def leaving_delete(request, booking_id):
    if request.method != 'POST':
        messages.error(request, "Invalid request method.")
        return redirect('pg_leaving_requests')
    if not _require_pg_admin(request.user):
        messages.error(request, "PG Admin access required.")
        return redirect('dashboard')
    # Allow delete for both APPROVED and COMPLETED bookings
    booking = get_object_or_404(Booking, pk=booking_id, status__in=[Booking.APPROVED, Booking.COMPLETED])
    if not _admin_pgs(request.user).filter(id=(getattr(booking, 'pg_id', None) or getattr(getattr(booking, 'room', None), 'pg_id', None))).exists():
        messages.error(request, "PG Admin access required for this PG.")
        return redirect('dashboard')
    if not booking.leaving_confirmed_date:
        messages.error(request, "Confirm the leaving request before deleting resident data.")
        return redirect('pg_leaving_requests')
    summary = _cleanup_booking_after_leave(booking, actor=request.user, origin='manual_button')
    if summary.get('failed_files'):
        messages.warning(
            request,
            (
                f"Resident data deleted but {summary['failed_files']} file(s) could not be removed from Drive. "
                "Please review manually."
            ),
        )
    else:
        extra = ''
        if summary.get('deleted_files'):
            extra = f" {summary['deleted_files']} file(s) deleted from Drive."
        messages.success(request, f"Resident booking and application removed.{extra}")
    return redirect('pg_leaving_requests')


@login_required
@transaction.atomic
def leaving_reject(request, booking_id):
    if request.method != 'POST':
        messages.error(request, "Invalid request method.")
        return redirect('pg_leaving_requests')
    if not _require_pg_admin(request.user):
        messages.error(request, "PG Admin access required.")
        return redirect('dashboard')
    booking = get_object_or_404(Booking, pk=booking_id, status=Booking.APPROVED)
    if not _admin_pgs(request.user).filter(id=(getattr(booking, 'pg_id', None) or getattr(getattr(booking, 'room', None), 'pg_id', None))).exists():
        messages.error(request, "PG Admin access required for this PG.")
        return redirect('dashboard')
    if booking.leaving_confirmed_date:
        messages.error(request, "Leaving request already confirmed. Use delete to remove resident data if needed.")
        return redirect('pg_leaving_requests')
    old_leaving_date = booking.leaving_date
    booking.leaving_date = None
    booking.save(update_fields=['leaving_date'])
    try:
        share = RoomShareStatus.objects.get(room=booking.room, share_no=booking.share_no)
        share.status = RoomShareStatus.OCCUPIED
        share.vacant_from = None
        share.save(update_fields=['status', 'vacant_from'])
    except RoomShareStatus.DoesNotExist:
        pass
    log(request.user, 'booking_leave_rejected', 'Booking', booking.id, f"Leave request rejected for room {booking.room.room_no} bed {booking.share_no}")

    try:
        admin_path, admin_payload = _leave_admin_path_and_payload(booking)
        admin_users = _pg_admin_users_for_pg(booking.room.pg_id)
        tenant_name = booking.user.get_full_name() or booking.user.email
        user_detail_path = f"{reverse('booking_detail', args=[booking.id])}?{urlencode({'pg': booking.room.pg_id})}"

        Notification.objects.create(
            user=booking.user,
            title="Leave Request Rejected",
            message=f"Your leave request for Room {booking.room.room_no}, Bed {booking.share_no} was rejected.",
        )
        send_push_to_user(
            booking.user,
            title="Leave Request Rejected",
            body=f"Leave request for Room {booking.room.room_no}, Bed {booking.share_no} was rejected.",
            url=user_detail_path,
            extra_data={'type': 'leave_rejected', 'booking_id': booking.id, 'pg_id': booking.room.pg_id},
        )

        for admin_user in admin_users:
            Notification.objects.create(
                user=admin_user,
                title="Leave Request Rejected",
                message=f"{tenant_name}'s leave request for Room {booking.room.room_no}, Bed {booking.share_no} was rejected.",
            )
        send_push_to_users(
            admin_users,
            title="Leave Request Rejected",
            body=f"{tenant_name} leave request rejected for Room {booking.room.room_no}, Bed {booking.share_no}.",
            url=admin_path,
            extra_data={**admin_payload, 'type': 'leave_rejected', 'source': 'legacy_leave_reject'},
        )
    except Exception:
        _logger.exception('Legacy leave reject notification dispatch failed for booking %s', booking.id)

    messages.success(request, "Leave request rejected and cleared.")
    return redirect('pg_leaving_requests')
from django.shortcuts import render

# Create your views here.


@login_required
def resident_applications(request):
    # Show all approved bookings for this PG with application status/actions
    if not _require_pg_admin(request.user):
        messages.error(request, "PG Admin access required.")
        return redirect('dashboard')
    pg = _active_pg(request)
    bookings = []
    referral_options = []
    if pg:
        today = timezone.now().date()
        bookings = (
            Booking.objects
            .filter(
                room__pg=pg,
                status=Booking.APPROVED,
            )
            .filter(Q(leaving_date__isnull=True) | Q(leaving_date__gt=today))
            .select_related('user', 'room', 'application', 'application__referral_credit')
            .order_by('room__room_no', 'share_no')
        )
        referral_options = (
            Booking.objects
            .filter(
                room__pg=pg,
                status=Booking.APPROVED,
            )
            .filter(Q(leaving_date__isnull=True) | Q(leaving_date__gt=today))
            .select_related('user', 'room', 'user__profile')
            .order_by('user__first_name', 'user__last_name', 'room__room_no', 'share_no')
            .distinct()
        )
    if bookings:
        # Submitted: applications with submitted/resubmitted/refill_requested statuses
        submitted_count = bookings.filter(
            application__status__in=(
                ResidentApplication.SUBMITTED,
                ResidentApplication.RESUBMITTED,
                ResidentApplication.REFILL_REQUESTED,
            )
        ).count()
        # Pending: either no application exists or application exists with 'pending' status
        pending_count = bookings.filter(
            Q(application__isnull=True) | Q(application__status=ResidentApplication.PENDING)
        ).count()
    else:
        submitted_count = 0
        pending_count = 0
    
    # Check if PG admin has edit applications permission
    can_edit_applications = False
    try:
        from pgadmin.models import PGAdminPermission
        pg_admin_record = PGAdmin.objects.filter(user=request.user, pg=pg).first() if pg else None
        if pg_admin_record:
            perm = PGAdminPermission.objects.filter(pg_admin=pg_admin_record).first()
            if perm:
                can_edit_applications = perm.can_edit_applications
        # Website admins and superusers always have permission
        if getattr(request.user, 'is_superuser', False) or (hasattr(request.user, 'profile') and getattr(request.user.profile, 'is_website_admin', False)):
            can_edit_applications = True
    except Exception:
        pass
    
    return render(
        request,
        'pgadmin/resident_applications.html',
        {
            "pg": pg,
            "bookings": bookings,
            "pgs": list(_admin_pgs(request.user)),
            "submitted_count": submitted_count,
            "pending_count": pending_count,
            "referral_options": referral_options,
            "can_edit_applications": can_edit_applications,
        },
    )


@login_required
@transaction.atomic
def application_confirm(request, app_id):
    if not _require_pg_admin(request.user):
        messages.error(request, "PG Admin access required.")
        return redirect('dashboard')
    app = get_object_or_404(ResidentApplication, pk=app_id)
    if not _admin_pgs(request.user).filter(id=getattr(app, 'pg_id', None)).exists():
        messages.error(request, "PG Admin access required for this PG.")
        return redirect('dashboard')
    # Set status to confirmed
    app.status = ResidentApplication.CONFIRMED
    app.save(update_fields=['status'])
    # History
    from bookings.models import ApplicationStatusHistory
    ApplicationStatusHistory.objects.create(application=app, status=app.status, comment='Confirmed by PG admin', by_user=request.user)
    # Ensure referral credit record exists when applicable
    try:
        if app.referred_by_booking_id and not getattr(app, 'referral_credit', None):
            ref_booking = app.referred_by_booking
            if ref_booking and ref_booking.status == Booking.APPROVED:
                amount = Decimal(str(app.pg.referral_amount or 0))
                if amount > 0:
                    anchor = getattr(app.booking, 'payment_date', None) or getattr(app.booking, 'joining_date', None)
                    if anchor:
                        scheduled_month = anchor.replace(day=1)
                    else:
                        scheduled_month = timezone.now().date().replace(day=1)
                    ReferralCredit.objects.create(
                        pg=app.pg,
                        referrer_user=ref_booking.user,
                        referrer_booking=ref_booking,
                        referred_user=app.user,
                        referred_booking=app.booking,
                        application=app,
                        amount=amount,
                        scheduled_month=scheduled_month,
                        notes='Auto-created on confirmation',
                    )
    except Exception as exc:
        messages.warning(request, f"Referral credit could not be created automatically: {exc}")
    # Notify user
    try:
        from django.urls import reverse
        link = request.build_absolute_uri(reverse('my_application'))
        send_mail(
            subject="PG-MS: Application Confirmed",
            message=f"Your resident application has been confirmed. You can view it here: {link}\nFurther edits are disabled unless requested by admin.",
            from_email=None,
            recipient_list=[app.user.email],
            fail_silently=True,
        )
        send_push_to_user(
            app.user,
            title="Application Confirmed",
            body="Your resident application has been confirmed.",
            url=reverse('my_application'),
            extra_data={'type': 'application_confirmed', 'application_id': app.id},
        )
    except Exception:
        pass
    messages.success(request, "Application confirmed.")
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.POST.get('ajax'):
        return JsonResponse({'ok': True, 'action': 'application_confirm', 'application_id': app.id, 'status': app.status})
    return redirect('pg_resident_applications')


@login_required
@transaction.atomic
def application_reject(request, app_id):
    if not _require_pg_admin(request.user):
        messages.error(request, "PG Admin access required.")
        return redirect('dashboard')
    app = get_object_or_404(ResidentApplication, pk=app_id)
    if not _admin_pgs(request.user).filter(id=getattr(app, 'pg_id', None)).exists():
        messages.error(request, "PG Admin access required for this PG.")
        return redirect('dashboard')
    reason = (request.POST.get('reason') or '').strip()
    if request.method != 'POST' or not reason:
        messages.error(request, "Rejection reason is required.")
        return redirect('pg_resident_applications')
    app.status = ResidentApplication.REJECTED
    app.save(update_fields=['status'])
    from bookings.models import ApplicationStatusHistory
    ApplicationStatusHistory.objects.create(application=app, status=app.status, comment=reason, by_user=request.user)
    # Email user with reason and link to edit
    try:
        from django.urls import reverse
        link = request.build_absolute_uri(reverse('application_fill', args=[app.booking_id]))
        send_mail(
            subject="PG-MS: Application Rejected",
            message=f"Your resident application was rejected for the following reason:\n\n{reason}\n\nPlease re-fill the form here: {link}",
            from_email=None,
            recipient_list=[app.user.email],
            fail_silently=False,
        )
    except Exception as e:
        messages.error(request, f"Failed to send rejection email: {e}")
    messages.info(request, "Application rejected and user notified.")
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.POST.get('ajax'):
        return JsonResponse({'ok': True, 'action': 'application_reject', 'application_id': app.id, 'status': app.status})
    return redirect('pg_resident_applications')


@login_required
@transaction.atomic
def application_refill_request(request, app_id):
    if not _require_pg_admin(request.user):
        messages.error(request, "PG Admin access required.")
        return redirect('dashboard')
    app = get_object_or_404(ResidentApplication, pk=app_id)
    if not _admin_pgs(request.user).filter(id=getattr(app, 'pg_id', None)).exists():
        messages.error(request, "PG Admin access required for this PG.")
        return redirect('dashboard')
    comment = (request.POST.get('comment') or '').strip()
    app.status = ResidentApplication.REFILL_REQUESTED
    app.save(update_fields=['status'])
    from bookings.models import ApplicationStatusHistory
    ApplicationStatusHistory.objects.create(application=app, status=app.status, comment=comment, by_user=request.user)
    # Email user to edit (this is commonly triggered after confirmation to request changes)
    try:
        from django.urls import reverse
        link = request.build_absolute_uri(reverse('application_fill', args=[app.booking_id]))
        send_mail(
            subject="PG-MS: Application Update Requested",
            message=f"Updates are requested on your resident application. Please edit here: {link}\n\nDetails: {comment or 'Please update your details as discussed.'}",
            from_email=None,
            recipient_list=[app.user.email],
            fail_silently=False,
        )
        send_push_to_user(
            app.user,
            title="Application Update Requested",
            body="PG admin requested updates on your application.",
            url=reverse('application_fill', args=[app.booking_id]),
            extra_data={'type': 'application_refill_requested', 'application_id': app.id},
        )
    except Exception as e:
        messages.error(request, f"Failed to send update request email: {e}")
    messages.success(request, "Re-Fill request sent to user.")
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.POST.get('ajax'):
        return JsonResponse({'ok': True, 'action': 'application_refill', 'application_id': app.id, 'status': app.status})
    return redirect('pg_resident_applications')


@login_required
@transaction.atomic
def admin_application_edit(request, app_id):
    """Allow PG admin with permission to edit tenant applications."""
    if not _require_pg_admin(request.user):
        messages.error(request, "PG Admin access required.")
        return redirect('dashboard')
    
    app = get_object_or_404(
        ResidentApplication.objects.select_related('pg', 'booking', 'user', 'room'),
        pk=app_id,
    )
    
    if not _admin_pgs(request.user).filter(id=getattr(app, 'pg_id', None)).exists():
        messages.error(request, "PG Admin access required for this PG.")
        return redirect('dashboard')
    
    # Check edit applications permission
    has_permission = False
    if getattr(request.user, 'is_superuser', False) or (hasattr(request.user, 'profile') and getattr(request.user.profile, 'is_website_admin', False)):
        has_permission = True
    else:
        try:
            from pgadmin.models import PGAdminPermission
            pg_admin_record = PGAdmin.objects.filter(user=request.user, pg=app.pg).first()
            if pg_admin_record:
                perm = PGAdminPermission.objects.filter(pg_admin=pg_admin_record).first()
                if perm and perm.can_edit_applications:
                    has_permission = True
        except Exception:
            pass
    
    if not has_permission:
        messages.error(request, "You don't have permission to edit applications.")
        return redirect('pg_resident_applications')
    
    booking = app.booking
    from bookings.application_forms import ResidentApplicationForm
    
    if request.method == 'POST':
        form = ResidentApplicationForm(request.POST, request.FILES, instance=app)
        
        if form.is_valid():
            # Track what fields changed for logging
            old_values = {}
            for field in form.changed_data:
                old_values[field] = getattr(app, field, None)
            
            inst = form.save(commit=False)

            from core.drive import drive_upload, drive_delete

            def _replace_drive_file(uploaded_file, old_url, name_prefix, warn_prefix):
                """Upload replacement file; delete old file only after successful upload."""
                folder_id = getattr(app.pg, 'drive_folder_id', None) or 'root'
                original_name = (getattr(uploaded_file, 'name', '') or '').strip()
                ext = ''
                if '.' in original_name:
                    ext = '.' + original_name.rsplit('.', 1)[-1].lower()
                    if len(ext) > 10:
                        ext = ''

                filename = f"{name_prefix}_{app.id}_{timezone.now().strftime('%Y%m%d%H%M%S%f')}{ext}"

                try:
                    uploaded = drive_upload(uploaded_file, filename, folder_id)
                    if not uploaded:
                        messages.warning(request, f"{warn_prefix} upload failed: Google Drive upload service unavailable.")
                        return None

                    _new_file_id, preview_url = uploaded

                    # Delete old file best-effort only after replacement succeeded.
                    if old_url:
                        try:
                            drive_delete(old_url)
                        except Exception:
                            pass

                    return preview_url
                except Exception as e:
                    messages.warning(request, f"{warn_prefix} upload failed: {e}")
                    return None
            
            # Handle selfie upload
            if 'selfie' in request.FILES:
                selfie_file = request.FILES['selfie']
                selfie_preview = _replace_drive_file(selfie_file, app.selfie_url, 'selfie', 'Selfie')
                if selfie_preview:
                    inst.selfie_url = selfie_preview
            
            # Handle aadhaar file upload
            if 'aadhaar_pdf' in request.FILES:
                aadhaar_file = request.FILES['aadhaar_pdf']
                aadhaar_preview = _replace_drive_file(aadhaar_file, app.aadhaar_file_url, 'aadhaar_front', 'Document')
                if aadhaar_preview:
                    inst.aadhaar_file_url = aadhaar_preview
            
            # Handle aadhaar file 2 upload
            if 'aadhaar_pdf_2' in request.FILES:
                aadhaar_file_2 = request.FILES['aadhaar_pdf_2']
                aadhaar_2_preview = _replace_drive_file(aadhaar_file_2, app.aadhaar_file_url_2, 'aadhaar_back', 'Document 2')
                if aadhaar_2_preview:
                    inst.aadhaar_file_url_2 = aadhaar_2_preview
            
            inst.save()
            
            # Log the edit in AuditLog
            from core.models import AuditLog
            changed_fields = form.changed_data
            if changed_fields:
                change_summary = ", ".join(changed_fields)
                AuditLog.objects.create(
                    actor=request.user,
                    action='application_edited_by_admin',
                    target_type='ResidentApplication',
                    target_id=app.id,
                    message=f"Application edited by PG Admin. Changed fields: {change_summary}",
                    meta={
                        'changed_fields': changed_fields,
                        'booking_id': booking.id,
                        'tenant_email': app.user.email if app.user else None,
                        'admin_email': request.user.email,
                    }
                )
            
            # Also log in ApplicationStatusHistory
            from bookings.models import ApplicationStatusHistory
            ApplicationStatusHistory.objects.create(
                application=app,
                status=app.status,
                comment=f"Application edited by PG Admin ({request.user.email}). Changed: {', '.join(changed_fields) if changed_fields else 'No changes'}",
                by_user=request.user
            )
            
            messages.success(request, "Application updated successfully.")
            return redirect('pg_resident_applications')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = ResidentApplicationForm(instance=app)
    
    return render(request, 'pgadmin/admin_application_edit.html', {
        'form': form,
        'app': app,
        'booking': booking,
        'pg': app.pg,
    })


@login_required
@transaction.atomic
def application_update_referral(request, app_id):
    if not _require_pg_admin(request.user):
        messages.error(request, "PG Admin access required.")
        return redirect('dashboard')
    app = get_object_or_404(
        ResidentApplication.objects.select_related('pg', 'booking', 'user', 'referral_credit'),
        pk=app_id,
    )
    if not _admin_pgs(request.user).filter(id=getattr(app, 'pg_id', None)).exists():
        messages.error(request, "PG Admin access required for this PG.")
        return redirect('dashboard')
    if request.method != 'POST':
        messages.error(request, "Invalid request method.")
        return redirect('pg_resident_applications')

    redirect_target = redirect('pg_resident_applications')
    action = (request.POST.get('action') or '').strip().lower()
    credit = getattr(app, 'referral_credit', None)

    # Clear referral
    if action == 'clear' or not (request.POST.get('referred_by_booking') or '').strip():
        if credit and credit.redeemed_on:
            messages.error(request, "Referral credit already redeemed; cannot clear it.")
            return redirect_target
        app.referred_by_booking = None
        app.save(update_fields=['referred_by_booking'])
        if credit:
            credit.delete()
        messages.success(request, "Referral cleared for this application.")
        return redirect_target

    # Validate selected booking
    try:
        ref_booking_id = int(request.POST.get('referred_by_booking'))
    except (TypeError, ValueError):
        messages.error(request, "Select a valid referrer.")
        return redirect_target

    ref_booking = (
        Booking.objects.select_related('user')
        .filter(pk=ref_booking_id, room__pg_id=app.pg_id, status=Booking.APPROVED)
        .first()
    )
    if not ref_booking:
        messages.error(request, "Selected referrer is not valid for this PG.")
        return redirect_target
    if ref_booking.id == app.booking_id:
        messages.error(request, "An application cannot reference its own booking as referrer.")
        return redirect_target
    if ref_booking.user_id == app.user_id:
        messages.error(request, "A resident cannot refer themselves.")
        return redirect_target

    amount_raw = (request.POST.get('amount') or '').strip()
    if amount_raw:
        try:
            amount = Decimal(amount_raw)
        except (InvalidOperation, ValueError):
            messages.error(request, "Enter a valid referral amount.")
            return redirect_target
    else:
        amount = Decimal(str(app.pg.referral_amount or 0))

    if amount <= 0:
        messages.error(request, "Set a referral amount greater than zero to record credits.")
        return redirect_target

    scheduled_input = (request.POST.get('scheduled_month') or '').strip()
    scheduled_month = None
    if scheduled_input:
        try:
            year_str, month_str = scheduled_input.split('-', 1)
            scheduled_month = date(int(year_str), int(month_str), 1)
        except Exception:
            messages.error(request, "Invalid referral month selection.")
            return redirect_target

    notes = (request.POST.get('notes') or '').strip()[:255]

    app.referred_by_booking = ref_booking
    app.save(update_fields=['referred_by_booking'])

    default_month = scheduled_month
    if not default_month:
        anchor = getattr(app.booking, 'payment_date', None) or getattr(app.booking, 'joining_date', None)
        if anchor:
            default_month = anchor.replace(day=1)
        else:
            today = timezone.now().date()
            default_month = today.replace(day=1)

    credit = getattr(app, 'referral_credit', None)
    if credit and credit.redeemed_on:
        # Allow updating notes or schedule if it aligns with redeemed month
        if credit.referrer_booking_id != ref_booking.id:
            messages.error(request, "Referral credit already redeemed; cannot change the referrer.")
            return redirect_target
        if amount != credit.amount:
            messages.error(request, "Referral credit already redeemed; cannot change the amount.")
            return redirect_target
        if scheduled_month and credit.redeemed_for_month and credit.redeemed_for_month != scheduled_month:
            messages.error(request, "Referral credit already redeemed for a different month.")
            return redirect_target
        update_fields = []
        if scheduled_month and credit.scheduled_month != scheduled_month:
            credit.scheduled_month = scheduled_month
            update_fields.append('scheduled_month')
        if notes != credit.notes:
            credit.notes = notes
            update_fields.append('notes')
        if update_fields:
            credit.save(update_fields=update_fields)
        messages.success(request, "Referral details updated.")
        return redirect_target

    if not credit:
        credit = ReferralCredit.objects.create(
            pg=app.pg,
            referrer_user=ref_booking.user,
            referrer_booking=ref_booking,
            referred_user=app.user,
            referred_booking=app.booking,
            application=app,
            amount=amount,
            scheduled_month=default_month,
            notes=notes,
        )
    else:
        credit.referrer_user = ref_booking.user
        credit.referrer_booking = ref_booking
        credit.referred_user = app.user
        credit.referred_booking = app.booking
        credit.amount = amount
        credit.notes = notes
        if scheduled_month or not credit.scheduled_month:
            credit.scheduled_month = scheduled_month or default_month
        credit.save(update_fields=[
            'referrer_user',
            'referrer_booking',
            'referred_user',
            'referred_booking',
            'amount',
            'notes',
            'scheduled_month',
        ])

    messages.success(
        request,
        f"Referral recorded: {ref_booking.user.get_full_name() or ref_booking.user.email} will receive Rs.{amount:.2f}.",
    )
    return redirect_target


# ============================================================================
# ASYNC PDF GENERATION WITH PROGRESS TRACKING
# ============================================================================

@login_required
def tenants_export_pdf_async_start(request):
    """Start async PDF generation and return task ID"""
    from .pdf_tasks import PDFTaskManager
    import threading
    
    if not _require_pg_admin(request.user):
        return JsonResponse({'error': 'Forbidden'}, status=403)

    pg = _active_pg(request)
    if not pg:
        return JsonResponse({'error': 'No active PG selected'}, status=400)
    
    # Check if user already has an active task for this PG
    task_id, existing_task = PDFTaskManager.get_user_active_task(request.user.id, pg.id)
    if existing_task:
        return JsonResponse({
            'task_id': task_id,
            'status': existing_task['status'],
            'message': 'Task already in progress'
        })
    
    # Create new task
    task_id = PDFTaskManager.create_task(request.user.id, pg.id, pg.name)
    
    # Start background thread for PDF generation
    thread = threading.Thread(
        target=_generate_pdf_async,
        args=(task_id, request.user.id, pg.id),
        daemon=True
    )
    thread.start()
    
    return JsonResponse({
        'task_id': task_id,
        'status': 'pending',
        'message': 'PDF generation started'
    })


@login_required
def tenants_export_pdf_async_progress(request, task_id):
    """Check PDF generation progress"""
    from .pdf_tasks import PDFTaskManager
    
    task = PDFTaskManager.get_task(task_id)
    
    if not task:
        return JsonResponse({'error': 'Task not found'}, status=404)
    
    # Verify user owns this task
    if task['user_id'] != request.user.id:
        return JsonResponse({'error': 'Forbidden'}, status=403)
    
    return JsonResponse({
        'task_id': task_id,
        'status': task['status'],
        'progress': task['progress'],
        'message': task['message'],
        'error': task.get('error'),
    })


@login_required
def tenants_export_pdf_async_download(request, task_id):
    """Download completed PDF and delete task"""
    from .pdf_tasks import PDFTaskManager, get_pdf_storage_dir
    from datetime import datetime
    import os

    task = PDFTaskManager.get_task(task_id)
    fallback_mode = False

    if not task:
        # Recover from scenarios where the in-memory task dictionary is cleared (e.g., different worker)
        parts = task_id.split('_') if task_id else []
        fallback_user_id = None
        fallback_pg_id = None
        if len(parts) >= 4 and parts[0] == 'pdf':
            try:
                fallback_user_id = int(parts[1])
                fallback_pg_id = int(parts[2])
            except ValueError:
                fallback_user_id = None
        if fallback_user_id != request.user.id:
            return HttpResponse('Task not found', status=404)
        if fallback_pg_id is None or not _admin_pgs(request.user).filter(id=fallback_pg_id).exists():
            return HttpResponse('Task not found', status=404)

        fallback_file = os.path.join(get_pdf_storage_dir(), f"pdf_{task_id}.pdf")
        if not os.path.exists(fallback_file):
            return HttpResponse('Task not found', status=404)

        pg_name = PG.objects.filter(id=fallback_pg_id).values_list('name', flat=True).first() or 'PG'
        task = {
            'user_id': fallback_user_id,
            'pg_id': fallback_pg_id,
            'pg_name': pg_name,
            'status': 'completed',
            'file_path': fallback_file,
        }
        fallback_mode = True

    # Verify user owns this task
    if task['user_id'] != request.user.id:
        return HttpResponse('Forbidden', status=403)

    if task['status'] != 'completed':
        return HttpResponse('PDF not ready yet', status=400)

    file_path = task.get('file_path')
    if not file_path or not os.path.exists(file_path):
        return HttpResponse('PDF file not found', status=404)

    with open(file_path, 'rb') as f:
        pdf_data = f.read()

    if fallback_mode:
        try:
            os.remove(file_path)
        except OSError:
            pass
    else:
        PDFTaskManager.delete_task(task_id)

    response = HttpResponse(pdf_data, content_type='application/pdf')
    filename = f"{task['pg_name'].replace(' ', '_')}_tenants_{datetime.now().strftime('%B_%Y')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


@login_required
def tenants_export_pdf_async_cancel(request, task_id):
    """Cancel/delete a PDF generation task"""
    from .pdf_tasks import PDFTaskManager
    
    task = PDFTaskManager.get_task(task_id)
    
    if not task:
        return JsonResponse({'error': 'Task not found'}, status=404)
    
    # Verify user owns this task
    if task['user_id'] != request.user.id:
        return JsonResponse({'error': 'Forbidden'}, status=403)
    
    # Delete task
    PDFTaskManager.delete_task(task_id)
    
    return JsonResponse({'message': 'Task cancelled'})


def _generate_pdf_async(task_id, user_id, pg_id):
    """
    Background function to generate PDF with progress tracking.
    This runs in a separate thread.
    """
    from .pdf_tasks import PDFTaskManager, get_pdf_storage_dir
    import os
    from datetime import datetime
    
    try:
        PDFTaskManager.update_task(
            task_id,
            status='processing',
            progress=5,
            message='Fetching PG data...'
        )
        
        # Get PG object
        pg = PG.objects.get(id=pg_id)
        
        PDFTaskManager.update_task(
            task_id,
            progress=10,
            message='Fetching rooms and tenants...'
        )
        
        # Import required libraries
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm, inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether, Flowable
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_LEFT, TA_CENTER
        from reportlab.lib import colors
        import requests
        from concurrent.futures import ThreadPoolExecutor, as_completed
        from io import BytesIO
        from PIL import Image as PILImage
        
        # Get current date and month/year
        today = timezone.now().date()
        current_month_year = datetime.now().strftime('%B %Y')
        
        # Calculate current month range for expected rent calculation
        import calendar
        year = today.year
        month = today.month
        m_first = date(year, month, 1)
        last_day = calendar.monthrange(year, month)[1]
        m_last = date(year, month, last_day)
        
        PDFTaskManager.update_task(
            task_id,
            progress=15,
            message='Loading room data...'
        )
        
        # Fetch all rooms (we'll get bookings separately)
        rooms = list(Room.objects.filter(pg=pg).order_by('room_no'))
        
        total_rooms = len(rooms)
        
        PDFTaskManager.update_task(
            task_id,
            progress=20,
            message=f'Processing {total_rooms} rooms...'
        )
        
        # Create PDF file path
        pdf_dir = get_pdf_storage_dir()
        filename = f"pdf_{task_id}.pdf"
        file_path = os.path.join(pdf_dir, filename)
        
        # Create PDF document
        doc = SimpleDocTemplate(file_path, pagesize=A4, topMargin=10*mm, bottomMargin=10*mm, 
                               leftMargin=12*mm, rightMargin=12*mm)
        story = []
        
        # Styles
        styles = getSampleStyleSheet()
        pg_name_style = ParagraphStyle(
            'PGName',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1a237e'),
            alignment=TA_CENTER,
            spaceAfter=3*mm,
            fontName='Helvetica-Bold'
        )
        pg_addr_style = ParagraphStyle(
            'PGAddr',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_CENTER,
            spaceAfter=2*mm
        )
        pg_phone_style = ParagraphStyle(
            'PGPhone',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_CENTER,
            spaceAfter=2*mm,
            fontName='Helvetica-Bold'
        )
        pg_month_style = ParagraphStyle(
            'PGMonth',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_CENTER,
            spaceAfter=4*mm,
            fontName='Helvetica-Bold'
        )
        room_header_style = ParagraphStyle(
            'RoomHeader',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#0d47a1'),
            spaceAfter=2*mm,
            spaceBefore=3*mm,
            fontName='Helvetica-Bold'
        )
        label_style = ParagraphStyle(
            'Label',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            leading=10
        )
        value_style = ParagraphStyle(
            'Value',
            parent=styles['Normal'],
            fontSize=9,
            leading=11,
            fontName='Helvetica-Bold'
        )
        
        # Header
        story.append(Paragraph(pg.name, pg_name_style))
        story.append(Paragraph(pg.address or '', pg_addr_style))
        if pg.phone:
            story.append(Paragraph(f"Phone: {pg.phone}", pg_phone_style))
        story.append(Paragraph(current_month_year, pg_month_style))
        story.append(Spacer(1, 2*mm))
        
        PDFTaskManager.update_task(
            task_id,
            progress=25,
            message='Downloading tenant images...'
        )
        
        # Image cache
        _image_cache = {}
        
        # Custom Flowable for checkbox outline
        class OutlinedCheckbox(Flowable):
            def __init__(self, size=5*mm, checked=False):
                Flowable.__init__(self)
                self.size = size
                self.checked = checked
                self.width = size
                self.height = size
            
            def draw(self):
                self.canv.setStrokeColor(colors.black)
                self.canv.setLineWidth(0.5)
                self.canv.rect(0, 0, self.size, self.size, stroke=1, fill=0)
                if self.checked:
                    self.canv.setStrokeColor(colors.black)
                    self.canv.setLineWidth(1)
                    self.canv.line(0.15*self.size, 0.5*self.size, 0.4*self.size, 0.25*self.size)
                    self.canv.line(0.4*self.size, 0.25*self.size, 0.85*self.size, 0.75*self.size)
        
        def _get_image(url, default_width=18*mm, default_height=22*mm):
            """Download and cache image with extended timeout and retry support"""
            if url in _image_cache:
                return _image_cache[url]
            
            if not url:
                return None
            
            # Normalize Drive/Dropbox URLs
            if 'drive.google.com' in url and '/file/d/' in url:
                fid = url.split('/file/d/')[1].split('/')[0]
                url = f"https://drive.google.com/uc?export=download&id={fid}"
            elif 'dropbox.com' in url:
                url = url.replace('?dl=0', '?dl=1')
            
            try:
                # Increased timeout to 10s for very reliable image loading
                resp = requests.get(url, timeout=10, stream=True)
                resp.raise_for_status()
                
                img_data = BytesIO(resp.content)
                pil_img = PILImage.open(img_data)
                
                # Resize if needed
                max_w, max_h = int(default_width * 2.83), int(default_height * 2.83)
                if pil_img.width > max_w or pil_img.height > max_h:
                    pil_img.thumbnail((max_w, max_h), PILImage.Resampling.LANCZOS)
                
                # Convert to RGB if needed
                if pil_img.mode in ('RGBA', 'LA', 'P'):
                    bg = PILImage.new('RGB', pil_img.size, (255, 255, 255))
                    if pil_img.mode == 'P':
                        pil_img = pil_img.convert('RGBA')
                    bg.paste(pil_img, mask=pil_img.split()[-1] if pil_img.mode in ('RGBA', 'LA') else None)
                    pil_img = bg
                
                # Save to BytesIO
                buf = BytesIO()
                pil_img.save(buf, format='JPEG', quality=85)
                buf.seek(0)
                
                from reportlab.platypus import Image as RLImage
                rl_img = RLImage(buf, width=default_width, height=default_height)
                _image_cache[url] = rl_img
                return rl_img
            
            except Exception as e:
                # Cache the failure to avoid retrying in the same session
                _image_cache[url] = None
                return None
        
        PDFTaskManager.update_task(
            task_id,
            progress=30,
            message='Loading images sequentially for best quality...'
        )
        
        # Pre-fetch all shares to get booking details including future bookings
        image_urls = set()
        for room in rooms:
            total_shares = room.total_shares or 1
            for share_no in range(1, total_shares + 1):
                share = room.shares.filter(share_no=share_no).first()
                share_detail = _build_share_detail(room, share) if share else {}
                booking = share_detail.get('booking')
                future_booking = share_detail.get('future_booking')
                
                # Collect image URLs from both current and future bookings
                for bk in [booking, future_booking]:
                    if bk:
                        user = bk.user
                        app = share_detail.get('application') if bk == booking else share_detail.get('future_application')
                        selfie_url = getattr(app, 'selfie_url', None) or getattr(getattr(user, 'profile', None), 'selfie_url', None)
                        if selfie_url:
                            image_urls.add(selfie_url)
        
        # Download images one by one with retry logic for perfect loading
        if image_urls:
            completed = 0
            total_images = len(image_urls)
            
            for url in image_urls:
                # Try to load image with up to 3 retries
                max_retries = 3
                success = False
                
                for attempt in range(max_retries):
                    try:
                        result = _get_image(url)
                        if result is not None:
                            success = True
                            break
                        # If result is None, retry
                        import time
                        time.sleep(1)  # Wait 1 second before retry
                    except Exception as e:
                        if attempt < max_retries - 1:
                            import time
                            time.sleep(1)  # Wait before retry
                        continue
                
                completed += 1
                progress = 30 + int((completed / total_images) * 20)  # 30-50%
                status = "✓" if success else "✗"
                PDFTaskManager.update_task(
                    task_id,
                    progress=min(progress, 50),
                    message=f'Loading images: {completed}/{total_images} {status}'
                )
        
        PDFTaskManager.update_task(
            task_id,
            progress=55,
            message='Generating PDF pages...'
        )
        
        # Generate PDF content for each room
        for idx, room in enumerate(rooms):
            # Room header
            story.append(Paragraph(f"Room {room.room_no}", room_header_style))
            
            # Get shares for this room
            total_shares = room.total_shares or 1
            
            # Collect all cards for this room
            all_cards = []
            
            for share_no in range(1, total_shares + 1):
                # Use _build_share_detail to get both current and future bookings
                share = room.shares.filter(share_no=share_no).first()
                share_detail = _build_share_detail(room, share) if share else {}
                booking = share_detail.get('booking')
                future_booking = share_detail.get('future_booking')
                
                # Build single card with 3 columns: [selfie | details | checkbox]
                if booking:
                    user = booking.user
                    app = share_detail.get('application')
                    selfie_url = getattr(app, 'selfie_url', None) or getattr(getattr(user, 'profile', None), 'selfie_url', None)
                    
                    # Normalize URL before cache lookup (must match normalization during download)
                    normalized_url = None
                    if selfie_url:
                        normalized_url = selfie_url
                        # Google Drive: file/d/ID/view → uc?export=download&id=ID
                        if 'drive.google.com/file/d/' in normalized_url:
                            parts = normalized_url.split('/d/')
                            if len(parts) > 1:
                                file_id = parts[1].split('/')[0]
                                normalized_url = f'https://drive.google.com/uc?export=download&id={file_id}'
                        # Dropbox: ?dl=0 → ?dl=1
                        elif 'dropbox.com' in normalized_url and '?dl=0' in normalized_url:
                            normalized_url = normalized_url.replace('?dl=0', '?dl=1')
                    
                    selfie_img = _image_cache.get(normalized_url) if normalized_url else None
                    
                    name = f"{user.first_name} {user.last_name}".strip() or user.email
                    phone = getattr(app, 'phone', None) or getattr(getattr(user, 'profile', None), 'phone', '') or ''
                    joining = booking.joining_date or booking.start_date
                    joining_str = joining.strftime('%d/%m/%y') if joining else '—'
                    payment = booking.payment_date
                    payment_str = payment.strftime('%d/%m/%y') if payment else '—'
                    leaving = booking.leaving_date
                    leaving_str = leaving.strftime('%d/%m/%y') if leaving else '—'
                    
                    # Get monthly fee for this tenant (expected rent for current month)
                    # This calculation works even with leaving_date - it will be pro-rated
                    from finance.views import _expected_rent_for_user_pg_month
                    try:
                        monthly_fee = _expected_rent_for_user_pg_month(user, pg, booking, m_first, m_last, today=today)
                        
                        # If current month expected is 0 (e.g., tenant joining next month or already left)
                        # calculate next payment cycle
                        next_cycle_fee = 0.0
                        next_payment_date = None
                        if monthly_fee == 0 or monthly_fee < 0.01:
                            # Calculate next month range
                            next_month = month + 1 if month < 12 else 1
                            next_year = year if month < 12 else year + 1
                            next_m_first = date(next_year, next_month, 1)
                            next_last_day = calendar.monthrange(next_year, next_month)[1]
                            next_m_last = date(next_year, next_month, next_last_day)
                            
                            # Only calculate next cycle if tenant will still be present
                            # (no leaving date or leaving date is after next cycle starts)
                            if not leaving or leaving >= next_m_first:
                                try:
                                    next_cycle_fee = _expected_rent_for_user_pg_month(user, pg, booking, next_m_first, next_m_last, today=today)
                                    # Get the payment date for next cycle
                                    if booking.payment_date:
                                        next_payment_date = booking.payment_date
                                except Exception:
                                    pass
                    except Exception:
                        # Fallback to static fee if calculation fails
                        from finance.models import ResidentRate, Fees
                        resident_rate = ResidentRate.objects.filter(user=user, pg=pg, active=True).first()
                        if resident_rate:
                            monthly_fee = float(resident_rate.amount)
                        else:
                            room_share_type = str(room.total_shares or '')
                            fee_obj = Fees.objects.filter(pg=pg, share_type=room_share_type).first()
                            monthly_fee = float(fee_obj.amount) if fee_obj else 0.0
                        next_cycle_fee = 0.0
                        next_payment_date = None
                    
                    # Column 1: Selfie
                    if selfie_img:
                        selfie_cell = selfie_img
                    else:
                        selfie_cell = Paragraph("<i>No Photo</i>", ParagraphStyle('TinyText', parent=styles['Normal'], fontSize=7, alignment=TA_CENTER))
                    
                    # Column 2: Details
                    detail_style = ParagraphStyle('CardDetail', parent=styles['Normal'], fontSize=7, leading=9, wordWrap='CJK')
                    advance_amount = booking.advance_paid if booking.advance_paid is not None else Decimal('0')
                    advance_str = f"Rs.{advance_amount:.2f}"

                    # Build monthly fee display
                    if monthly_fee > 0:
                        monthly_fee_str = f"Monthly Fee: Rs.{monthly_fee:.0f}"
                    elif next_cycle_fee > 0:
                        # Show next cycle fee with payment date
                        if next_payment_date:
                            monthly_fee_str = f"Monthly Fee: Rs.{next_cycle_fee:.0f} (Next: {next_payment_date.strftime('%d/%m/%y')})"
                        else:
                            monthly_fee_str = f"Monthly Fee: Rs.{next_cycle_fee:.0f} (Next cycle)"
                    else:
                        monthly_fee_str = "Monthly Fee: Rs.0"

                    details_lines = [
                        f"<b>{name}</b>",
                        f"Phone: {phone}",
                        f"Join: {joining_str}",
                        f"Pay: {payment_str}",
                        f"Leave: {leaving_str}",
                        f"Advance: {advance_str}",
                        monthly_fee_str
                    ]
                    
                    details_cell = Paragraph("<br/>".join(details_lines), detail_style)
                    
                    # Column 3: Checkbox
                    checkbox_cell = OutlinedCheckbox(size=4*mm)
                    
                    # Build card for current booking
                    single_card_data = [[selfie_cell, details_cell, checkbox_cell]]
                    single_card = Table(single_card_data, colWidths=[18*mm, 35*mm, 7*mm], rowHeights=[26*mm])
                    single_card.setStyle(TableStyle([
                        ('VALIGN', (0, 0), (0, 0), 'MIDDLE'),
                        ('VALIGN', (1, 0), (1, 0), 'TOP'),
                        ('VALIGN', (2, 0), (2, 0), 'MIDDLE'),
                        ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                        ('ALIGN', (2, 0), (2, 0), 'CENTER'),
                        ('BOX', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('LEFTPADDING', (0, 0), (-1, -1), 2),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                        ('TOPPADDING', (0, 0), (-1, -1), 2),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                    ]))
                    all_cards.append(single_card)
                    
                    # Add future booking details if exists - as a separate card
                    if future_booking:
                        future_user = future_booking.user
                        future_name = f"{future_user.first_name} {future_user.last_name}".strip() or future_user.email
                        future_joining = future_booking.joining_date
                        future_joining_str = future_joining.strftime('%d/%m/%y') if future_joining else '—'
                        future_leaving = future_booking.leaving_date
                        
                        # Get future tenant's photo
                        future_app = share_detail.get('future_application')
                        future_selfie_url = getattr(future_app, 'selfie_url', None) or getattr(getattr(future_user, 'profile', None), 'selfie_url', None)
                        
                        # Normalize future selfie URL
                        future_normalized_url = None
                        if future_selfie_url:
                            future_normalized_url = future_selfie_url
                            if 'drive.google.com/file/d/' in future_normalized_url:
                                parts = future_normalized_url.split('/d/')
                                if len(parts) > 1:
                                    file_id = parts[1].split('/')[0]
                                    future_normalized_url = f'https://drive.google.com/uc?export=download&id={file_id}'
                            elif 'dropbox.com' in future_normalized_url and '?dl=0' in future_normalized_url:
                                future_normalized_url = future_normalized_url.replace('?dl=0', '?dl=1')
                        
                        future_selfie_img = _image_cache.get(future_normalized_url) if future_normalized_url else None
                        
                        # Get monthly fee for future tenant (expected rent for current month)
                        # This calculation works even with leaving_date - it will be pro-rated
                        try:
                            future_monthly_fee = _expected_rent_for_user_pg_month(future_user, pg, future_booking, m_first, m_last, today=today)
                            
                            # If current month expected is 0, calculate next payment cycle
                            future_next_cycle_fee = 0.0
                            future_next_payment_date = None
                            if future_monthly_fee == 0 or future_monthly_fee < 0.01:
                                # Calculate next month range
                                next_month = month + 1 if month < 12 else 1
                                next_year = year if month < 12 else year + 1
                                next_m_first = date(next_year, next_month, 1)
                                next_last_day = calendar.monthrange(next_year, next_month)[1]
                                next_m_last = date(next_year, next_month, next_last_day)
                                
                                # Only calculate next cycle if tenant will still be present
                                if not future_leaving or future_leaving >= next_m_first:
                                    try:
                                        future_next_cycle_fee = _expected_rent_for_user_pg_month(future_user, pg, future_booking, next_m_first, next_m_last, today=today)
                                        # Get the payment date for next cycle
                                        if future_booking.payment_date:
                                            future_next_payment_date = future_booking.payment_date
                                    except Exception:
                                        pass
                        except Exception:
                            # Fallback: try to get custom rate or use same as current tenant
                            from finance.models import ResidentRate
                            future_resident_rate = ResidentRate.objects.filter(user=future_user, pg=pg, active=True).first()
                            if future_resident_rate:
                                future_monthly_fee = float(future_resident_rate.amount)
                            else:
                                future_monthly_fee = monthly_fee  # Use same as current if not set
                            future_next_cycle_fee = 0.0
                            future_next_payment_date = None
                        
                        # Future tenant selfie
                        if future_selfie_img:
                            future_selfie_cell = future_selfie_img
                        else:
                            future_selfie_cell = Paragraph("<i>No Photo</i>", ParagraphStyle('TinyText', parent=styles['Normal'], fontSize=7, alignment=TA_CENTER))
                        
                        # Future tenant details
                        # Build monthly fee display for future tenant
                        if future_monthly_fee > 0:
                            future_fee_str = f"Monthly Fee: Rs.{future_monthly_fee:.0f}"
                        elif future_next_cycle_fee > 0:
                            # Show next cycle fee with payment date
                            if future_next_payment_date:
                                future_fee_str = f"Monthly Fee: Rs.{future_next_cycle_fee:.0f} (Next: {future_next_payment_date.strftime('%d/%m/%y')})"
                            else:
                                future_fee_str = f"Monthly Fee: Rs.{future_next_cycle_fee:.0f} (Next cycle)"
                        else:
                            future_fee_str = "Monthly Fee: Rs.0"
                        
                        future_details_lines = [
                            "<b>---NEXT---</b>",
                            f"<b>{future_name}</b>",
                            f"Join: {future_joining_str}",
                            future_fee_str
                        ]
                        future_details_cell = Paragraph("<br/>".join(future_details_lines), detail_style)
                        
                        # Future tenant checkbox
                        future_checkbox_cell = OutlinedCheckbox(size=4*mm)
                        
                        # Build separate card for future booking
                        future_card_data = [[future_selfie_cell, future_details_cell, future_checkbox_cell]]
                        future_card = Table(future_card_data, colWidths=[18*mm, 35*mm, 7*mm], rowHeights=[22*mm])
                        future_card.setStyle(TableStyle([
                            ('VALIGN', (0, 0), (0, 0), 'MIDDLE'),
                            ('VALIGN', (1, 0), (1, 0), 'TOP'),
                            ('VALIGN', (2, 0), (2, 0), 'MIDDLE'),
                            ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                            ('ALIGN', (2, 0), (2, 0), 'CENTER'),
                            ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#90EE90')),
                            ('LEFTPADDING', (0, 0), (-1, -1), 2),
                            ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                            ('TOPPADDING', (0, 0), (-1, -1), 2),
                            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                        ]))
                        all_cards.append(future_card)
                else:
                    # Empty bed card
                    vacant_text = Paragraph("<i>VACANT</i>", ParagraphStyle('VacantText', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER, textColor=colors.grey))
                    empty_detail = Paragraph(f"<i>Bed {share_no}</i>", ParagraphStyle('EmptyDetail', parent=styles['Normal'], fontSize=7, textColor=colors.grey))
                    checkbox_cell = OutlinedCheckbox(size=4*mm)
                    
                    empty_card_data = [[vacant_text, empty_detail, checkbox_cell]]
                    empty_card = Table(empty_card_data, colWidths=[18*mm, 35*mm, 7*mm], rowHeights=[26*mm])
                    empty_card.setStyle(TableStyle([
                        ('VALIGN', (0, 0), (0, 0), 'MIDDLE'),
                        ('VALIGN', (1, 0), (1, 0), 'MIDDLE'),
                        ('VALIGN', (2, 0), (2, 0), 'MIDDLE'),
                        ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                        ('ALIGN', (1, 0), (1, 0), 'LEFT'),
                        ('ALIGN', (2, 0), (2, 0), 'CENTER'),
                        ('BOX', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('LEFTPADDING', (0, 0), (-1, -1), 2),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                        ('TOPPADDING', (0, 0), (-1, -1), 2),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                    ]))
                    all_cards.append(empty_card)
            
            # Arrange cards in rows of 3
            cards_per_row = 3
            for i in range(0, len(all_cards), cards_per_row):
                row_cards = all_cards[i:i+cards_per_row]
                # Pad row if less than 3 cards
                while len(row_cards) < cards_per_row:
                    row_cards.append(Paragraph("", styles['Normal']))  # empty placeholder
                
                # Create row table
                row_table = Table([row_cards], colWidths=[60*mm, 60*mm, 60*mm])
                row_table.setStyle(TableStyle([
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 2),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                    ('TOPPADDING', (0, 0), (-1, -1), 0),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
                ]))
                story.append(row_table)
                story.append(Spacer(1, 2*mm))
            
            # Add space between rooms
            story.append(Spacer(1, 3*mm))
            
            # Update progress
            progress = 55 + int(((idx + 1) / total_rooms) * 35)  # 55-90%
            PDFTaskManager.update_task(
                task_id,
                progress=min(progress, 90),
                message=f'Generated {idx + 1}/{total_rooms} rooms...'
            )
        
        PDFTaskManager.update_task(
            task_id,
            progress=95,
            message='Building final PDF...'
        )
        
        # Build PDF
        doc.build(story)
        
        PDFTaskManager.update_task(
            task_id,
            status='completed',
            progress=100,
            message='PDF generated successfully!',
            file_path=file_path,
            completed_at=timezone.now()
        )
        
    except Exception as e:
        PDFTaskManager.update_task(
            task_id,
            status='failed',
            progress=0,
            message='PDF generation failed',
            error=str(e)
        )


# ============================================================================
# ENHANCED LEAVE MANAGEMENT (PG ADMIN)
# ============================================================================

@login_required
def leaving_requests(request):
    """Enhanced leaving requests page with advance management
    
    Shows ALL bookings with leaving_date set, including:
    - Pending (leaving_date set, not confirmed)
    - Confirmed (leaving_confirmed_date set)
    - COMPLETED status (already left, for advance return/re-continue/delete operations)
    """
    if not _require_pg_admin(request.user):
        messages.error(request, "You must be a PG Admin.")
        return redirect('dashboard')
    
    pg = _active_pg(request)
    if not pg:
        messages.error(request, "No PG assigned.")
        return redirect('dashboard')
    
    today = date.today()
    
    # Get ALL leave requests including COMPLETED bookings
    # This allows admin to manage advance returns, deletions, and re-continue operations
    leave_requests = Booking.objects.filter(
        room__pg=pg,
        leaving_date__isnull=False,
        status__in=[Booking.APPROVED, Booking.COMPLETED]  # Include both APPROVED and COMPLETED
    ).select_related(
        'user', 'user__profile', 'room', 'application'
    ).order_by('-leaving_date', '-leaving_initiated_at')  # Most recent leaving dates first
    
    context = {
        'pg': pg,
        'pgs': list(_admin_pgs(request.user)),
        'leave_requests': leave_requests,
        'today': today,
    }
    
    return render(request, 'pgadmin/leaving_requests_enhanced.html', context)


@login_required
def refresh_old_tenants(request):
    """Refresh OldTenant records by:
    1. Finding APPROVED bookings with past leaving dates and marking them COMPLETED
    2. Archiving all COMPLETED bookings not yet in OldTenant
    """
    from pgadmin.models import OldTenant
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    if not _require_pg_admin(request.user):
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    pg = _active_pg(request)
    if not pg:
        return JsonResponse({'error': 'No PG assigned'}, status=400)
    
    today = date.today()
    stats = {
        'bookings_marked_completed': 0,
        'bookings_archived': 0,
        'already_archived': 0,
    }
    
    # Step 1: Mark APPROVED bookings with past leaving dates as COMPLETED
    approved_past_leaving = Booking.objects.filter(
        pg=pg,
        status=Booking.APPROVED,
        leaving_date__lt=today
    ).select_related('room', 'user', 'application')
    
    for booking in approved_past_leaving:
        booking.status = Booking.COMPLETED
        booking.save(update_fields=['status'])
        stats['bookings_marked_completed'] += 1
    
    # Step 2: Archive all COMPLETED bookings not yet in OldTenant
    completed_bookings = Booking.objects.filter(
        pg=pg,
        status=Booking.COMPLETED
    ).select_related('room', 'user', 'application')
    
    for booking in completed_bookings:
        # Check if already archived (by original_booking_id)
        existing = OldTenant.objects.filter(
            pg=pg,
            original_booking_id=booking.id
        ).exists()
        
        if existing:
            stats['already_archived'] += 1
            continue
        
        # Archive the booking
        app = getattr(booking, 'application', None)
        
        full_name = ''
        father_name = ''
        mother_name = ''
        email = ''
        phone = ''
        whatsapp_number = ''
        address = ''
        
        if app:
            full_name = app.name or ''
            father_name = app.father_name or ''
            mother_name = app.mother_name or ''
            email = app.email or ''
            phone = app.phone or ''
            whatsapp_number = app.whatsapp_number or ''
            address = app.address or ''
        
        # Fallback to user data
        if not full_name and booking.user:
            full_name = f"{booking.user.first_name or ''} {booking.user.last_name or ''}".strip() or booking.user.email.split('@')[0]
        if not email and booking.user:
            email = booking.user.email or ''
        
        # Create OldTenant record
        if full_name:
            OldTenant.objects.create(
                pg=pg,
                full_name=full_name,
                father_name=father_name,
                mother_name=mother_name,
                email=email,
                phone=phone,
                whatsapp_number=whatsapp_number,
                address=address,
                room_no=getattr(getattr(booking, 'room', None), 'room_no', ''),
                bed_no=str(booking.share_no) if booking.share_no else '',
                joining_date=booking.joining_date,
                leaving_date=booking.leaving_date,
                leaving_reason=booking.leaving_reason or '',
                advance_paid=booking.advance_paid or 0,
                advance_returned=booking.advance_returned_amount if booking.advance_returned else 0,
                original_user=booking.user,
                original_booking_id=booking.id,
                archived_by=request.user,
                dob=app.dob if app else None,
                age=app.age if app else None,
                father_phone=app.father_phone if app else '',
                mother_phone=app.mother_phone if app else '',
                emergency_contact=app.emergency_contact if app else '',
                food_pref=app.food_pref if app else '',
                marital_status=app.marital_status if app else '',
                education=app.education if app else '',
                occupation=app.occupation if app else '',
                org_name=app.org_name if app else '',
                org_address=app.org_address if app else '',
                has_vehicle=app.has_vehicle if app else False,
                vehicle_number=app.vehicle_number if app else '',
                vehicle_model=app.vehicle_model if app else '',
                aadhaar_number=app.aadhaar_number if app else '',
                selfie_url=app.selfie_url if app else getattr(booking.user, 'profile', None).selfie_url if hasattr(booking.user, 'profile') else '',
                aadhaar_file_url=app.aadhaar_file_url if app else getattr(booking.user, 'profile', None).aadhaar_file_url if hasattr(booking.user, 'profile') else '',
                aadhaar_file_url_2=app.aadhaar_file_url_2 if app else getattr(booking.user, 'profile', None).aadhaar_file_url_2 if hasattr(booking.user, 'profile') else '',
            )
            stats['bookings_archived'] += 1
    
    return JsonResponse({
        'success': True,
        'message': f"Refresh complete. Marked {stats['bookings_marked_completed']} as completed, archived {stats['bookings_archived']} new records ({stats['already_archived']} already existed).",
        'stats': stats
    })


@login_required
def old_tenants(request):
    """View archived old tenant records with statistics
    
    Shows all archived tenant data from deleted bookings with:
    - Statistics on tenants joined vs left per month
    - Filters and search functionality
    - Sortable columns
    """
    from pgadmin.models import OldTenant
    from django.db.models import Count
    from django.db.models.functions import TruncMonth
    from collections import defaultdict
    import json
    
    if not _require_pg_admin(request.user):
        messages.error(request, "You must be a PG Admin.")
        return redirect('dashboard')
    
    pg = _active_pg(request)
    if not pg:
        messages.error(request, "No PG assigned.")
        return redirect('dashboard')
    
    # Get all old tenants for this PG
    old_tenant_records = OldTenant.objects.filter(pg=pg).order_by('-archived_at')
    
    # Calculate statistics for the last 12 months
    today = date.today()
    twelve_months_ago = today.replace(year=today.year - 1) if today.month == today.day == 1 else date(today.year - 1, today.month, 1)
    
    # Get monthly statistics from OldTenant records
    # Tenants who left per month (based on leaving_date)
    left_by_month = (
        OldTenant.objects.filter(
            pg=pg,
            leaving_date__gte=twelve_months_ago
        )
        .annotate(month=TruncMonth('leaving_date'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )
    
    # Tenants who joined per month (based on joining_date from old tenants)
    joined_from_old = (
        OldTenant.objects.filter(
            pg=pg,
            joining_date__gte=twelve_months_ago
        )
        .annotate(month=TruncMonth('joining_date'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )
    
    # Also get current active tenants who joined in the last 12 months
    joined_from_active = (
        Booking.objects.filter(
            room__pg=pg,
            status__in=[Booking.APPROVED, Booking.COMPLETED],
            joining_date__gte=twelve_months_ago
        )
        .annotate(month=TruncMonth('joining_date'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )
    
    # Build chart data - last 12 months
    chart_months = []
    chart_joined = []
    chart_left = []
    
    # Create a dict for quick lookup
    left_dict = {item['month']: item['count'] for item in left_by_month if item['month']}
    joined_old_dict = {item['month']: item['count'] for item in joined_from_old if item['month']}
    joined_active_dict = {item['month']: item['count'] for item in joined_from_active if item['month']}
    
    # Generate last 12 months
    for i in range(11, -1, -1):
        month_date = today.replace(day=1)
        for _ in range(i):
            # Go back one month
            if month_date.month == 1:
                month_date = month_date.replace(year=month_date.year - 1, month=12)
            else:
                month_date = month_date.replace(month=month_date.month - 1)
        
        month_key = month_date
        chart_months.append(month_date.strftime('%b %Y'))
        
        # Combined joined count from old tenants and active bookings
        joined_old = joined_old_dict.get(month_key, 0)
        joined_active = joined_active_dict.get(month_key, 0)
        # Avoid double counting - use max or combine intelligently
        # Since old tenants' joining data might overlap with deleted bookings that were once active
        chart_joined.append(joined_old + joined_active)
        
        chart_left.append(left_dict.get(month_key, 0))
    
    # Calculate summary statistics
    total_old_tenants = old_tenant_records.count()
    avg_stay_days = None
    if total_old_tenants > 0:
        stays = [ot.stay_duration_days for ot in old_tenant_records if ot.stay_duration_days is not None]
        if stays:
            avg_stay_days = sum(stays) // len(stays)
    
    # Get most common leaving reasons
    leaving_reasons = old_tenant_records.exclude(leaving_reason='').values_list('leaving_reason', flat=True)
    reason_counts = defaultdict(int)
    for reason in leaving_reasons:
        # Normalize and count common words/phrases
        reason_lower = reason.lower().strip()
        if reason_lower:
            reason_counts[reason_lower] += 1
    top_reasons = sorted(reason_counts.items(), key=lambda x: -x[1])[:5]
    
    context = {
        'pg': pg,
        'pgs': list(_admin_pgs(request.user)),
        'old_tenants': old_tenant_records,
        'total_old_tenants': total_old_tenants,
        'avg_stay_days': avg_stay_days,
        'top_reasons': top_reasons,
        'chart_months': json.dumps(chart_months),
        'chart_joined': json.dumps(chart_joined),
        'chart_left': json.dumps(chart_left),
        'total_joined_12m': sum(chart_joined),
        'total_left_12m': sum(chart_left),
    }
    
    return render(request, 'pgadmin/old_tenants.html', context)


@login_required
def confirm_leave(request, booking_id):
    """Confirm leave request"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    booking = get_object_or_404(
        Booking.objects.select_related('room', 'room__pg', 'user'),
        id=booking_id
    )
    
    if not _require_pg_admin(request.user) or not _admin_pgs(request.user).filter(id=booking.room.pg.id).exists():
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    if not booking.leaving_date:
        return JsonResponse({'error': 'No leaving date set'}, status=400)
    
    if booking.leaving_confirmed_date:
        return JsonResponse({'error': 'Already confirmed'}, status=400)
    
    # Confirm the leave
    booking.leaving_confirmed_date = booking.leaving_date
    booking.save(update_fields=['leaving_confirmed_date'])
    
    # Update RoomShareStatus to vacant_from
    share = RoomShareStatus.objects.filter(room=booking.room, share_no=booking.share_no).first()
    if share and share.status != RoomShareStatus.VACANT:
        share.status = RoomShareStatus.VACANT_FROM
        share.vacant_from = booking.leaving_date
        share.save(update_fields=['status', 'vacant_from'])
    
    # Notify tenant and all PG admins.
    user_detail_path = f"{reverse('booking_detail', args=[booking.id])}?{urlencode({'pg': booking.room.pg_id})}"
    Notification.objects.create(
        user=booking.user,
        title="Leave Request Confirmed",
        message=f"Your leave request for {booking.room.room_no}, Bed {booking.share_no} on {booking.leaving_date.strftime('%B %d, %Y')} has been confirmed."
    )
    send_push_to_user(
        booking.user,
        title="Leave Request Confirmed",
        body=f"Leave for Room {booking.room.room_no}, Bed {booking.share_no} on {booking.leaving_date.strftime('%b %d, %Y')} is confirmed.",
        url=user_detail_path,
        extra_data={'type': 'leave_confirmed', 'booking_id': booking.id, 'pg_id': booking.room.pg_id},
    )

    try:
        admin_path, admin_payload = _leave_admin_path_and_payload(booking)
        admin_users = _pg_admin_users_for_pg(booking.room.pg_id)
        tenant_name = booking.user.get_full_name() or booking.user.email

        for admin_user in admin_users:
            Notification.objects.create(
                user=admin_user,
                title="Leave Request Confirmed",
                message=(
                    f"{tenant_name}'s leave request for Room {booking.room.room_no}, "
                    f"Bed {booking.share_no} has been confirmed."
                ),
            )

        send_push_to_users(
            admin_users,
            title="Leave Request Confirmed",
            body=f"{tenant_name} leave confirmed for Room {booking.room.room_no}, Bed {booking.share_no}.",
            url=admin_path,
            extra_data={**admin_payload, 'type': 'leave_confirmed', 'source': 'pg_leave_confirm'},
        )
        
        # Send emails
        from django.core.mail import send_mail
        from django.conf import settings
        from django.utils import timezone
        
        subject = f"Leave Request Confirmed - {booking.room.pg.name}"
        tenant_name = booking.user.get_full_name() or booking.user.email
        eligibility = "Eligible" if booking.advance_eligible else "Not Eligible"
        message_body = (
            f"A leave request has been confirmed.\n\n"
            f"Tenant Name: {tenant_name}\n"
            f"Tenant Email: {booking.user.email}\n"
            f"PG Name: {booking.room.pg.name}\n"
            f"Room Number: {booking.room.room_no}\n"
            f"Bed: {booking.share_no}\n"
            f"Joining Date: {booking.joining_date}\n"
            f"Leave Initiation Time: {timezone.localtime(booking.leaving_initiated_at).strftime('%I:%M %p') if booking.leaving_initiated_at else 'N/A'}\n"
            f"Confirmed Leave Date: {booking.leaving_date}\n"
            f"Advance Refund Eligibility: {eligibility}\n"
        )
        admin_emails = [a.email for a in admin_users if getattr(a, 'email', None)]
        if booking.user and booking.user.email:
            send_mail(
                subject,
                message_body,
                settings.DEFAULT_FROM_EMAIL,
                [booking.user.email] + admin_emails,
                fail_silently=True,
            )
        elif admin_emails:
            send_mail(
                subject,
                message_body,
                settings.DEFAULT_FROM_EMAIL,
                admin_emails,
                fail_silently=True,
            )

    except Exception:
        _logger.exception('Leave confirm admin notification dispatch failed for booking %s', booking.id)
    
    # Audit log
    log(
        actor=request.user,
        action='leave_confirmed',
        target_type='Booking',
        target_id=booking.id,
        message=f"Leave confirmed for {booking.user.get_full_name()}, leaving on {booking.leaving_date}",
        meta={
            'leaving_date': booking.leaving_date.isoformat(),
            'tenant': booking.user.get_full_name()
        }
    )
    
    messages.success(request, f"Leave confirmed for {booking.user.get_full_name()} on {booking.leaving_date}.")
    return JsonResponse({'success': True, 'leaving_confirmed_date': booking.leaving_confirmed_date.isoformat()})


@login_required
def reject_leave(request, booking_id):
    """Reject leave request"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    booking = get_object_or_404(
        Booking.objects.select_related('room', 'room__pg', 'user'),
        id=booking_id
    )
    
    if not _require_pg_admin(request.user) or not _admin_pgs(request.user).filter(id=booking.room.pg.id).exists():
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    if booking.leaving_confirmed_date:
        return JsonResponse({'error': 'Already confirmed - cannot reject'}, status=400)
    
    # Clear leave request
    old_leaving_date = booking.leaving_date
    booking.leaving_date = None
    booking.leaving_reason = ''
    booking.leaving_initiated_at = None
    booking.advance_eligible = True
    booking.save(update_fields=[
        'leaving_date', 'leaving_reason', 'leaving_initiated_at', 'advance_eligible'
    ])
    
    # Notify tenant and all PG admins.
    user_detail_path = f"{reverse('booking_detail', args=[booking.id])}?{urlencode({'pg': booking.room.pg_id})}"
    Notification.objects.create(
        user=booking.user,
        title="Leave Request Rejected",
        message=f"Your leave request for {booking.room.room_no}, Bed {booking.share_no} on {old_leaving_date} has been rejected. Please contact admin for details."
    )
    send_push_to_user(
        booking.user,
        title="Leave Request Rejected",
        body=f"Leave request for Room {booking.room.room_no}, Bed {booking.share_no} was rejected.",
        url=user_detail_path,
        extra_data={'type': 'leave_rejected', 'booking_id': booking.id, 'pg_id': booking.room.pg_id},
    )

    try:
        admin_path, admin_payload = _leave_admin_path_and_payload(booking)
        admin_users = _pg_admin_users_for_pg(booking.room.pg_id)
        tenant_name = booking.user.get_full_name() or booking.user.email

        for admin_user in admin_users:
            Notification.objects.create(
                user=admin_user,
                title="Leave Request Rejected",
                message=(
                    f"{tenant_name}'s leave request for Room {booking.room.room_no}, "
                    f"Bed {booking.share_no} was rejected."
                ),
            )

        send_push_to_users(
            admin_users,
            title="Leave Request Rejected",
            body=f"{tenant_name} leave request rejected for Room {booking.room.room_no}, Bed {booking.share_no}.",
            url=admin_path,
            extra_data={**admin_payload, 'type': 'leave_rejected', 'source': 'pg_leave_reject'},
        )
    except Exception:
        _logger.exception('Leave reject admin notification dispatch failed for booking %s', booking.id)
    
    # Audit log
    log(
        actor=request.user,
        action='leave_rejected',
        target_type='Booking',
        target_id=booking.id,
        message=f"Leave request rejected for {booking.user.get_full_name()}",
        meta={
            'rejected_date': old_leaving_date.isoformat() if old_leaving_date else None,
            'tenant': booking.user.get_full_name()
        }
    )
    
    messages.success(request, f"Leave request rejected for {booking.user.get_full_name()}.")
    return JsonResponse({'success': True})


@login_required
def edit_leave_date(request, booking_id):
    """Edit leave date for a booking"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    booking = get_object_or_404(
        Booking.objects.select_related('room', 'room__pg', 'user'),
        id=booking_id
    )
    
    if not _require_pg_admin(request.user) or not _admin_pgs(request.user).filter(id=booking.room.pg.id).exists():
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    new_date_str = request.POST.get('new_date')
    if not new_date_str:
        return JsonResponse({'error': 'new_date required'}, status=400)
    
    try:
        new_date = parse_date(new_date_str)
        if not new_date:
            raise ValueError("Invalid date format")
    except:
        return JsonResponse({'error': 'Invalid date format'}, status=400)
    
    # Allow past dates for record-keeping, but require explicit acknowledgement
    # when changing from a current/future leaving date to a past date.
    if new_date < date.today():
        old_date = booking.leaving_date
        # If the original leaving date was today or in the future, require an explicit acknowledgement
        if old_date and old_date >= date.today():
            if request.POST.get('past_ack') != '1':
                return JsonResponse({'error': 'Past date acknowledgement required'}, status=400)
        # otherwise (editing an already-past date or no prior date) allow the change
    
    # Recalculate advance eligibility
    if booking.leaving_initiated_at:
        notice_period = getattr(booking.room.pg, 'notice_period', 30)
        days_diff = (new_date - booking.leaving_initiated_at.date()).days
        booking.advance_eligible = days_diff >= notice_period
    
    old_date = booking.leaving_date
    booking.leaving_date = new_date
    
    # Update confirmed date if already confirmed
    if booking.leaving_confirmed_date:
        booking.leaving_confirmed_date = new_date
        # Update share vacant_from
        share = RoomShareStatus.objects.filter(room=booking.room, share_no=booking.share_no).first()
        if share:
            share.vacant_from = new_date
            share.save(update_fields=['vacant_from'])
    
    booking.save(update_fields=['leaving_date', 'leaving_confirmed_date', 'advance_eligible'])
    
    # Notify user
    Notification.objects.create(
        user=booking.user,
        title="Leave Date Updated",
        message=f"Your leave date has been updated from {old_date} to {new_date} by PG admin."
    )
    send_push_to_user(
        booking.user,
        title="Leave Date Updated",
        body=f"Leave date updated to {new_date} for Room {booking.room.room_no}, Bed {booking.share_no}.",
        url=reverse('booking_detail', args=[booking.id]),
        extra_data={'type': 'leave_date_updated', 'booking_id': booking.id},
    )
    
    # Audit log
    log(
        actor=request.user,
        action='leave_date_edited',
        target_type='Booking',
        target_id=booking.id,
        message=f"Leave date updated from {old_date} to {new_date} for {booking.user.get_full_name()}",
        meta={
            'old_date': old_date.isoformat() if old_date else None,
            'new_date': new_date.isoformat(),
            'tenant': booking.user.get_full_name()
        }
    )
    
    return JsonResponse({
        'success': True,
        'new_date': new_date.isoformat(),
        'advance_eligible': booking.advance_eligible
    })


@login_required
def mark_advance_returned(request, booking_id):
    """Mark advance as returned and create expenditure"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    booking = get_object_or_404(
        Booking.objects.select_related('room', 'room__pg', 'user'),
        id=booking_id
    )
    
    if not _require_pg_admin(request.user) or not _admin_pgs(request.user).filter(id=booking.room.pg.id).exists():
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    if not booking.advance_eligible:
        return JsonResponse({'error': 'Not eligible for advance return'}, status=400)
    
    if booking.advance_returned:
        return JsonResponse({'error': 'Advance already marked as returned'}, status=400)
    
    amount_str = request.POST.get('amount', '')
    try:
        amount = Decimal(amount_str)
        if amount <= 0:
            raise ValueError("Amount must be positive")
    except:
        return JsonResponse({'error': 'Invalid amount'}, status=400)
    
    # Validate amount doesn't exceed advance_paid
    if amount > booking.advance_paid:
        return JsonResponse({
            'warning': True,
            'message': f'Amount (Rs.{amount}) exceeds advance paid (Rs.{booking.advance_paid}). Continue anyway?'
        })
    
    # Create expenditure entry
    from finance.models import Expenditure
    
    expenditure = Expenditure.objects.create(
        pg=booking.room.pg,
        category='advance_return',
        amount=amount,
        date=date.today(),
        notes=f"Advance returned to {booking.user.get_full_name()} (Room {booking.room.room_no}, Bed {booking.share_no}). Leaving date: {booking.leaving_date}",
        booking=booking
    )
    
    # Mark advance as returned
    booking.advance_returned = True
    booking.advance_returned_at = timezone.now()
    booking.advance_returned_amount = amount
    booking.save(update_fields=['advance_returned', 'advance_returned_at', 'advance_returned_amount'])
    
    # Notify user
    Notification.objects.create(
        user=booking.user,
        title="Advance Amount Returned",
        message=f"Your advance amount of Rs.{amount} has been returned for Room {booking.room.room_no}, Bed {booking.share_no}."
    )
    send_push_to_user(
        booking.user,
        title="Advance Amount Returned",
        body=f"Rs.{amount} returned for Room {booking.room.room_no}, Bed {booking.share_no}.",
        url=reverse('booking_detail', args=[booking.id]),
        extra_data={'type': 'advance_returned', 'booking_id': booking.id},
    )
    
    # Audit log
    log(
        actor=request.user,
        action='advance_returned',
        target_type='Booking',
        target_id=booking.id,
        message=f"Advance of Rs.{amount} returned to {booking.user.get_full_name()} for room {booking.room.room_no} bed {booking.share_no}",
        meta={
            'amount': str(amount),
            'expenditure_id': expenditure.id,
            'tenant': booking.user.get_full_name()
        }
    )
    
    messages.success(request, f"Advance of Rs.{amount} marked as returned to {booking.user.get_full_name()}. Expenditure entry created.")
    return JsonResponse({
        'success': True,
        'amount': str(amount),
        'returned_at': booking.advance_returned_at.isoformat()
    })


@login_required
def edit_advance_returned_amount(request, booking_id):
    """Edit advance returned amount and update associated expenditure"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    booking = get_object_or_404(
        Booking.objects.select_related('room', 'room__pg', 'user'),
        id=booking_id
    )
    
    if not _require_pg_admin(request.user) or not _admin_pgs(request.user).filter(id=booking.room.pg.id).exists():
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    if not booking.advance_returned:
        return JsonResponse({'error': 'Advance not yet marked as returned'}, status=400)
    
    amount_str = request.POST.get('amount', '')
    try:
        amount = Decimal(amount_str)
        if amount < 0:
            raise ValueError("Amount cannot be negative")
    except:
        return JsonResponse({'error': 'Invalid amount'}, status=400)
    
    old_amount = booking.advance_returned_amount
    
    # Update booking
    booking.advance_returned_amount = amount
    booking.save(update_fields=['advance_returned_amount'])
    
    # Find and update the associated expenditure by matching notes
    from finance.models import Expenditure
    
    # Search for expenditure with matching booking reference in notes
    expenditures = Expenditure.objects.filter(
        pg=booking.room.pg,
        category='advance_return',
        booking=booking
    ).order_by('-date')
    
    if expenditures.exists():
        # Update the most recent expenditure
        expenditure = expenditures.first()
        expenditure.amount = amount
        expenditure.notes = f"Advance returned to {booking.user.get_full_name()} (Room {booking.room.room_no}, Bed {booking.share_no}). Leaving date: {booking.leaving_date}. [Edited from Rs.{old_amount}]"
        expenditure.save(update_fields=['amount', 'notes'])
    
    # Audit log
    log(
        actor=request.user,
        action='advance_amount_edited',
        target_type='Booking',
        target_id=booking.id,
        message=f"Advance returned amount edited from Rs.{old_amount} to Rs.{amount} for {booking.user.get_full_name()}, room {booking.room.room_no} bed {booking.share_no}",
        meta={
            'old_amount': str(old_amount),
            'new_amount': str(amount),
            'tenant': booking.user.get_full_name()
        }
    )
    
    messages.success(request, f"Advance amount updated from Rs.{old_amount} to Rs.{amount}.")
    return JsonResponse({
        'success': True,
        'old_amount': str(old_amount),
        'new_amount': str(amount)
    })


# ============================================================================
# RE-CONTINUE FEATURE
# ============================================================================

def re_continue_booking(request, booking_id):
    """Allow user to re-continue (cancel leaving) before or on leaving date"""
    from bookings.models import RoomSwap
    
    # Check authentication - return JSON for AJAX requests
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'error': 'Please login to continue'}, status=401)
    
    try:
        booking = Booking.objects.select_related('room', 'room__pg').get(id=booking_id)
    except Booking.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Booking not found'}, status=404)
    
    if not _require_pg_admin(request.user) or not _admin_pgs(request.user).filter(id=booking.room.pg.id).exists():
        return JsonResponse({'success': False, 'error': 'Unauthorized'}, status=403)
    
    # Can only re-continue if leaving_confirmed_date exists
    if not booking.leaving_confirmed_date:
        return JsonResponse({'success': False, 'error': 'No confirmed leaving date'}, status=400)
    
    # Note: Re-continue is now allowed even after leaving date has passed
    # This allows PG admins to bring back tenants who left but want to return
    
    if request.method == 'GET':
        # First, sync bed statuses to ensure we have accurate data
        from bookings.utils import sync_room_share_statuses
        try:
            sync_room_share_statuses(pg=booking.room.pg)
        except Exception as e:
            # Log error but continue - we'll still check for conflicts
            pass
        
        # Show options: same room or change room
        # Check for conflicts in same room
        same_room_conflicts = []
        
        # Use today's date if leaving date has already passed, otherwise use leaving_confirmed_date
        reference_date = max(booking.leaving_confirmed_date, date.today())
        
        # CRITICAL: Check if the bed is currently occupied by another user
        # This handles the case where User A left and User B moved in
        current_share_status = RoomShareStatus.objects.filter(
            room=booking.room,
            share_no=booking.share_no
        ).first()
        
        if current_share_status and current_share_status.status == RoomShareStatus.OCCUPIED:
            # Bed is occupied - check if it's occupied by a different user
            current_occupant = Booking.objects.filter(
                room=booking.room,
                share_no=booking.share_no,
                status=Booking.APPROVED,
                joining_date__lte=date.today()
            ).exclude(id=booking.id).exclude(
                # Exclude bookings that have left (have confirmed leaving date in the past)
                leaving_confirmed_date__lt=date.today()
            ).select_related('user').first()
            
            if current_occupant:
                same_room_conflicts.append({
                    'type': 'current_occupant',
                    'detail': f"Bed is currently occupied by {current_occupant.user.get_full_name()} (since {current_occupant.joining_date})"
                })
        
        # Check for active APPROVED bookings on this bed (users who haven't left yet)
        active_bookings = Booking.objects.filter(
            room=booking.room,
            share_no=booking.share_no,
            status=Booking.APPROVED,
        ).exclude(id=booking.id).exclude(
            # Exclude bookings that have confirmed leaving date in the past
            leaving_confirmed_date__lt=date.today()
        ).select_related('user')
        
        for ab in active_bookings:
            # Check if this booking represents a current/active occupancy
            if ab.joining_date <= date.today():
                conflict_already_added = any(
                    c.get('type') == 'current_occupant' and ab.user.get_full_name() in c.get('detail', '')
                    for c in same_room_conflicts
                )
                if not conflict_already_added:
                    same_room_conflicts.append({
                        'type': 'active_booking',
                        'detail': f"Active booking for {ab.user.get_full_name()} (joined {ab.joining_date})"
                    })
        
        # Check future bookings on this bed (PENDING or APPROVED with future joining date)
        future_bookings = Booking.objects.filter(
            room=booking.room,
            share_no=booking.share_no,
            status__in=[Booking.PENDING, Booking.APPROVED],
            joining_date__gte=reference_date
        ).exclude(id=booking.id).select_related('user')
        
        if future_bookings.exists():
            for fb in future_bookings:
                same_room_conflicts.append({
                    'type': 'booking',
                    'detail': f"Booking for {fb.user.get_full_name()} from {fb.joining_date}"
                })
        
        # Check future swaps (both PENDING and APPROVED)
        future_swaps = RoomSwap.objects.filter(
            to_room=booking.room,
            to_share_no=booking.share_no,
            effective_date__gte=reference_date,
            status__in=[RoomSwap.PENDING, RoomSwap.APPROVED],
            is_future_swap=True
        ).select_related('booking', 'booking__user')
        
        if future_swaps.exists():
            for fs in future_swaps:
                swap_status = "Pending" if fs.status == RoomSwap.PENDING else "Approved"
                same_room_conflicts.append({
                    'type': 'swap',
                    'detail': f"Future Swap ({swap_status}) for {fs.booking.user.get_full_name()} effective {fs.effective_date}"
                })
        
        # Get available rooms for change option
        # First, collect beds that have pending/approved future swaps
        beds_with_future_swaps = set()
        all_future_swaps = RoomSwap.objects.filter(
            to_room__pg=booking.room.pg,
            effective_date__gte=date.today(),
            status__in=[RoomSwap.PENDING, RoomSwap.APPROVED],
            is_future_swap=True
        ).values_list('to_room_id', 'to_share_no')
        for room_id, share_no in all_future_swaps:
            beds_with_future_swaps.add((room_id, share_no))
        
        vacant_rooms = []
        all_rooms = Room.objects.filter(pg=booking.room.pg).prefetch_related('shares')
        pending_booking_keys = pending_booking_share_keys(pg=booking.room.pg)
        for room in all_rooms:
            for share in room.shares.all():
                if share.status == RoomShareStatus.VACANT:
                    if (room.id, share.share_no) in pending_booking_keys:
                        continue
                    # Skip beds that have pending/approved future swaps
                    if (room.id, share.share_no) in beds_with_future_swaps:
                        continue
                    vacant_rooms.append({
                        'room_id': room.id,
                        'room_no': room.room_no,
                        'share_no': share.share_no
                    })
        
        return JsonResponse({
            'success': True,
            'same_room_available': len(same_room_conflicts) == 0,
            'same_room_conflicts': same_room_conflicts,
            'vacant_rooms': vacant_rooms,
            'current_room': booking.room.room_no,
            'current_share': booking.share_no
        })
    
    elif request.method == 'POST':
        option = request.POST.get('option')  # 'same' or 'change'
        
        # First, sync bed statuses to ensure we have accurate data
        from bookings.utils import sync_room_share_statuses
        try:
            sync_room_share_statuses(pg=booking.room.pg)
        except Exception as e:
            # Log error but continue
            pass
        
        # Use today's date if leaving date has already passed, otherwise use leaving_confirmed_date
        reference_date = max(booking.leaving_confirmed_date, date.today())
        
        if option == 'same':
            # CRITICAL: First check if bed is currently occupied by another user
            current_share_status = RoomShareStatus.objects.filter(
                room=booking.room,
                share_no=booking.share_no
            ).first()
            
            if current_share_status and current_share_status.status == RoomShareStatus.OCCUPIED:
                # Check if it's occupied by a different active user
                current_occupant = Booking.objects.filter(
                    room=booking.room,
                    share_no=booking.share_no,
                    status=Booking.APPROVED,
                    joining_date__lte=date.today()
                ).exclude(id=booking.id).exclude(
                    leaving_confirmed_date__lt=date.today()
                ).first()
                
                if current_occupant:
                    return JsonResponse({
                        'error': f'Bed is currently occupied by {current_occupant.user.get_full_name()}. Please select a different room.'
                    }, status=400)
            
            # Validate no conflicts (including both PENDING and APPROVED future swaps)
            conflicts_exist = (
                Booking.objects.filter(
                    room=booking.room,
                    share_no=booking.share_no,
                    status__in=[Booking.PENDING, Booking.APPROVED],
                    joining_date__gte=reference_date
                ).exclude(id=booking.id).exists() or
                RoomSwap.objects.filter(
                    to_room=booking.room,
                    to_share_no=booking.share_no,
                    effective_date__gte=reference_date,
                    status__in=[RoomSwap.PENDING, RoomSwap.APPROVED],
                    is_future_swap=True
                ).exists()
            )
            
            if conflicts_exist:
                return JsonResponse({'error': 'Conflicts exist - cannot continue in same room'}, status=400)
            
            # Clear leaving dates and restore status to APPROVED if needed
            booking.leaving_date = None
            booking.leaving_confirmed_date = None
            booking.leaving_reason = ''
            booking.leaving_initiated_at = None
            # Restore status to APPROVED if it was COMPLETED
            if booking.status == Booking.COMPLETED:
                booking.status = Booking.APPROVED
                booking.save(update_fields=[
                    'leaving_date', 'leaving_confirmed_date', 'leaving_reason', 'leaving_initiated_at', 'status'
                ])
            else:
                booking.save(update_fields=[
                    'leaving_date', 'leaving_confirmed_date', 'leaving_reason', 'leaving_initiated_at'
                ])
            
            # Update room share status back to OCCUPIED
            share = RoomShareStatus.objects.filter(room=booking.room, share_no=booking.share_no).first()
            if share:
                share.status = RoomShareStatus.OCCUPIED
                share.vacant_from = None
                share.save(update_fields=['status', 'vacant_from'])
            
            # Notify user
            Notification.objects.create(
                user=booking.user,
                title="Re-Continue Approved",
                message=f"Your request to continue staying in Room {booking.room.room_no}, Bed {booking.share_no} has been approved."
            )
            send_push_to_user(
                booking.user,
                title="Re-Continue Approved",
                body=f"Continue approved for Room {booking.room.room_no}, Bed {booking.share_no}.",
                url=reverse('booking_detail', args=[booking.id]),
                extra_data={'type': 're_continue_approved', 'booking_id': booking.id},
            )
            
            # Remove from Old Tenants if exists (they're coming back)
            try:
                from .models import OldTenant
                OldTenant.objects.filter(
                    pg=booking.room.pg,
                    original_booking_id=booking.id
                ).delete()
            except Exception:
                pass  # Ignore errors - it's okay if record doesn't exist
            
            # Audit log
            log(
                actor=request.user,
                action='re_continue_same_room',
                target_type='Booking',
                target_id=booking.id,
                message=f"Re-continue approved for {booking.user.get_full_name()} in same room",
                meta={'tenant': booking.user.get_full_name()}
            )
            
            messages.success(request, f"{booking.user.get_full_name()} will continue in same room.")
            return JsonResponse({'success': True, 'message': 'Re-continue approved - same room'})
        
        elif option == 'change':
            new_room_id = request.POST.get('new_room_id')
            new_share_no = request.POST.get('new_share_no')
            
            try:
                new_room = Room.objects.get(id=new_room_id, pg=booking.room.pg)
                new_share = RoomShareStatus.objects.get(room=new_room, share_no=new_share_no)
            except:
                return JsonResponse({'error': 'Invalid room/share selection'}, status=400)
            
            if new_share.status != RoomShareStatus.VACANT:
                return JsonResponse({'error': 'Selected bed is not vacant'}, status=400)
            
            # Free old room/share
            old_share = RoomShareStatus.objects.filter(room=booking.room, share_no=booking.share_no).first()
            if old_share:
                old_share.status = RoomShareStatus.VACANT
                old_share.vacant_from = None
                old_share.save(update_fields=['status', 'vacant_from'])
            
            # Update booking to new room/share
            old_room_no = booking.room.room_no
            old_share_no = booking.share_no
            
            booking.room = new_room
            booking.share_no = new_share_no
            booking.leaving_date = None
            booking.leaving_confirmed_date = None
            booking.leaving_reason = ''
            booking.leaving_initiated_at = None
            # Restore status to APPROVED if it was COMPLETED
            if booking.status == Booking.COMPLETED:
                booking.status = Booking.APPROVED
                booking.save(update_fields=[
                    'room', 'share_no', 'leaving_date', 'leaving_confirmed_date', 
                    'leaving_reason', 'leaving_initiated_at', 'status'
                ])
            else:
                booking.save(update_fields=[
                    'room', 'share_no', 'leaving_date', 'leaving_confirmed_date', 
                    'leaving_reason', 'leaving_initiated_at'
                ])
            
            # Mark new share as occupied
            new_share.status = RoomShareStatus.OCCUPIED
            new_share.save(update_fields=['status'])
            
            # Remove from Old Tenants if exists (they're coming back)
            try:
                from .models import OldTenant
                OldTenant.objects.filter(
                    pg=booking.room.pg,
                    original_booking_id=booking.id
                ).delete()
            except Exception:
                pass  # Ignore errors - it's okay if record doesn't exist
            
            # Notify user
            Notification.objects.create(
                user=booking.user,
                title="Re-Continue with Room Change",
                message=f"Your request to continue has been approved. You have been moved from Room {old_room_no}, Bed {old_share_no} to Room {new_room.room_no}, Bed {new_share_no}."
            )
            send_push_to_user(
                booking.user,
                title="Room Changed",
                body=f"Moved to Room {new_room.room_no}, Bed {new_share_no}.",
                url=reverse('booking_detail', args=[booking.id]),
                extra_data={'type': 're_continue_room_change', 'booking_id': booking.id},
            )
            
            # Audit log
            log(
                actor=request.user,
                action='re_continue_change_room',
                target_type='Booking',
                target_id=booking.id,
                message=f"Re-continue with room change for {booking.user.get_full_name()}: {old_room_no}/{old_share_no} → {new_room.room_no}/{new_share_no}",
                meta={
                    'tenant': booking.user.get_full_name(),
                    'old_room': f"{old_room_no}/{old_share_no}",
                    'new_room': f"{new_room.room_no}/{new_share_no}"
                }
            )
            
            messages.success(request, f"{booking.user.get_full_name()} moved to Room {new_room.room_no}, Bed {new_share_no}.")
            return JsonResponse({'success': True, 'message': 'Re-continue approved - room changed'})
        
        return JsonResponse({'error': 'Invalid option'}, status=400)


# ============================================================================
# FUTURE SWAP FEATURE
# ============================================================================

@login_required
def create_future_swap(request, booking_id):
    """Create future swap request based on confirmed leaving date"""
    from bookings.models import RoomSwap
    
    booking = get_object_or_404(
        Booking.objects.select_related('room', 'room__pg', 'user'),
        id=booking_id
    )
    
    if not _require_pg_admin(request.user) or not _admin_pgs(request.user).filter(id=booking.room.pg.id).exists():
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    if request.method == 'GET':
        # Get available target beds (vacant or with confirmed leaving)
        available_targets = []
        today = date.today()
        
        # Get all rooms in PG
        all_rooms = Room.objects.filter(pg=booking.room.pg).prefetch_related('shares')
        pending_booking_keys = pending_booking_share_keys(pg=booking.room.pg)
        
        for room in all_rooms:
            for share in room.shares.all():
                # Skip current bed
                if room.id == booking.room.id and share.share_no == booking.share_no:
                    continue
                if (room.id, share.share_no) in pending_booking_keys:
                    continue
                
                # Check if vacant
                if share.status == RoomShareStatus.VACANT:
                    available_targets.append({
                        'room_id': room.id,
                        'room_no': room.room_no,
                        'share_no': share.share_no,
                        'available_from': 'Now',
                        'status': 'vacant'
                    })
                # Check if has confirmed leaving
                elif share.status in [RoomShareStatus.OCCUPIED, RoomShareStatus.VACANT_FROM]:
                    occupant_booking = Booking.objects.filter(
                        room=room,
                        share_no=share.share_no,
                        status=Booking.APPROVED,
                        leaving_confirmed_date__isnull=False
                    ).first()
                    
                    if occupant_booking and occupant_booking.leaving_confirmed_date >= today:
                        available_targets.append({
                            'room_id': room.id,
                            'room_no': room.room_no,
                            'share_no': share.share_no,
                            'available_from': occupant_booking.leaving_confirmed_date.isoformat(),
                            'status': 'leaving',
                            'leaving_user': occupant_booking.user.get_full_name()
                        })
        
        return JsonResponse({
            'success': True,
            'available_targets': available_targets,
            'current_room': booking.room.room_no,
            'current_share': booking.share_no
        })
    
    elif request.method == 'POST':
        to_room_id = request.POST.get('to_room_id')
        to_share_no = request.POST.get('to_share_no')
        effective_date_str = request.POST.get('effective_date')
        reason = request.POST.get('reason', '')
        
        try:
            to_room = Room.objects.get(id=to_room_id, pg=booking.room.pg)
            to_share = RoomShareStatus.objects.get(room=to_room, share_no=to_share_no)
            effective_date = parse_date(effective_date_str)
            
            if not effective_date:
                raise ValueError("Invalid date")
        except:
            return JsonResponse({'error': 'Invalid swap parameters'}, status=400)
        
        if effective_date < date.today():
            return JsonResponse({'error': 'Effective date cannot be in the past'}, status=400)
        
        # Validate target bed availability
        # Check if bed will be vacant by effective_date
        target_booking = Booking.objects.filter(
            room=to_room,
            share_no=to_share_no,
            status=Booking.APPROVED
        ).first()
        
        if target_booking:
            if not target_booking.leaving_confirmed_date:
                return JsonResponse({'error': 'Target bed has no confirmed leaving date'}, status=400)
            if target_booking.leaving_confirmed_date > effective_date:
                return JsonResponse({
                    'error': f'Target bed will be available only from {target_booking.leaving_confirmed_date}'
                }, status=400)
        
        # Check for overlapping swaps
        overlapping_swaps = RoomSwap.objects.filter(
            Q(booking=booking) | Q(to_room=to_room, to_share_no=to_share_no),
            effective_date=effective_date,
            status__in=[RoomSwap.PENDING, RoomSwap.APPROVED]
        )
        
        if overlapping_swaps.exists():
            return JsonResponse({'error': 'Overlapping swap exists'}, status=400)
        
        # Create swap request
        swap = RoomSwap.objects.create(
            booking=booking,
            from_room=booking.room,
            from_share_no=booking.share_no,
            to_room=to_room,
            to_share_no=to_share_no,
            effective_date=effective_date,
            is_future_swap=True,
            reason=reason,
            status=RoomSwap.PENDING
        )
        
        # Notify user
        Notification.objects.create(
            user=booking.user,
            title="Room Swap Scheduled",
            message=f"A future room swap has been scheduled for you from Room {booking.room.room_no}/Bed {booking.share_no} to Room {to_room.room_no}/Bed {to_share_no}, effective {effective_date}. Pending approval."
        )
        send_push_to_user(
            booking.user,
            title="Room Swap Scheduled",
            body=f"Swap to Room {to_room.room_no}, Bed {to_share_no} on {effective_date} is pending approval.",
            url=reverse('booking_detail', args=[booking.id]),
            extra_data={'type': 'future_swap_scheduled', 'swap_id': swap.id, 'booking_id': booking.id},
        )
        
        # Audit log
        log(
            actor=request.user,
            action='future_swap_created',
            target_type='RoomSwap',
            target_id=swap.id,
            message=f"Future swap created: {booking.room.room_no}/{booking.share_no} → {to_room.room_no}/{to_share_no} on {effective_date}",
            meta={
                'booking_id': booking.id,
                'from': f"{booking.room.room_no}/{booking.share_no}",
                'to': f"{to_room.room_no}/{to_share_no}",
                'effective_date': effective_date.isoformat()
            }
        )
        
        messages.success(request, f"Future swap created for {booking.user.get_full_name()}.")
        return JsonResponse({'success': True, 'swap_id': swap.id, 'trigger_sync': True})


@login_required
def approve_future_swap(request, swap_id):
    """Approve a future swap request"""
    from bookings.models import RoomSwap
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    swap = get_object_or_404(
        RoomSwap.objects.select_related('booking', 'booking__room', 'booking__room__pg', 'from_room', 'to_room'),
        id=swap_id
    )
    
    if not _require_pg_admin(request.user) or not _admin_pgs(request.user).filter(id=swap.booking.room.pg.id).exists():
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    if swap.status != RoomSwap.PENDING:
        return JsonResponse({'error': 'Swap is not pending'}, status=400)
    
    # Approve swap
    swap.status = RoomSwap.APPROVED
    swap.processed_at = timezone.now()
    swap.processed_by = request.user
    swap.save(update_fields=['status', 'processed_at', 'processed_by'])
    
    # Notify user
    Notification.objects.create(
        user=swap.booking.user,
        title="Room Swap Approved",
        message=f"Your room swap from {swap.from_room.room_no}/Bed {swap.from_share_no} to {swap.to_room.room_no}/Bed {swap.to_share_no} effective {swap.effective_date} has been approved."
    )
    send_push_to_user(
        swap.booking.user,
        title="Room Swap Approved",
        body=f"Swap to Room {swap.to_room.room_no}, Bed {swap.to_share_no} is approved.",
        url=reverse('booking_detail', args=[swap.booking_id]),
        extra_data={'type': 'future_swap_approved', 'swap_id': swap.id, 'booking_id': swap.booking_id},
    )
    
    # Audit log
    log(
        actor=request.user,
        action='future_swap_approved',
        target_type='RoomSwap',
        target_id=swap.id,
        message=f"Future swap approved for booking {swap.booking.id}",
        meta={'booking_id': swap.booking.id}
    )
    
    messages.success(request, "Future swap approved.")
    return JsonResponse({'success': True, 'trigger_sync': True})


@login_required
def reject_future_swap(request, swap_id):
    """Reject a future swap request"""
    from bookings.models import RoomSwap
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    swap = get_object_or_404(
        RoomSwap.objects.select_related('booking', 'booking__room', 'booking__room__pg'),
        id=swap_id
    )
    
    if not _require_pg_admin(request.user) or not _admin_pgs(request.user).filter(id=swap.booking.room.pg.id).exists():
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    if swap.status != RoomSwap.PENDING:
        return JsonResponse({'error': 'Swap is not pending'}, status=400)
    
    # Reject swap
    swap.status = RoomSwap.REJECTED
    swap.processed_at = timezone.now()
    swap.processed_by = request.user
    swap.save(update_fields=['status', 'processed_at', 'processed_by'])
    
    # Notify user
    Notification.objects.create(
        user=swap.booking.user,
        title="Room Swap Rejected",
        message=f"Your room swap request has been rejected. Please contact admin for details."
    )
    send_push_to_user(
        swap.booking.user,
        title="Room Swap Rejected",
        body="Your room swap request was rejected.",
        url=reverse('booking_detail', args=[swap.booking_id]),
        extra_data={'type': 'future_swap_rejected', 'swap_id': swap.id, 'booking_id': swap.booking_id},
    )
    
    # Audit log
    log(
        actor=request.user,
        action='future_swap_rejected',
        target_type='RoomSwap',
        target_id=swap.id,
        message=f"Future swap rejected for booking {swap.booking.id}",
        meta={'booking_id': swap.booking.id}
    )
    
    messages.success(request, "Future swap rejected.")
    return JsonResponse({'success': True})


@login_required
def execute_swap(request, swap_id):
    """Manually execute/complete a swap (normally done automatically on effective date)"""
    from bookings.models import RoomSwap
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    swap = get_object_or_404(
        RoomSwap.objects.select_related('booking', 'booking__room', 'from_room', 'to_room'),
        id=swap_id
    )
    
    if not _require_pg_admin(request.user) or not _admin_pgs(request.user).filter(id=swap.booking.room.pg.id).exists():
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    if swap.status != RoomSwap.APPROVED:
        return JsonResponse({'error': 'Swap is not approved'}, status=400)
    
    # Execute swap
    with transaction.atomic():
        # Update booking room/share
        old_share = RoomShareStatus.objects.filter(room=swap.from_room, share_no=swap.from_share_no).first()
        if old_share:
            old_share.status = RoomShareStatus.VACANT
            old_share.save(update_fields=['status'])
        
        new_share = RoomShareStatus.objects.filter(room=swap.to_room, share_no=swap.to_share_no).first()
        if new_share:
            new_share.status = RoomShareStatus.OCCUPIED
            new_share.vacant_from = None
            new_share.save(update_fields=['status', 'vacant_from'])
        
        swap.booking.room = swap.to_room
        swap.booking.share_no = swap.to_share_no
        swap.booking.save(update_fields=['room', 'share_no'])
        
        swap.status = RoomSwap.COMPLETED
        swap.save(update_fields=['status'])
    
    # Notify user
    Notification.objects.create(
        user=swap.booking.user,
        title="Room Swap Completed",
        message=f"Your room swap is complete. You are now in Room {swap.to_room.room_no}, Bed {swap.to_share_no}."
    )
    send_push_to_user(
        swap.booking.user,
        title="Room Swap Completed",
        body=f"You are now in Room {swap.to_room.room_no}, Bed {swap.to_share_no}.",
        url=reverse('booking_detail', args=[swap.booking_id]),
        extra_data={'type': 'future_swap_completed', 'swap_id': swap.id, 'booking_id': swap.booking_id},
    )
    
    # Audit log
    log(
        actor=request.user,
        action='swap_executed',
        target_type='RoomSwap',
        target_id=swap.id,
        message=f"Swap executed for booking {swap.booking.id}",
        meta={'booking_id': swap.booking.id}
    )
    
    messages.success(request, "Swap executed successfully.")
    return JsonResponse({'success': True})


@login_required
def sync_bed_statuses(request):
    """
    Sync RoomShareStatus based on actual Booking data for the active PG.
    Also executes any pending future swaps that are due (effective_date <= today).
    Triggered by the Refresh button in Beds Overview or Tenants page, or after swap submit.
    
    Handles chain swaps properly:
    - Example: user0 leaves room 201/bed 1, user1 -> 201/bed 1, user2 -> user1's old bed, etc.
    - Executes swaps in correct dependency order
    - Updates bed statuses to reflect final state after all swaps
    """
    if not _require_pg_admin(request.user):
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'ok': False, 'error': 'Access denied.'}, status=403)
        messages.error(request, "Access denied.")
        return redirect('dashboard')
    
    pg = _active_pg(request)
    if not pg:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'ok': False, 'error': 'No PG selected.'}, status=400)
        messages.error(request, "No PG selected.")
        return redirect('dashboard')
    
    today = timezone.now().date()
    
    # First, execute any pending future swaps for this PG that are due
    pending_swaps = RoomSwap.objects.filter(
        status=RoomSwap.PENDING,
        is_future_swap=True,
        effective_date__lte=today,
        booking__room__pg=pg
    ).select_related('booking', 'from_room', 'to_room', 'booking__user').order_by('effective_date', 'requested_at')
    
    # Order swaps to handle chains correctly
    ordered_swaps = _get_swap_chain_order(pending_swaps)
    
    swap_results = {
        'executed': 0,
        'failed': 0,
        'messages': []
    }
    
    for swap in ordered_swaps:
        try:
            result = _execute_future_swap(swap, request.user)
            if result['success']:
                swap_results['executed'] += 1
                swap_results['messages'].append(
                    f"✓ Executed swap for {swap.booking.user.get_full_name() or swap.booking.user.email}: "
                    f"Room {swap.from_room.room_no} Bed {swap.from_share_no} → "
                    f"Room {swap.to_room.room_no} Bed {swap.to_share_no} "
                    f"(scheduled for {swap.effective_date.strftime('%Y-%m-%d')})"
                )
            else:
                swap_results['failed'] += 1
                swap_results['messages'].append(
                    f"✗ Failed swap for {swap.booking.user.get_full_name() or swap.booking.user.email}: {result['error']}"
                )
        except Exception as e:
            swap_results['failed'] += 1
            swap_results['messages'].append(
                f"✗ Error executing swap #{swap.id}: {str(e)}"
            )
    
    # Then, sync bed statuses (now includes future swap adjustments automatically)
    from bookings.utils import sync_room_share_statuses
    
    try:
        stats = sync_room_share_statuses(pg=pg)
        
        success_msg = (
            f"Bed statuses synced successfully! "
            f"Processed {stats['total_processed']} beds: "
            f"{stats['vacant']} vacant, {stats['reserved']} reserved, "
            f"{stats['occupied']} occupied, {stats['vacant_from']} leaving."
        )
        
        if stats.get('bookings_completed', 0) > 0:
            success_msg += f" | {stats['bookings_completed']} past bookings marked completed."
        
        if swap_results['executed'] > 0 or swap_results['failed'] > 0:
            success_msg += f" | Future Swaps: {swap_results['executed']} executed, {swap_results['failed']} failed."
        
        # For AJAX requests, return JSON
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                'ok': True,
                'message': success_msg,
                'stats': stats,
                'swaps_executed': swap_results['executed'],
                'swaps_failed': swap_results['failed'],
                'swap_messages': swap_results['messages']
            })
        
        messages.success(request, success_msg)
        
        # Show detailed swap messages if any
        for msg in swap_results['messages']:
            if msg.startswith('✓'):
                messages.success(request, msg)
            else:
                messages.warning(request, msg)
                
    except Exception as e:
        error_msg = f"Error syncing bed statuses: {str(e)}"
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'ok': False, 'error': error_msg}, status=500)
        messages.error(request, error_msg)
    
    # If caller provided a 'next' indicator (e.g. from tenants page), honor it
    next_target = request.POST.get('next')
    if next_target == 'pg_tenants':
        return redirect('pg_tenants')
    return redirect('dashboard')


def _execute_future_swap(swap: RoomSwap, executor) -> dict:
    """
    Execute a pending future swap.
    Returns: {'success': bool, 'error': str or None}
    
    This function performs the actual room swap that was scheduled for a future date.
    It uses the effective_date from the swap record for logging, even if executed later.
    """
    try:
        with transaction.atomic():
            # Re-fetch with lock to ensure consistency
            swap = RoomSwap.objects.select_for_update().get(pk=swap.id)
            
            # Verify swap is still pending
            if swap.status != RoomSwap.PENDING:
                return {'success': False, 'error': f'Swap status is {swap.status}, not PENDING'}
            
            booking = Booking.objects.select_for_update().get(pk=swap.booking_id, status=Booking.APPROVED)
            from_room = Room.objects.select_for_update().get(pk=swap.from_room_id)
            to_room = Room.objects.select_for_update().get(pk=swap.to_room_id)
            from_share = RoomShareStatus.objects.select_for_update().get(room=from_room, share_no=swap.from_share_no)
            to_share = RoomShareStatus.objects.select_for_update().get(room=to_room, share_no=swap.to_share_no)
            
            # Verify booking is still in the from_room (user hasn't moved elsewhere)
            if booking.room_id != from_room.id or booking.share_no != swap.from_share_no:
                swap.status = RoomSwap.CANCELLED
                swap.reason += f" | Auto-cancelled: booking no longer at source location (now at room {booking.room.room_no} bed {booking.share_no})"
                swap.processed_at = timezone.now()
                swap.save(update_fields=['status', 'reason', 'processed_at'])
                return {'success': False, 'error': 'Booking has moved from original location'}
            
            today = timezone.now().date()
            
            # Check if target bed is available
            # We need to verify by checking actual bookings, not just the status flag
            # Allow if there's a leaving booking with leaving_date <= today (they've left/leaving)
            blocking_booking = Booking.objects.filter(
                room=to_room,
                share_no=swap.to_share_no,
                status=Booking.APPROVED,
                joining_date__lte=today
            ).filter(
                Q(leaving_date__isnull=True) | Q(leaving_date__gt=today)
            ).exclude(pk=booking.pk).first()
            
            if blocking_booking:
                # There's still someone in this bed - cannot execute
                swap.status = RoomSwap.CANCELLED
                blocker_name = blocking_booking.user.get_full_name() or blocking_booking.user.email
                swap.reason += f" | Auto-cancelled: target bed is occupied by {blocker_name}"
                swap.processed_at = timezone.now()
                swap.save(update_fields=['status', 'reason', 'processed_at'])
                return {'success': False, 'error': f'Target bed is occupied by {blocker_name}'}
            
            # Mark any leaving booking at the target bed as COMPLETED
            # (booking with leaving_date <= today that hasn't been marked completed yet)
            leaving_bookings = Booking.objects.filter(
                room=to_room,
                share_no=swap.to_share_no,
                status=Booking.APPROVED,
                leaving_date__lte=today
            ).exclude(pk=booking.pk)
            
            for leaving_booking in leaving_bookings:
                leaving_booking.status = Booking.COMPLETED
                leaving_booking.save(update_fields=['status'])
            
            # Execute the swap
            # Update booking to new room
            booking.room = to_room
            booking.share_no = swap.to_share_no
            try:
                booking.pg = to_room.pg
            except Exception:
                pass
            booking.save(update_fields=['room', 'pg', 'share_no'])
            
            # Update application if exists
            app = getattr(booking, 'application', None)
            if app and app.room_id != to_room.id:
                app.room = to_room
                app.save(update_fields=['room'])
            
            # Update share statuses
            from_share.status = RoomShareStatus.VACANT
            from_share.vacant_from = None
            from_share.save(update_fields=['status', 'vacant_from'])
            
            to_share.status = RoomShareStatus.OCCUPIED
            to_share.vacant_from = None
            to_share.save(update_fields=['status', 'vacant_from'])
            
            # Mark swap as completed
            swap.status = RoomSwap.COMPLETED
            swap.processed_at = timezone.now()
            swap.processed_by = executor
            # NOTE: effective_date remains as originally scheduled - important for billing/logs
            swap.save(update_fields=['status', 'processed_at', 'processed_by'])
            
            # Log the execution
            log(
                executor,
                'future_swap_executed',
                'RoomSwap',
                swap.id,
                f"Future swap executed: {booking.user.get_full_name() or booking.user.email} moved from room {from_room.room_no} bed {swap.from_share_no} to room {to_room.room_no} bed {swap.to_share_no}. Originally scheduled for {swap.effective_date.strftime('%Y-%m-%d')}, executed on {timezone.now().date().strftime('%Y-%m-%d')}."
            )
            
            return {'success': True, 'error': None}
            
    except Booking.DoesNotExist:
        return {'success': False, 'error': 'Booking no longer exists'}
    except Room.DoesNotExist:
        return {'success': False, 'error': 'Room no longer exists'}
    except RoomShareStatus.DoesNotExist:
        return {'success': False, 'error': 'Room share no longer exists'}
    except Exception as e:
        return {'success': False, 'error': str(e)}


@login_required
def future_swaps(request):
    """
    List and manage all pending future swaps for the admin's PG(s).
    Allows canceling or modifying scheduled swaps.
    """
    if not _require_pg_admin(request.user):
        messages.error(request, "Access denied.")
        return redirect('dashboard')
    
    pg = _active_pg(request)
    if not pg:
        messages.error(request, "No PG selected.")
        return redirect('dashboard')
    
    today = timezone.now().date()
    
    # Get all future swaps for this PG
    pending_swaps = RoomSwap.objects.filter(
        booking__room__pg=pg,
        is_future_swap=True,
        status__in=[RoomSwap.PENDING, RoomSwap.APPROVED]
    ).select_related(
        'booking',
        'booking__user',
        'booking__user__profile',
        'from_room',
        'to_room',
        'processed_by'
    ).order_by('effective_date', 'requested_at')
    
    # Separate into due and upcoming
    due_swaps = []
    upcoming_swaps = []
    
    for swap in pending_swaps:
        if swap.effective_date <= today:
            due_swaps.append(swap)
        else:
            upcoming_swaps.append(swap)
    
    # Get completed future swaps for history (last 30 days)
    from datetime import timedelta
    thirty_days_ago = today - timedelta(days=30)
    
    completed_swaps = RoomSwap.objects.filter(
        booking__room__pg=pg,
        is_future_swap=True,
        status=RoomSwap.COMPLETED,
        processed_at__gte=thirty_days_ago
    ).select_related(
        'booking',
        'booking__user',
        'from_room',
        'to_room',
        'processed_by'
    ).order_by('-processed_at')[:20]
    
    context = {
        'pg': pg,
        'pgs': list(_admin_pgs(request.user)),
        'due_swaps': due_swaps,
        'upcoming_swaps': upcoming_swaps,
        'completed_swaps': completed_swaps,
        'today': today,
    }
    
    return render(request, 'pgadmin/future_swaps.html', context)


@login_required
@transaction.atomic
def cancel_future_swap(request, swap_id):
    """
    Cancel a pending future swap.
    
    Handles chain swap dependencies:
    - If this swap's destination bed (to_room/to_share) is the source for another swap,
      that dependent swap must be cancelled first (or together).
    - Shows warning to user about dependent swaps and proper cancellation order.
    """
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.POST.get('ajax') == '1'
    
    if not _require_pg_admin(request.user):
        if is_ajax:
            return JsonResponse({'ok': False, 'error': 'Access denied.'}, status=403)
        messages.error(request, "Access denied.")
        return redirect('pg_future_swaps')
    
    swap = get_object_or_404(
        RoomSwap.objects.select_for_update(),
        pk=swap_id,
        is_future_swap=True,
        status__in=[RoomSwap.PENDING, RoomSwap.APPROVED]
    )
    
    # Verify admin has access to this PG
    pg_id = getattr(swap.booking.room, 'pg_id', None)
    if not _admin_pgs(request.user).filter(id=pg_id).exists():
        if is_ajax:
            return JsonResponse({'ok': False, 'error': 'Access denied for this PG.'}, status=403)
        messages.error(request, "Access denied for this PG.")
        return redirect('pg_future_swaps')
    
    # Check for force_cancel parameter (to cancel with all dependents)
    force_cancel = request.POST.get('force_cancel') == '1'
    
    # Find dependent swaps (swaps that depend on this swap completing)
    # A swap S depends on swap T if S.to_room/to_share == T.from_room/from_share
    # (S needs to move TO the bed that T is vacating)
    dependent_swaps = _find_dependent_swaps(swap)
    
    if dependent_swaps and not force_cancel:
        # There are dependent swaps - warn the user
        dependent_info = []
        for dep_swap in dependent_swaps:
            dep_user = dep_swap.booking.user.get_full_name() or dep_swap.booking.user.email
            dependent_info.append({
                'id': dep_swap.id,
                'user': dep_user,
                'from': f"Room {dep_swap.from_room.room_no} Bed {dep_swap.from_share_no}",
                'to': f"Room {dep_swap.to_room.room_no} Bed {dep_swap.to_share_no}",
                'date': dep_swap.effective_date.strftime('%Y-%m-%d')
            })
        
        # Include the current swap being cancelled for complete chain display
        current_swap_user = swap.booking.user.get_full_name() or swap.booking.user.email
        current_swap_info = {
            'id': swap.id,
            'user': current_swap_user,
            'from': f"Room {swap.from_room.room_no} Bed {swap.from_share_no}",
            'to': f"Room {swap.to_room.room_no} Bed {swap.to_share_no}",
            'date': swap.effective_date.strftime('%Y-%m-%d')
        }
        
        warning_msg = _build_chain_cancel_warning(swap, dependent_swaps)
        
        if is_ajax:
            return JsonResponse({
                'ok': False,
                'action': 'chain_dependency_warning',
                'error': warning_msg,
                'warning': True,
                'current_swap': current_swap_info,
                'dependent_swaps': dependent_info,
                'swap_id': swap.id,
                'total_chain_length': len(dependent_swaps) + 1,
                'message': 'This swap has dependent chain swaps. Please review before cancelling.'
            }, status=400)
        
        messages.warning(request, warning_msg)
        return redirect('pg_future_swaps')
    
    # Store data for response
    booking_id = swap.booking_id
    room_id = swap.booking.room_id
    share_no = swap.booking.share_no
    
    cancelled_swaps = []
    
    # If force_cancel, cancel all dependent swaps first (in reverse dependency order)
    if force_cancel and dependent_swaps:
        # Cancel dependents in reverse order (most dependent first)
        for dep_swap in reversed(dependent_swaps):
            dep_swap.status = RoomSwap.CANCELLED
            dep_swap.reason += f" | Chain-cancelled due to parent swap #{swap.id} cancellation by {request.user.get_full_name() or request.user.email} on {timezone.now().date()}"
            dep_swap.processed_at = timezone.now()
            dep_swap.save(update_fields=['status', 'reason', 'processed_at'])
            
            log(
                request.user,
                'future_swap_chain_cancelled',
                'RoomSwap',
                dep_swap.id,
                f"Chain-cancelled future swap for {dep_swap.booking.user.get_full_name() or dep_swap.booking.user.email} (dependent on swap #{swap.id})"
            )
            Notification.objects.create(
                user=dep_swap.booking.user,
                title="Room Swap Cancelled",
                message=f"Your scheduled room swap (ID #{dep_swap.id}) was cancelled because a dependent parent swap was cancelled by PG admin."
            )
            send_push_to_user(
                dep_swap.booking.user,
                title="Room Swap Cancelled",
                body=f"Scheduled swap #{dep_swap.id} was cancelled by PG admin.",
                url=reverse('booking_detail', args=[dep_swap.booking_id]),
                extra_data={'type': 'future_swap_cancelled', 'swap_id': dep_swap.id, 'booking_id': dep_swap.booking_id},
            )
            cancelled_swaps.append({
                'id': dep_swap.id,
                'user': dep_swap.booking.user.get_full_name() or dep_swap.booking.user.email
            })
    
    # Cancel the main swap
    swap.status = RoomSwap.CANCELLED
    swap.reason += f" | Cancelled by {request.user.get_full_name() or request.user.email} on {timezone.now().date()}"
    swap.processed_at = timezone.now()
    swap.save(update_fields=['status', 'reason', 'processed_at'])
    
    log(
        request.user,
        'future_swap_cancelled',
        'RoomSwap',
        swap.id,
        f"Cancelled future swap for {swap.booking.user.get_full_name() or swap.booking.user.email} from room {swap.from_room.room_no} bed {swap.from_share_no} to room {swap.to_room.room_no} bed {swap.to_share_no}"
    )
    Notification.objects.create(
        user=swap.booking.user,
        title="Room Swap Cancelled",
        message=f"Your scheduled room swap from Room {swap.from_room.room_no}/Bed {swap.from_share_no} to Room {swap.to_room.room_no}/Bed {swap.to_share_no} has been cancelled by PG admin."
    )
    send_push_to_user(
        swap.booking.user,
        title="Room Swap Cancelled",
        body=f"Swap from Room {swap.from_room.room_no}/Bed {swap.from_share_no} to Room {swap.to_room.room_no}/Bed {swap.to_share_no} was cancelled.",
        url=reverse('booking_detail', args=[swap.booking_id]),
        extra_data={'type': 'future_swap_cancelled', 'swap_id': swap.id, 'booking_id': swap.booking_id},
    )
    
    # Trigger bed status sync after cancellation
    from bookings.utils import sync_room_share_statuses
    try:
        sync_room_share_statuses(pg=swap.booking.room.pg)
    except Exception as e:
        _logger.warning(f"Failed to sync bed statuses after swap cancellation: {e}")
    
    if is_ajax:
        message = 'Future swap cancelled successfully.'
        if cancelled_swaps:
            message = f'Cancelled {len(cancelled_swaps) + 1} chain swaps successfully.'
        
        return JsonResponse({
            'ok': True,
            'action': 'future_swap_cancelled',
            'message': message,
            'booking_id': booking_id,
            'room_id': room_id,
            'share_no': share_no,
            'cancelled_swaps': cancelled_swaps,
            'trigger_sync': True
        })
    
    # Regular redirect for future swaps admin page
    if cancelled_swaps:
        messages.success(request, f"Cancelled {len(cancelled_swaps) + 1} chain swaps successfully.")
    else:
        messages.success(request, f"Future swap cancelled successfully.")
    return redirect('pg_future_swaps')


def _find_dependent_swaps(swap):
    """
    Find all swaps that depend on the given swap completing.
    
    A swap S depends on swap T if:
    - S.to_room/to_share == T.from_room/from_share (S needs the bed T is moving FROM - the bed T will vacate)
    - S is still pending/approved
    
    Example: Bed 4 → Bed 1, Bed 2 → Bed 4
    - When cancelling "Bed 4 → Bed 1", we need to find who depends on Bed 4 becoming vacant
    - "Bed 2 → Bed 4" depends on "Bed 4 → Bed 1" because Bed 2 needs to move TO Bed 4 (which Bed 4's swap vacates)
    
    Returns swaps in dependency order (direct dependents first, then their dependents, etc.)
    """
    dependent_swaps = []
    visited = {swap.id}
    to_process = [swap]
    
    while to_process:
        current_swap = to_process.pop(0)
        
        # Find swaps that are moving TO the bed this swap is moving FROM
        # These swaps depend on the current swap vacating that bed
        direct_dependents = RoomSwap.objects.filter(
            to_room_id=current_swap.from_room_id,
            to_share_no=current_swap.from_share_no,
            status__in=[RoomSwap.PENDING, RoomSwap.APPROVED],
            is_future_swap=True
        ).exclude(pk__in=visited).select_related('booking', 'booking__user', 'from_room', 'to_room')
        
        for dep in direct_dependents:
            if dep.id not in visited:
                visited.add(dep.id)
                dependent_swaps.append(dep)
                to_process.append(dep)
    
    return dependent_swaps


def _build_chain_cancel_warning(swap, dependent_swaps):
    """Build a warning message explaining the chain dependency."""
    swap_user = swap.booking.user.get_full_name() or swap.booking.user.email
    
    if len(dependent_swaps) == 1:
        dep = dependent_swaps[0]
        dep_user = dep.booking.user.get_full_name() or dep.booking.user.email
        return (
            f"Cannot cancel this swap yet! "
            f"{dep_user}'s swap (Room {dep.from_room.room_no} Bed {dep.from_share_no} → "
            f"Room {dep.to_room.room_no} Bed {dep.to_share_no}) depends on this swap completing. "
            f"Cancel {dep_user}'s swap first, or use 'Cancel All Chain' to cancel both."
        )
    else:
        msg = f"Cannot cancel this swap yet! {len(dependent_swaps)} dependent swaps found:\n"
        for i, dep in enumerate(dependent_swaps, 1):
            dep_user = dep.booking.user.get_full_name() or dep.booking.user.email
            msg += f"{i}. {dep_user}: Room {dep.from_room.room_no}/{dep.from_share_no} → {dep.to_room.room_no}/{dep.to_share_no}\n"
        msg += "\nCancel in reverse order (last in chain first), or use 'Cancel All Chain' to cancel all."
        return msg


# =============================================
# WhatsApp Group Management Views
# =============================================

@login_required
def whatsapp_management(request):
    """
    WhatsApp Group Management page - shows all approved bookings with WhatsApp invite status.
    Allows PG admin to send WhatsApp invites and mark them as sent.
    """
    if not _require_pg_admin(request.user):
        messages.error(request, "PG Admin access required.")
        return redirect('dashboard')
    
    pg = _active_pg(request)
    if not pg:
        messages.error(request, "No PG found.")
        return redirect('dashboard')
    
    # Get all approved bookings for this PG (both active and completed for filtering)
    bookings = Booking.objects.filter(
        pg=pg,
        status__in=[Booking.APPROVED, Booking.COMPLETED]
    ).select_related(
        'user', 'room'
    ).prefetch_related(
        'application'
    ).order_by('-joining_date')
    
    # Build booking data with WhatsApp info
    booking_data = []
    for b in bookings:
        # Get WhatsApp number from application
        whatsapp_number = ''
        if hasattr(b, 'application') and b.application:
            whatsapp_number = b.application.whatsapp_number or b.application.phone or ''
        
        # Determine if booking is active (approved and not left yet)
        is_active = b.status == Booking.APPROVED and (not b.leaving_date or b.leaving_date >= timezone.now().date())
        
        booking_data.append({
            'id': b.id,
            'user_name': b.user.get_full_name() or b.user.email,
            'user_email': b.user.email,
            'room_no': b.room.room_no if b.room else '-',
            'bed_no': b.share_no,
            'whatsapp_number': whatsapp_number,
            'whatsapp_invite_sent': b.whatsapp_invite_sent,
            'whatsapp_invite_sent_at': b.whatsapp_invite_sent_at.strftime('%Y-%m-%d %H:%M') if b.whatsapp_invite_sent_at else None,
            'joining_date': b.joining_date.strftime('%Y-%m-%d') if b.joining_date else '-',
            'is_active': is_active,
            'status': b.status,
        })
    
    # Calculate stats
    total_bookings = len([b for b in booking_data if b['is_active']])
    sent_count = len([b for b in booking_data if b['is_active'] and b['whatsapp_invite_sent']])
    pending_count = total_bookings - sent_count
    
    context = {
        'pg': pg,
        'pgs': list(_admin_pgs(request.user)),
        'bookings_json': json.dumps(booking_data),
        'whatsapp_invite_link': pg.whatsapp_invite_link or '',
        'whatsapp_invite_message': pg.whatsapp_invite_message or '',
        'total_bookings': total_bookings,
        'sent_count': sent_count,
        'pending_count': pending_count,
    }
    
    return render(request, 'pgadmin/whatsapp_management.html', context)


@login_required
def whatsapp_mark_sent(request, booking_id):
    """
    AJAX endpoint to mark WhatsApp invite as sent for a booking.
    """
    if not _require_pg_admin(request.user):
        return JsonResponse({'ok': False, 'error': 'PG Admin access required.'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST required.'}, status=405)
    
    pg = _active_pg(request)
    if not pg:
        return JsonResponse({'ok': False, 'error': 'No PG found.'}, status=400)
    
    try:
        booking = Booking.objects.get(pk=booking_id, pg=pg)
    except Booking.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Booking not found.'}, status=404)
    
    # Toggle or set the status
    data = json.loads(request.body) if request.body else {}
    mark_as_sent = data.get('mark_sent', True)
    
    booking.whatsapp_invite_sent = mark_as_sent
    if mark_as_sent:
        booking.whatsapp_invite_sent_at = timezone.now()
    else:
        booking.whatsapp_invite_sent_at = None
    booking.save(update_fields=['whatsapp_invite_sent', 'whatsapp_invite_sent_at'])
    
    # Log the action
    log(
        request.user,
        'whatsapp_invite_marked' if mark_as_sent else 'whatsapp_invite_unmarked',
        'Booking',
        booking.id,
        f"WhatsApp invite {'marked as sent' if mark_as_sent else 'unmarked'} for {booking.user.get_full_name() or booking.user.email}"
    )
    
    return JsonResponse({
        'ok': True,
        'whatsapp_invite_sent': booking.whatsapp_invite_sent,
        'whatsapp_invite_sent_at': booking.whatsapp_invite_sent_at.strftime('%Y-%m-%d %H:%M') if booking.whatsapp_invite_sent_at else None,
    })


@login_required
def whatsapp_stats(request):
    """
    AJAX endpoint to get updated WhatsApp stats for the current PG.
    """
    if not _require_pg_admin(request.user):
        return JsonResponse({'ok': False, 'error': 'PG Admin access required.'}, status=403)
    
    pg = _active_pg(request)
    if not pg:
        return JsonResponse({'ok': False, 'error': 'No PG found.'}, status=400)
    
    # Get active approved bookings
    active_bookings = Booking.objects.filter(
        pg=pg,
        status=Booking.APPROVED
    ).filter(
        Q(leaving_date__isnull=True) | Q(leaving_date__gte=timezone.now().date())
    )
    
    total_active = active_bookings.count()
    sent_count = active_bookings.filter(whatsapp_invite_sent=True).count()
    pending_count = total_active - sent_count
    
    return JsonResponse({
        'ok': True,
        'total_active': total_active,
        'sent_count': sent_count,
        'pending_count': pending_count,
    })


# execute_future_swap_manually removed - swaps now only execute automatically on scheduled date via sync_bed_statuses


@login_required
@transaction.atomic
def booking_delete(request, booking_id):
    """
    Delete a confirmed/approved booking. Requires PG Admin permission: can_delete_confirmed_bookings.
    This will:
    1. Free up the bed/share (set to VACANT)
    2. Delete associated ResidentApplication if exists
    3. Delete uploaded documents from Google Drive
    4. Delete the booking record
    """
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST request required.'}, status=405)
    
    if not _require_pg_admin(request.user):
        messages.error(request, "PG Admin access required.")
        return redirect('dashboard')
    
    booking = get_object_or_404(Booking, pk=booking_id)
    
    # Get the PG for this booking
    booking_pg = getattr(booking, 'pg', None) or (booking.room.pg if booking.room else None)
    
    if not booking_pg:
        messages.error(request, "Could not determine PG for this booking.")
        return redirect('dashboard')
    
    # Check if admin has access to this PG
    if not _admin_pgs(request.user).filter(id=booking_pg.id).exists():
        messages.error(request, "You don't have admin access to this PG.")
        return redirect('dashboard')
    
    # Check specific permission for deleting confirmed bookings
    from .models import PGAdminPermission
    try:
        permission = PGAdminPermission.objects.get(pg_admin__user=request.user, pg_admin__pg=booking_pg)
        can_delete = permission.can_delete_confirmed_bookings
    except PGAdminPermission.DoesNotExist:
        can_delete = False
    
    # Site admins can always delete
    is_site_admin = request.user.is_staff or request.user.is_superuser
    
    if not can_delete and not is_site_admin:
        messages.error(request, "You don't have permission to delete confirmed bookings.")
        return redirect('booking_detail', booking_id=booking_id)
    
    # Store details for logging before deletion
    user_name = booking.user.get_full_name() or booking.user.email
    room_info = f"Room {booking.room.room_number}, Share {booking.share_no}" if booking.room else "N/A"
    booking_status = booking.status
    
    # Free up the bed/share if booking was approved and has a room assigned
    if booking.room and booking.share_no:
        try:
            share = RoomShareStatus.objects.get(room=booking.room, share_no=booking.share_no)
            # Only free up if this booking was using the bed
            if booking.status == Booking.APPROVED:
                share.status = RoomShareStatus.VACANT
                share.vacant_from = None
                share.save(update_fields=['status', 'vacant_from'])
                _logger.info(f"Freed up bed: Room {booking.room.room_number}, Share {booking.share_no}")
        except RoomShareStatus.DoesNotExist:
            pass  # Share might not exist
    
    # Delete uploaded documents from Google Drive
    drive_files_deleted = []
    if booking.aadhaar_front_url:
        try:
            drive_delete(booking.aadhaar_front_url)
            drive_files_deleted.append('aadhaar_front')
        except Exception as e:
            _logger.warning(f"Failed to delete aadhaar_front from Drive: {e}")
    
    if booking.aadhaar_back_url:
        try:
            drive_delete(booking.aadhaar_back_url)
            drive_files_deleted.append('aadhaar_back')
        except Exception as e:
            _logger.warning(f"Failed to delete aadhaar_back from Drive: {e}")
    
    if booking.photo_url:
        try:
            drive_delete(booking.photo_url)
            drive_files_deleted.append('photo')
        except Exception as e:
            _logger.warning(f"Failed to delete photo from Drive: {e}")
    
    # Delete associated ResidentApplication if exists
    application_deleted = False
    if hasattr(booking, 'application') and booking.application:
        booking.application.delete()
        application_deleted = True
    
    # Delete the booking
    booking.delete()
    
    # Log the deletion
    log(
        request.user,
        'booking_deleted',
        'Booking',
        booking_id,
        f"Deleted booking for {user_name} ({room_info}). Status was: {booking_status}. "
        f"Application deleted: {application_deleted}. Drive files deleted: {drive_files_deleted}"
    )
    
    messages.success(request, f"Booking for {user_name} has been deleted successfully.")
    
    # Handle AJAX response
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.POST.get('ajax'):
        return JsonResponse({
            'ok': True,
            'action': 'booking_delete',
            'booking_id': booking_id,
            'message': f'Booking for {user_name} deleted successfully.',
            'redirect_url': reverse('pg_bookings_list')
        })
    
    return redirect('pg_bookings_list')
